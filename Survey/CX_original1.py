#####测斜还原原始数据 测试CX1
import numpy
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
import os
import shutil
####1、制定测斜模板
moban_excel_book=openpyxl.Workbook()
manban_excel_sheet=moban_excel_book.create_sheet('测斜原始数据')



####2、根据数据库来生成测斜原始数据