import numpy
import pandas as pd
from openpyxl import load_workbook
import os
import re
import random
# def weiyi1(src_path,src_data_name,Cnumber):
src_path="D:\\2022\\"
src_data_name="k110-k290.xlsx" ###这里的xlsx文件必须再src_path文件下
output_name="报表"  ###输出的文件夹名字
if(os.path.isdir(src_path+output_name)):  ###判断输出文件架是否存在，如果存在不创建
    pass
    print(output_name+"已经存在！")
else:
    print("创建文件夹"+output_name+"!")
    os.makedirs(src_path+output_name)
book1=load_workbook(src_path+src_data_name)
sheet_names=book1.sheetnames
wy_name="坡顶水平位移"
cj_name="坡顶沉降"
wy_name1="桩顶水平位移"
cj_name1="桩顶沉降"
for i in range(len(sheet_names)):
    if(wy_name in sheet_names[i] or wy_name1 in sheet_names[i]):
        wy_sheet=sheet_names[i]
    if(cj_name in sheet_names[i] or cj_name1 in sheet_names[i]):
        cj_sheet=sheet_names[i]
wy_data=pd.read_excel(src_path+src_data_name,wy_sheet)
cj_data=pd.read_excel(src_path+src_data_name,cj_sheet)
Cnumber='1-10' ###表示输入期次范围
Cnumber1=int(re.findall(r'(.*)-',Cnumber,flags=0)[0])
Cnumber2=int(re.findall(r'-(.*)',Cnumber,flags=0)[0])
Cdate_range='' ###表示日期的范围
Cznumber=cj_data.shape[0]-11 ###总的期次 一次性生成完
Cdname=[]
for i in range(3,cj_data.shape[1]):
    Cdname.append(cj_data.iloc[10,i])
for i in range(10+Cnumber1,10+Cnumber2+1):
    ####创建文件，文件的格式为dat文件或者txt文件
    ###判断文件是否存在
    date1_name=str(i-10)+'观测记录'+str(cj_data.iloc[i,1]).replace(" 00:00:00","")+'坐标.txt'  ###数据库中的日期单元格格式需要保持日期的格式 年月日 ///
    fid=open(src_path+output_name+"\\"+date1_name,'w')
    fid.write("坡顶水平位移"+str(cj_data.iloc[i,1]).replace(" 00:00:00","")+':'+"\n")
    WY1=[]
    WY2=[]
    WY3=[]
    for j in range(0,cj_data.shape[1]-3):
        bx=wy_data.iloc[i,3+j*2]
        by=wy_data.iloc[i,3+j*2+1]
        if(numpy.isnan(bx)):
            continue
        wy_x_random1=round(bx+random.randint(-5, 15) * 0.0001,4)##这里对X方向均值及Y方向的均值进行取值计算
        wy_x_random2=round(bx+random.randint(-5, 15) * 0.0001,4)
        wy_x_random3=round(3*bx-wy_x_random1-wy_x_random2,4)
        wy_y_random1=round(by+random.randint(-5,15)*0.0001,4)
        wy_y_random2=round(by+random.randint(-5,15)*0.0001,4)
        wy_y_random3=round(3*by-wy_y_random1-wy_y_random2,4)
        WY1.append([wy_x_random1,wy_y_random1])
        WY2.append([wy_x_random2,wy_y_random2])
        WY3.append([wy_x_random3,wy_y_random3])
    k=0
    for j in range(3,cj_data.shape[1]):
        a=cj_data.iloc[i,j]
        if(numpy.isnan(a)):
            continue
        Cdname1=Cdname[j-3]
        name1=Cdname1+'-1'
        name2=Cdname1+'-2'
        name3=Cdname1+'-3'
        cj_random1=round(a+random.randint(-5,15)*0.0001,4) ###这里只对沉降数据进行了一次随机数的选取
        cj_random2=round(a+random.randint(-5,15)*0.0001,4)
        cj_random3=round(a+random.randint(-5,15)*0.0001,4)
        a=str(WY1[k][0])
        len_num=100
        ling1='0'
        ling2='00'
        ling3='000'
        if(re.findall('[.](.*)',a,flags=0)==[]):   ########保留小数4为 字符串
            ag =a+'0000'
        elif(len(re.findall('[.](.*)',a,flags=0)[0])==1):
            ag=a+ling3
        elif(len(re.findall('[.](.*)',a,flags=0)[0])==2):
            ag=a+ling2
        elif(len(re.findall('[.](.*)',a,flags=0)[0])==3):
            ag=a+ling1
        else:
            ag=a
        b=str(WY1[k][1])
        if(re.findall('[.](.*)',b,flags=0)==[]):
            bg =b+'0000'
        elif(len(re.findall('[.](.*)',b,flags=0)[0])==1):
            bg=b+ling3
        elif(len(re.findall('[.](.*)',b,flags=0)[0])==2):
            bg=b+ling2
        elif(len(re.findall('[.](.*)',b,flags=0)[0])==3):
            bg=b+ling1
        else:
            bg=b
        c=str(cj_random1)
        if(re.findall('[.](.*)',c,flags=0)==[]):
            cg =c+'0000'
        elif(len(re.findall('[.](.*)',c,flags=0)[0])==1):
            cg=c+ling3
        elif(len(re.findall('[.](.*)',c,flags=0)[0])==2):
            cg=c+ling2
        elif(len(re.findall('[.](.*)',c,flags=0)[0])==3):
            cg=c+ling1
        else:
            cg=c
        a1=str(WY2[k][0])
        if(re.findall('[.](.*)',a1,flags=0)==[]):
            ag1 =a1+'0000'
        elif(len(re.findall('[.](.*)',a1,flags=0)[0])==1):
            ag1=a1+ling3
        elif(len(re.findall('[.](.*)',a1,flags=0)[0])==2):
            ag1=a1+ling2
        elif(len(re.findall('[.](.*)',a1,flags=0)[0])==3):
            ag1=a1+ling1
        else:
            ag1=a1
        b1=str(WY2[k][1])
        if(re.findall('[.](.*)',b1,flags=0)==[]):
            bg1 =b1+'0000'
        elif(len(re.findall('[.](.*)',b1,flags=0)[0])==1):
            bg1=b1+ling3
        elif(len(re.findall('[.](.*)',b1,flags=0)[0])==2):
            bg1=b1+ling2
        elif(len(re.findall('[.](.*)',b1,flags=0)[0])==3):
            bg1=b1+ling1
        else:
            bg1=b1
        c1=str(cj_random2)
        if(re.findall('[.](.*)',c1,flags=0)==[]):
            cg1 =c1+'0000'
        elif(len(re.findall('[.](.*)',c1,flags=0)[0])==1):
            cg1=c1+ling3
        elif(len(re.findall('[.](.*)',c1,flags=0)[0])==2):
            cg1=c1+ling2
        elif(len(re.findall('[.](.*)',c1,flags=0)[0])==3):
            cg1=c1+ling1
        else:
            cg1=c1
        a2=str(WY3[k][0])
        if(re.findall('[.](.*)',a2,flags=0)==[]):
            ag2 =a2+'0000'
        elif(len(re.findall('[.](.*)',a2,flags=0)[0])==1):
            ag2=a2+ling3
        elif(len(re.findall('[.](.*)',a2,flags=0)[0])==2):
            ag2=a2+ling2
        elif(len(re.findall('[.](.*)',a2,flags=0)[0])==3):
            ag2=a2+ling1
        else:
            ag2=a2
        b2=str(WY3[k][1])
        if(re.findall('[.](.*)',b2,flags=0)==[]):
            bg2 =b2+'0000'
        elif(len(re.findall('[.](.*)',b2,flags=0)[0])==1):
            bg2=b2+ling3
        elif(len(re.findall('[.](.*)',b2,flags=0)[0])==2):
            bg2=b2+ling2
        elif(len(re.findall('[.](.*)',b2,flags=0)[0])==3):
            bg2=b2+ling1
        else:
            bg2=b2
        c2=str(cj_random3)
        if(re.findall('[.](.*)',c2,flags=0)==[]):
            cg2 =c2+'0000'
        elif(len(re.findall('[.](.*)',c2,flags=0)[0])==1):
            cg2=c2+ling3
        elif(len(re.findall('[.](.*)',c2,flags=0)[0])==2):
            cg2=c2+ling2
        elif(len(re.findall('[.](.*)',c2,flags=0)[0])==3):
            cg2=c2+ling1
        else:
            cg2=c2
        fid.write(name1+'，'+ag+'，'+bg+'，'+cg+"\n")
        fid.write(name2 + '，' + ag1 + '，' + bg1 + '，' + cg1+"\n")
        fid.write(name3 + '，' + ag2 + '，' + bg2 + '，' + cg2+"\n")
        k=k+1
    fid.close()



