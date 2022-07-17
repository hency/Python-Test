import os
import openpyxl
import re
book1=openpyxl.Workbook()
sheet1=book1.create_sheet('sheet')
src_path="D:\\省院测量公司数据处理\\三组\\G03原始数据\\"
src_path="D:\\省院测量公司数据处理\三组\\H12原始数据\\原始数据\\"
filename=os.listdir(src_path)
#排序：type1没有商业街的
for i in range(len(filename)-1):
    for j in range(i+1,len(filename)):
        a=int(re.findall(r'\d+',filename[i])[0])
        b=int(re.findall(r'\d+',filename[j])[0])
        if(a>b):
            mid1=filename[i]
            filename[i]=filename[j]
            filename[j]=mid1
#排序：type2有商业街的 暂时不写
k=1
for i in range(len(filename)):
    filename1=src_path+filename[i]
    filename2=os.listdir(filename1)
    k1=1
    for j in range(len(filename2)):
        for z in range(len(filename2)):
            if('第'+str(j+1)+'期' in filename2[z]):
                filename3=os.listdir(filename1+'\\'+filename2[z])
                for z1 in range(len(filename3)):
                    if('dat' in filename3[z1]):
                        filename4=filename3[z1]
                f = open(filename1+'\\'+filename2[z]+'\\'+filename4, 'r')
                file1 = f.readlines()
                count1 = len(file1)
                r1 = file1[count1 - 3 - 1][58:66]
                # print(file1[count1-3-1])
                f.close()
                sheet1.cell(k1,k).value=float(r1)*1000
                k1=k1+1
    k=k+1
book1.save('闭合差1.xlsx')

