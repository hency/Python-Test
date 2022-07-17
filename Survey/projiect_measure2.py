#coding=utf-8
import sys
from PyQt5.QtWidgets import QWidget,QPushButton,QVBoxLayout,QHBoxLayout,QGridLayout,QFormLayout,QLineEdit,QLabel,QMessageBox,QApplication,\
    QDialog,QAction,QDateTimeEdit,QFileDialog
from PyQt5.QtCore import QDate, QTime, QDateTime
import numpy
import pandas as pd
import numpy as np
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import os
import shutil
from openpyxl import load_workbook
import math
import datetime
class Data_Process(QWidget):
    def __init__(self):
        super(Data_Process,self).__init__()
        self.resize(800,400)
        self.setWindowTitle('测量公司内业数据处理软件')
        self.label1=QLabel(self)
        self.label1.setText('主体沉降数据处理')
        self.pushbutton1=QPushButton(self)
        self.pushbutton1.setText('进入程序')
        self.label2=QLabel(self)
        self.label2.setText('基坑监测数据处理')
        self.pushbutton2=QPushButton(self)
        self.pushbutton2.setText('进入程序')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout_init()
    def Layout_init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout2.addWidget(self.pushbutton1)
        self.V_layout2.addWidget(self.pushbutton2)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.setLayout(self.H_layout1)
        self.pushbutton1.clicked.connect(self.enter_pushbutton1)
        self.pushbutton2.clicked.connect(self.enter_pushbutton2)
    def enter_pushbutton1(self):
        demo = Zhuti_Data_Process()
        demo.exec_()

    def enter_pushbutton2(self):
        demo1 = Jikeng_Data_Process()
        demo1.exec_()
class Zhuti_Data_Process(QDialog):
    def __init__(self):
        super(Zhuti_Data_Process,self).__init__()
        self.resize(1000,500)
        self.setWindowTitle('主体沉降数据处理')
        self.label1=QLabel(self)
        self.label1.setText('(1)主体沉降原始数据还原：提供3种文件格式 支点EXCEl格式、无支点EXCEL格式、支点DAT格式')
        self.pushbutton1=QPushButton(self)
        self.pushbutton1.setText('原始数据')
        self.label2=QLabel(self)
        self.label2.setText('(2)创建文件夹：需要根据已有的模板进行创建')
        self.pushbutton2=QPushButton(self)
        self.pushbutton2.setText('创建文件夹')
        self.label3=QLabel(self)
        self.label3.setText('(3)闭合差提取并生成：高差偶然中误差及高差全中误差')
        self.pushbutton3=QPushButton(self)
        self.pushbutton3.setText('闭合差提取')
        self.label4=QLabel(self)
        self.label4.setText('(4)生成基准点复测表')
        self.pushbutton4=QPushButton(self)
        self.pushbutton4.setText('基准点复测表')
        self.label5=QLabel(self)
        self.label5.setText('(5)生成主体沉降日报：根据浇筑时间模板、及初始值、拟合模型曲线系数')
        self.pushbutton5=QPushButton(self)
        self.pushbutton5.setText('主体日报')
        self.label6=QLabel(self)
        self.label6.setText('(6)主体沉降数据库创建：根据最终日报确定')
        self.pushbutton6=QPushButton(self)
        self.pushbutton6.setText('生成数据库')
        self.label7=QLabel(self)
        self.label7.setText('(7)主体时间-荷载-沉降量曲线:根据最终日报确定')
        self.pushbutton7=QPushButton(self)
        self.pushbutton7.setText('累计变化曲线图')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout_init()
    def Layout_init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout1.addWidget(self.label4)
        self.V_layout1.addWidget(self.label5)
        self.V_layout1.addWidget(self.label6)
        self.V_layout1.addWidget(self.label7)
        self.V_layout2.addWidget(self.pushbutton1)
        self.V_layout2.addWidget(self.pushbutton2)
        self.V_layout2.addWidget(self.pushbutton3)
        self.V_layout2.addWidget(self.pushbutton4)
        self.V_layout2.addWidget(self.pushbutton5)
        self.V_layout2.addWidget(self.pushbutton6)
        self.V_layout2.addWidget(self.pushbutton7)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.setLayout(self.H_layout1)
        self.pushbutton1.clicked.connect(self.enter_pushbutton1)
        self.pushbutton2.clicked.connect(self.enter_pushbutton2)
        self.pushbutton3.clicked.connect(self.enter_pushbutton3)
        self.pushbutton4.clicked.connect(self.enter_pushbutton4)
        self.pushbutton5.clicked.connect(self.enter_pushbutton5)
        self.pushbutton6.clicked.connect(self.enter_pushbutton6)
        self.pushbutton7.clicked.connect(self.enter_pushbutton7)
    def enter_pushbutton1(self):
        demo1=Zhuti_original_cj()
        demo1.exec_()

    def enter_pushbutton2(self):
        demo1 = Create_zhuti_dir()
        demo1.exec_()

    def enter_pushbutton3(self):
        demo1 = Zhuti_bihecha()
        demo1.exec_()

    def enter_pushbutton4(self):
        pass

    def enter_pushbutton5(self):
        demo1 = Zhuti_ribao()
        demo1.exec_()

    def enter_pushbutton6(self):
        demo1 = Zhuti_dataset()
        demo1.exec_()

    def enter_pushbutton7(self):
        demo1 = Zhuti_curv()
        demo1.exec_()


class Create_zhuti_dir(QDialog):
    def __init__(self):
        super(Create_zhuti_dir,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('创建主体沉降原始数据文件夹')
        self.label1=QLabel('第一步打开数据库路径：数据库命名需要按照规则命名',self)
        self.label2=QLabel('第二步打开原始数据存放路径：数据库命名需要按照规则命名', self)
        self.button1=QPushButton(self)
        self.button1.setText('执行创建文件夹程序')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.zhuti_dir)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zhuti_dir(self):
        pass

class Zhuti_bihecha(QDialog):
    def __init__(self):
        super(Zhuti_bihecha,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降闭合差提取：输出到EXCEL中')
        self.label1 = QLabel('打开原始数据存放路径：数据库命名需要按照规则命名', self)
        self.button1=QPushButton(self)
        self.button1.setText('执行主体沉降闭合差提取程序')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.zhuti_bihecha)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zhuti_bihecha(self):
        path1=os.path.abspath(QFileDialog.getExistingDirectory(self,"选取文件夹",''))+'\\' # 起始路径
        filename=os.listdir(path1)
        book1=openpyxl.Workbook()
        sheet1=book1.create_sheet('闭合差')
        for i in range(len(filename)-1):
            for i in range(len(filename) - 1):
                for j in range(i + 1, len(filename)):
                    a = int(re.findall(r'\d+', filename[i])[0])
                    b = int(re.findall(r'\d+', filename[j])[0])
                    if (a > b):
                        mid1 = filename[i]
                        filename[i] = filename[j]
                        filename[j] = mid1
        k = 1
        for i in range(len(filename)):
            filename1 = path1 + filename[i]
            filename2 = os.listdir(filename1)
            k1 = 1
            for j in range(len(filename2)):
                for z in range(len(filename2)):
                    if ('第' + str(j + 1) + '期' in filename2[z]):
                        filename3 = os.listdir(filename1 + '\\' + filename2[z])
                        for z1 in range(len(filename3)):
                            if ('dat' in filename3[z1]):
                                filename4 = filename3[z1]
                        f = open(filename1 + '\\' + filename2[z] + '\\' + filename4, 'r')
                        file1 = f.readlines()
                        count1 = len(file1)
                        r1 = file1[count1 - 3 - 1][58:66]
                        # print(file1[count1-3-1])
                        f.close()
                        sheet1.cell(k1, k).value = float(r1) * 1000
                        k1 = k1 + 1
            k = k + 1
        book1.save('闭合差1.xlsx')
class Zhuti_ribao(QDialog):
    def __init__(self):
        super(Zhuti_ribao,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降日报创建')
        self.label1 = QLabel('打开初始值数据存放路径：数据库命名需要按照规则命名', self)
        self.label2 = QLabel('打开浇筑时间数据存放路径：数据库命名需要按照规则命名', self)
        self.button1=QPushButton(self)
        self.button1.setText('执行主体沉降闭合差提取程序')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.zhuti_ribao)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zhuti_ribao(self):
        pass
class Zhuti_dataset(QDialog):
    def __init__(self):
        super(Zhuti_dataset,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降数据库创建')
        self.label1 = QLabel('打开日报存放路径：数据库命名将按照规则命名', self)
        self.button1=QPushButton(self)
        self.button1.setText('执行主体沉降闭合差提取程序')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.zhuti_dataset)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zhuti_dataset(self):
        pass

class Zhuti_curv(QDialog):
    def __init__(self):
        super(Zhuti_curv,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体时间-荷载-沉降量曲线')
        self.label1 = QLabel('打开日报存放路径：生成一个累计变化量的EXCEl，然后再根据EXCEl创建曲线图', self)
        self.button1=QPushButton(self)
        self.button1.setText('执行主体沉降闭合差提取程序')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.zhuti_curv)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zhuti_curv(self):
        pass

class Jikeng_Data_Process(QDialog):
    def __init__(self):
        super(Jikeng_Data_Process,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('基坑监测数据处理')
        self.label1=QLabel(self)
        self.label1.setText('坡顶位移原始数据还原')
        self.pushbutton1=QPushButton(self)
        self.pushbutton1.setText('坡顶位移')
        self.label2=QLabel(self)
        self.label2.setText('坡顶沉降、道路、管线、地表、建筑沉降原始数据还原')
        self.pushbutton2=QPushButton(self)
        self.pushbutton2.setText('沉降数据')
        self.label3=QLabel(self)
        self.label3.setText('深层水平位移还原')
        self.pushbutton3=QPushButton(self)
        self.pushbutton3.setText('测斜')
        self.label4=QLabel(self)
        self.label4.setText('水位原始数据还原')
        self.pushbutton4=QPushButton(self)
        self.pushbutton4.setText('水位')
        self.label5=QLabel(self)
        self.label5.setText('支撑轴力原始数据还原')
        self.pushbutton5=QPushButton(self)
        self.pushbutton5.setText('支撑内力')
        self.label6=QLabel(self)
        self.label6.setText('原始数据文件夹')
        self.pushbutton6=QPushButton(self)
        self.pushbutton6.setText('创建文件夹')
        self.label7=QLabel(self)
        self.label7.setText('更改文件的名字')
        self.pushbutton7=QPushButton(self)
        self.pushbutton7.setText('执行RENAME')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout_init()
    def Layout_init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout1.addWidget(self.label4)
        self.V_layout1.addWidget(self.label5)
        self.V_layout1.addWidget(self.label6)
        self.V_layout1.addWidget(self.label7)
        self.V_layout2.addWidget(self.pushbutton1)
        self.V_layout2.addWidget(self.pushbutton2)
        self.V_layout2.addWidget(self.pushbutton3)
        self.V_layout2.addWidget(self.pushbutton4)
        self.V_layout2.addWidget(self.pushbutton5)
        self.V_layout2.addWidget(self.pushbutton6)
        self.V_layout2.addWidget(self.pushbutton7)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.setLayout(self.H_layout1)
        self.pushbutton1.clicked.connect(self.enter_pushbutton1)
        self.pushbutton2.clicked.connect(self.enter_pushbutton2)
        self.pushbutton3.clicked.connect(self.enter_pushbutton3)
        self.pushbutton4.clicked.connect(self.enter_pushbutton4)
        self.pushbutton5.clicked.connect(self.enter_pushbutton5)
        self.pushbutton6.clicked.connect(self.enter_pushbutton6)
        self.pushbutton7.clicked.connect(self.enter_pushbutton7)
    def enter_pushbutton1(self):
        demo1=Jikeng_original_wy()
        demo1.exec_()

    def enter_pushbutton2(self):
        demo1 = Jikeng_original_cj()
        demo1.exec_()

    def enter_pushbutton3(self):
        demo1=Jikeng_original_cx()
        demo1.exec_()

    def enter_pushbutton4(self):
        demo1 =Jikeng_original_sw()
        demo1.exec_()

    def enter_pushbutton5(self):
        demo1=Jikeng_original_zl()
        demo1.exec_()

    def enter_pushbutton6(self):
        demo1 = Jikeng_make_dir()
        demo1.exec_()

    def enter_pushbutton7(self):
        demo1 = Jikeng_rename()
        demo1.exec_()

class Jikeng_original_cx(QDialog):
    def __init__(self):
        super(Jikeng_original_cx,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('深层水平位移还原')
        self.label1=QLabel('输入期次范围',self)
        self.editline1=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout2.addWidget(self.editline1)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('请输入期数')
        self.button1.clicked.connect(self.cx)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def cx(self):
        pass

class Jikeng_original_sw(QDialog):
    def __init__(self):
        super(Jikeng_original_sw,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('水位原始数据还原')
        self.label1=QLabel('输入期次范围',self)
        self.editline1=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout2.addWidget(self.editline1)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('请输入期数')
        self.button1.clicked.connect(self.sw)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def sw(self):
        pass
    
class Jikeng_original_zl(QDialog):
    def __init__(self):
        super(Jikeng_original_zl,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('支撑内力原始数据还原')
        self.label1=QLabel('输入期次范围',self)
        self.editline1=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout2.addWidget(self.editline1)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('请输入期数')
        self.button1.clicked.connect(self.zl)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def zl(self):
        pass

class Jikeng_original_cj(QDialog):
    def __init__(self):
        super(Jikeng_original_cj, self).__init__()
        self.resize(600, 300)
        self.setWindowTitle('基坑沉降原始数据还原')
        self.label1 = QLabel('闭合差范围', self)
        self.editline1 = QLineEdit(self)
        self.label2 = QLabel('总期数', self)
        self.editline2 = QLineEdit(self)
        self.button1 = QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.V_layout3 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout__init()
        self.display_editline()

    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout2.addWidget(self.editline1)
        self.V_layout2.addWidget(self.editline2)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)

    def display_editline(self):
        self.editline1.setPlaceholderText('X Y')
        self.editline2.setPlaceholderText('请输入期数:')
        self.button1.clicked.connect(self.jikeng_function_cj)
    def jikeng_function_cj(self):
        # measure_line_path = "D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"#D:\Desktop\测试期次1\测试线路文件1.xlsx##self.editline1.text()
        # dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'#D:\Desktop\测试期次1\测试数据库1.xlsx'#self.editline2.text()#self.editline2.text()
        # path_output1 = 'D:\\Desktop\\测试期次1\\测试output\\'#self.editline3.text()#'D:\Desktop\测试期次1\测试output\self.editline3.text()
        measure_line_path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
        dateset_path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
        src_path = os.path.abspath(
            os.path.join(measure_line_path, "..")) + '\\'  #############################################################
        output_name = "报表"  ###输出的文件夹名字
        if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
            pass
            print(output_name + "已经存在！")
        else:
            print("创建文件夹" + output_name + "!")
            os.makedirs(src_path + output_name)
        path_output1 = src_path + output_name
        qi_shu = self.editline2.text()
        BC_range1 = self.editline1.text()
        BC_range_min = re.findall(r'(.*) ', BC_range1, flags=0)[0]
        BC_range_max = re.findall(r' (.*)', BC_range1, flags=0)[0]

class Jikeng_original_wy(QDialog):
    def __init__(self):
        super(Jikeng_original_wy,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('坡顶位移数据还原')
        self.label1=QLabel('输入期次范围',self)
        self.editline1=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout2.addWidget(self.editline1)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('请输入期数')
        self.button1.clicked.connect(self.output_excel)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def output_excel(self):
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        src_path=os.path.abspath(os.path.join(openfile_name[0], ".."))+'\\' #############################################################
        dateset_file_path=os.path.abspath(openfile_name[0])
        # src_path = self.editline1.text()  # src_path="D:\\Desktop\\水平位移原始数据还原\\"
        # src_data_name = self.editline2.text()  # src_data_name="4-2数据库.xlsx" ###这里的xlsx文件必须再src_path文件下
        Cnumber = self.editline1.text()   # Cnumber='2-50' ###表示输入期次范围
        # src_path="D:\\Desktop\\水平位移原始数据还原\\"
        # src_data_name="4-2数据库.xlsx" ###这里的xlsx文件必须再src_path文件下
        output_name = "报表"  ###输出的文件夹名字
        if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
            pass
            print(output_name + "已经存在！")
        else:
            print("创建文件夹" + output_name + "!")
            os.makedirs(src_path + output_name)
        book1 = load_workbook(dateset_file_path)
        sheet_names = book1.sheetnames
        wy_name = "坡顶水平位移"
        cj_name = "坡顶沉降"
        wy_name1 = "桩顶水平位移"
        cj_name1 = "桩顶沉降"
        for i in range(len(sheet_names)):
            if (wy_name in sheet_names[i] or wy_name1 in sheet_names[i]):
                wy_sheet = sheet_names[i]
            if (cj_name in sheet_names[i] or cj_name1 in sheet_names[i]):
                cj_sheet = sheet_names[i]
        wy_data = pd.read_excel(dateset_file_path, wy_sheet)
        cj_data = pd.read_excel(dateset_file_path, cj_sheet)
        # Cnumber='2-50' ###表示输入期次范围
        Cnumber1 = int(re.findall(r'(.*)-', Cnumber, flags=0)[0])
        Cnumber2 = int(re.findall(r'-(.*)', Cnumber, flags=0)[0])
        Cdate_range = ''  ###表示日期的范围
        Cznumber = cj_data.shape[0] - 11  ###总的期次 一次性生成完
        Cdname = []
        for i in range(3, cj_data.shape[1]):
            Cdname.append(cj_data.iloc[10, i])
        for i in range(10 + Cnumber1, 10 + Cnumber2 + 1):
            ####创建文件，文件的格式为dat文件或者txt文件
            ###判断文件是否存在
            date1_name = str(i - 10) + '观测记录' + str(cj_data.iloc[i, 1]).replace(" 00:00:00",
                                                                                "") + '坐标.txt'  ###数据库中的日期单元格格式需要保持日期的格式 年月日 ///
            fid = open(src_path + output_name + "\\" + date1_name, 'w')
            fid.write("坡顶水平位移" + str(cj_data.iloc[i, 1]).replace(" 00:00:00", "") + ':' + "\n")
            WY1 = []
            WY2 = []
            WY3 = []
            for j in range(0, cj_data.shape[1] - 3):
                bx = wy_data.iloc[i, 3 + j * 2]
                by = wy_data.iloc[i, 3 + j * 2 + 1]
                if (numpy.isnan(bx)):
                    continue
                wy_x_random1 = round(bx + random.randint(-5, 15) * 0.0001, 4)  ##这里对X方向均值及Y方向的均值进行取值计算
                wy_x_random2 = round(bx + random.randint(-5, 15) * 0.0001, 4)
                wy_x_random3 = round(3 * bx - wy_x_random1 - wy_x_random2, 4)
                wy_y_random1 = round(by + random.randint(-5, 15) * 0.0001, 4)
                wy_y_random2 = round(by + random.randint(-5, 15) * 0.0001, 4)
                wy_y_random3 = round(3 * by - wy_y_random1 - wy_y_random2, 4)
                WY1.append([wy_x_random1, wy_y_random1])
                WY2.append([wy_x_random2, wy_y_random2])
                WY3.append([wy_x_random3, wy_y_random3])
            k = 0
            for j in range(3, cj_data.shape[1]):
                a = cj_data.iloc[i, j]
                if (numpy.isnan(a)):
                    continue
                Cdname1 = Cdname[j - 3]
                name1 = Cdname1 + '-1'
                name2 = Cdname1 + '-2'
                name3 = Cdname1 + '-3'
                cj_random1 = round(a + random.randint(-5, 15) * 0.0001, 4)  ###这里只对沉降数据进行了一次随机数的选取
                cj_random2 = round(a + random.randint(-5, 15) * 0.0001, 4)
                cj_random3 = round(a + random.randint(-5, 15) * 0.0001, 4)
                a = str(WY1[k][0])
                len_num = 100
                ling1 = '0'
                ling2 = '00'
                ling3 = '000'
                if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数4为 字符串
                    ag = a + '0000'
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
                    ag = a + ling3
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
                    ag = a + ling2
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
                    ag = a + ling1
                else:
                    ag = a
                b = str(WY1[k][1])
                if (re.findall('[.](.*)', b, flags=0) == []):
                    bg = b + '0000'
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 1):
                    bg = b + ling3
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 2):
                    bg = b + ling2
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 3):
                    bg = b + ling1
                else:
                    bg = b
                c = str(cj_random1)
                if (re.findall('[.](.*)', c, flags=0) == []):
                    cg = c + '0000'
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 1):
                    cg = c + ling3
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 2):
                    cg = c + ling2
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 3):
                    cg = c + ling1
                else:
                    cg = c
                a1 = str(WY2[k][0])
                if (re.findall('[.](.*)', a1, flags=0) == []):
                    ag1 = a1 + '0000'
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 1):
                    ag1 = a1 + ling3
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 2):
                    ag1 = a1 + ling2
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 3):
                    ag1 = a1 + ling1
                else:
                    ag1 = a1
                b1 = str(WY2[k][1])
                if (re.findall('[.](.*)', b1, flags=0) == []):
                    bg1 = b1 + '0000'
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 1):
                    bg1 = b1 + ling3
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 2):
                    bg1 = b1 + ling2
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 3):
                    bg1 = b1 + ling1
                else:
                    bg1 = b1
                c1 = str(cj_random2)
                if (re.findall('[.](.*)', c1, flags=0) == []):
                    cg1 = c1 + '0000'
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 1):
                    cg1 = c1 + ling3
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 2):
                    cg1 = c1 + ling2
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 3):
                    cg1 = c1 + ling1
                else:
                    cg1 = c1
                a2 = str(WY3[k][0])
                if (re.findall('[.](.*)', a2, flags=0) == []):
                    ag2 = a2 + '0000'
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 1):
                    ag2 = a2 + ling3
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 2):
                    ag2 = a2 + ling2
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 3):
                    ag2 = a2 + ling1
                else:
                    ag2 = a2
                b2 = str(WY3[k][1])
                if (re.findall('[.](.*)', b2, flags=0) == []):
                    bg2 = b2 + '0000'
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 1):
                    bg2 = b2 + ling3
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 2):
                    bg2 = b2 + ling2
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 3):
                    bg2 = b2 + ling1
                else:
                    bg2 = b2
                c2 = str(cj_random3)
                if (re.findall('[.](.*)', c2, flags=0) == []):
                    cg2 = c2 + '0000'
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 1):
                    cg2 = c2 + ling3
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 2):
                    cg2 = c2 + ling2
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 3):
                    cg2 = c2 + ling1
                else:
                    cg2 = c2
                fid.write(name1 + '，' + ag + '，' + bg + '，' + cg + "\n")
                fid.write(name2 + '，' + ag1 + '，' + bg1 + '，' + cg1 + "\n")
                fid.write(name3 + '，' + ag2 + '，' + bg2 + '，' + cg2 + "\n")
                k = k + 1
            fid.close()

class Zhuti_original_cj(QDialog):
    def __init__(self):
        super(Zhuti_original_cj, self).__init__()
        self.resize(600, 300)
        self.setWindowTitle('沉降原始数据还原')
        self.label1 = QLabel('闭合差范围', self)
        self.editline1 = QLineEdit(self)
        self.label2 = QLabel('总期数', self)
        self.editline2 = QLineEdit(self)
        self.button1 = QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.V_layout3 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout__init()
        self.display_editline()

    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout2.addWidget(self.editline1)
        self.V_layout2.addWidget(self.editline2)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)

    def display_editline(self):
        self.editline1.setPlaceholderText('X Y')
        self.editline2.setPlaceholderText('请输入期数:')
        self.button1.clicked.connect(self.function_cj)

    def read_measure_line_from_dataset(self,dateset_path, measure_line_path):
        mpath = load_workbook(measure_line_path)
        msheet = mpath.get_sheet_by_name('测线')
        msheet = mpath['测线']
        dpath = load_workbook(dateset_path)
        dsheet1 = dpath.get_sheet_by_name('日报')
        sheetname_from_dateset = []
        for name in dpath.sheetnames:
            if ('沉降' in name):
                sheetname_from_dateset.append(name)
                print('沉降数据库有：%s', name)
            else:
                pass
        print('选择指定的数据库进行对应的沉降观测')
        # 由于这里只有一个测线和一个建筑物沉降成果数据库，只对其进行原始数据的还原
        dsheet2 = dpath.get_sheet_by_name(sheetname_from_dateset[0])
        dsheet3 = dpath.get_sheet_by_name('日报')
        # 先进行每一栋每一期的数据进行测试
        max_row = 0
        for i in range(0, dsheet3.max_row):
            if (dsheet3.cell(i + 1, 2).value == '' or dsheet3.cell(i + 1, 3).value is None):
                max_row = dsheet3.max_row
                break
            elif (dsheet3.cell(dsheet3.max_row, 2).value != '' and dsheet3.cell(dsheet3.max_row, 3).value is not None):
                max_row = dsheet3.max_row
        date = []
        max_col = 0
        for i in range(2, max_row + 1):
            date.append(dsheet3.cell(i, 2).value)
            print('第%d期沉降观测' % (i - 1))
        for i in range(1, dsheet2.max_row):
            if (dsheet2.cell(i, 2).value == '日期'):
                date_start = i
        for j in range(1, dsheet2.max_column):
            if (dsheet2.cell(date_start, j) == '' or dsheet2.cell(date_start, j) is None):
                max_col = j - 1
                break
            elif (dsheet2.cell(date_start, dsheet2.max_column) != '' or dsheet2.cell(date_start,
                                                                                     dsheet2.max_column) is not None):
                max_col = dsheet2.max_column
        gc_name = []
        gc_start_num = 0
        for i in range(1, max_col):
            if (dsheet2.cell(date_start, i).value == '时间'):
                gc_start_num = i + 1
                for j in range(0, max_col - gc_start_num + 1):
                    gc_name.append(dsheet2.cell(date_start, gc_start_num + j).value)
                break
        cx_make = []
        for i in range(1, len(date) + 1):
            date1 = date_start + i
            print(
                '***********************************%s*******************************' % (dsheet2.cell(date1, 2).value))
            print('提取侧线文件及数据库文件，进行相应复制')
            cx_make1 = []
            for j in range(0, max_col - gc_start_num + 1):
                cx_make1.append(dsheet2.cell(date1, j + 4).value)
            cx_make.append(cx_make1)
        return gc_name, cx_make

    def position_index(self,df3):
        position = []
        num_1 = 0
        for i in range(df3.shape[0]):
            if 'Y' in df3.iloc[i, 0]:
                position.append([i, i - num_1 - 1])
                num_1 = num_1 + 1
        return position

    def position_index1(self,df3):
        num = []
        position1 = self.position_index(df3)
        flag = 0
        k1 = 0
        k3 = 0
        while (flag == 0):
            for q1 in range(k1, len(position1)):
                k = 0
                k2 = position1[q1][1]
                for q2 in range(len(position1)):
                    if (position1[q1][1] == position1[q2][1]):
                        k = k + 1
                        k1 = k1 + 1
                    else:
                        pass
                k3 = k3 + k
                num.append([k2, k, k3])
                break
            if (k1 == len(position1)):
                flag = 1
        return num

    def Z_H_function(self,df3, df4, num, BC1):
        dict1 = {i: num[i] for i in range(df4.shape[0])}
        position = self.position_index(df3)
        Z_name = []
        Z_H = []
        Z_HD = []
        Z_HF = []
        for zz1 in range(len(position)):
            pass
            Z_DH1 = df3.loc[dict1[position[zz1][1]], 1] - df3.loc[position[zz1][0], 1]
            Z_H.append(df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - Z_DH1)
        return Z_H

    def sight_height_distance(self,df3, df4, book1, sheet1, book2, sheet2, BC1, path1, path2, num, dict1):
        sheet1.cell(1, 1).value = '点位'
        sheet1.cell(1, 3).value = '时间'
        sheet1.cell(1, 5).value = '视线高'
        sheet1.cell(1, 7).value = '视距'
        sheet1.cell(1, 9).value = '高程'
        sheet1.cell(2, 1).value = df3.iloc[0, 0]
        sheet1.cell(2, 9).value = df4.iloc[0, 1]
        sheet2.cell(1, 1).value = '点位'
        sheet2.cell(1, 3).value = '时间'
        sheet2.cell(1, 5).value = '视线高'
        sheet2.cell(1, 7).value = '视距'
        sheet2.cell(1, 9).value = '高程'
        sheet2.cell(2, 1).value = df3.iloc[0, 0]
        sheet2.cell(2, 9).value = df4.iloc[0, 1]
        k = 3
        RH1_random = []
        RH2_random = []
        FH1 = []
        FH2 = []
        height1 = []
        position1 = self.position_index(df3)
        position2 = self.position_index1(df3)
        # BC1=BC2/(df4.shape[0]-1)
        Z_H1 = self.Z_H_function(df3, df4, num, BC1)
        for i in range(df4.shape[0] - 1):
            dh = df4.iloc[i + 1, 1] - df4.iloc[i, 1] - BC1
            if (dh > 1.13):
                print("高差大于1.13m，高差过大")  # 因为dh1_random的值不能超过1.15m
                xx1 = 1 / 0
            height1.append(df4.iloc[i + 1, 1] - (i + 1) * BC1)
            sheet1.cell(7 + 5 * i, 9).value = height1[i]
            # '高差值最大是1.8-0.55=1.25m' 且 dh_random的取值范围为0.55+dh~1.8
            Sight_Height_random1 = random.randint(-5, 15) * 0.00001  ##单位m
            Sight_Height_random2 = random.randint(-5, 15) * 0.00001  ##单位m0.15mm*2=0.3m
            dh1_random = dh + Sight_Height_random1  ##FH2-FH1=2*Sight_Height_random1+Sight_Height_random2
            dh2 = 2 * dh - dh1_random
            flagxx1 = 0
            for q in range(len(position2)):
                if (i == position2[q][0]):
                    #############加入
                    pass
                    flagxx1 = 1
                    if (dh1_random > 0):
                        RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        can_shu = 0
                        for p in range(position2[q][1]):
                            if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                    dh1_random - dh2) < 0.0006) and (
                                    RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                position2[q][2] - position2[q][1] + p] and Z_H1[
                                        position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                        dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                can_shu = can_shu + 1
                        if (can_shu == position2[q][1]):
                            RH1_random.append(
                                RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            print("输出")
                        else:
                            for z1 in range(100000):
                                can_shu1 = 0
                                Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                                RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1_1 = RH1_random_1 - dh1_random
                                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                                FH2_1 = RH2_random_1 - dh2
                                print("重新选择")
                                for p in range(position2[q][1]):
                                    if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                            FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                            RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                        position2[q][2] - position2[q][1] + p] and Z_H1[
                                                position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                                dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                        can_shu1 = can_shu1 + 1
                                if (can_shu1 == position2[q][1]):
                                    RH1_random.append(
                                        RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                    FH1.append(FH1_1)
                                    RH2_random.append(RH2_random_1)
                                    FH2.append(FH2_1)
                                    print("输出")
                                    break
                                else:
                                    continue
                        if ((i + 1) % 2 == 1):
                            sheet1.cell(k, 5, RH1_random[i])
                            sheet1.cell(k + 1, 5, FH1[i])
                            sheet1.cell(k + 2, 5, FH2[i])
                            sheet1.cell(k + 3, 5, RH2_random[i])
                            sheet1.cell(k, 6).value = 'RB'
                            sheet1.cell(k + 1, 6).value = 'RF'
                            sheet1.cell(k + 2, 6).value = 'RF'
                            sheet1.cell(k + 3, 6).value = 'RB'
                        else:
                            sheet1.cell(k, 5, FH1[i])
                            sheet1.cell(k + 1, 5, RH1_random[i])
                            sheet1.cell(k + 2, 5, RH2_random[i])
                            sheet1.cell(k + 3, 5, FH2[i])
                            sheet1.cell(k, 6).value = 'RF'
                            sheet1.cell(k + 1, 6).value = 'RB'
                            sheet1.cell(k + 2, 6).value = 'RB'
                            sheet1.cell(k + 3, 6).value = 'RF'
                        k = k + 5
                    elif (dh1_random < 0):
                        #############加入
                        RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        can_shu = 0
                        for p in range(position2[q][1]):
                            if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                    dh1_random - dh2) < 0.0006) and (
                                    RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                position2[q][2] - position2[q][1] + p] and Z_H1[
                                        position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                        dict1[i], 1] - i * BC1 - 0.6)):
                                # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                can_shu = can_shu + 1
                        if (can_shu == position2[q][1]):
                            RH1_random.append(
                                RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            print("输出")
                        else:
                            for z1 in range(100000):
                                can_shu1 = 0
                                Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                                RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1_1 = RH1_random_1 - dh1_random
                                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                                FH2_1 = RH2_random_1 - dh2
                                print("重新选择")
                                for p in range(position2[q][1]):
                                    if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                            FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                            RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                        position2[q][2] - position2[q][1] + p] and Z_H1[
                                                position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                                dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                        can_shu1 = can_shu1 + 1
                                if (can_shu1 == position2[q][1]):
                                    RH1_random.append(
                                        RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                    FH1.append(FH1_1)
                                    RH2_random.append(RH2_random_1)
                                    FH2.append(FH2_1)
                                    print("输出")
                                    break
                                else:
                                    continue
                        if ((i + 1) % 2 == 1):
                            sheet1.cell(k, 5, RH1_random[i])
                            sheet1.cell(k + 1, 5, FH1[i])
                            sheet1.cell(k + 2, 5, FH2[i])
                            sheet1.cell(k + 3, 5, RH2_random[i])
                            sheet1.cell(k, 6).value = 'RB'
                            sheet1.cell(k + 1, 6).value = 'RF'
                            sheet1.cell(k + 2, 6).value = 'RF'
                            sheet1.cell(k + 3, 6).value = 'RB'
                        else:
                            sheet1.cell(k, 5, RH1_random[i])
                            sheet1.cell(k + 1, 5, FH1[i])
                            sheet1.cell(k + 2, 5, FH2[i])
                            sheet1.cell(k + 3, 5, RH2_random[i])
                            sheet1.cell(k, 6).value = 'RF'
                            sheet1.cell(k + 1, 6).value = 'RB'
                            sheet1.cell(k + 2, 6).value = 'RB'
                            sheet1.cell(k + 3, 6).value = 'RF'
                        k = k + 5
                    elif (dh1_random == 0):
                        RH1_random_1 = round(random.uniform(0.6, 1.75),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        can_shu = 0
                        for p in range(position2[q][1]):
                            if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                    dh1_random - dh2) < 0.0006) and (
                                    RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                position2[q][2] - position2[q][1] + p] and Z_H1[
                                        position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                        dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                can_shu = can_shu + 1
                        if (can_shu == position2[q][1]):
                            RH1_random.append(
                                RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            print("输出")
                        else:
                            for z1 in range(100000):
                                can_shu1 = 0
                                Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                                RH1_random_1 = round(random.uniform(0.6, 1.75),
                                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1_1 = RH1_random_1 - dh1_random
                                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                                FH2_1 = RH2_random_1 - dh2
                                print("重新选择")
                                for p in range(position2[q][1]):
                                    if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                            FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                            RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                        position2[q][2] - position2[q][1] + p] and Z_H1[
                                                position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                                dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                        can_shu1 = can_shu1 + 1
                                if (can_shu1 == position2[q][1]):
                                    RH1_random.append(
                                        RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                    FH1.append(FH1_1)
                                    RH2_random.append(RH2_random_1)
                                    FH2.append(FH2_1)
                                    print("输出")
                                    break
                                else:
                                    continue
                        if ((i + 1) % 2 == 1):
                            sheet1.cell(k, 5, RH1_random[i])
                            sheet1.cell(k + 1, 5, FH1[i])
                            sheet1.cell(k + 2, 5, FH2[i])
                            sheet1.cell(k + 3, 5, RH2_random[i])
                            sheet1.cell(k, 6).value = 'RB'
                            sheet1.cell(k + 1, 6).value = 'RF'
                            sheet1.cell(k + 2, 6).value = 'RF'
                            sheet1.cell(k + 3, 6).value = 'RB'
                        else:
                            sheet1.cell(k, 5, RH1_random[i])
                            sheet1.cell(k + 1, 5, FH1[i])
                            sheet1.cell(k + 2, 5, FH2[i])
                            sheet1.cell(k + 3, 5, RH2_random[i])
                            sheet1.cell(k, 6).value = 'RF'
                            sheet1.cell(k + 1, 6).value = 'RB'
                            sheet1.cell(k + 2, 6).value = 'RB'
                            sheet1.cell(k + 3, 6).value = 'RF'
                        k = k + 5
            if (flagxx1 == 1):
                continue
            if (dh1_random > 0):
                RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1_1 = RH1_random_1 - dh1_random
                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                FH2_1 = RH2_random_1 - dh2
                if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                        dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                    print("输出")
                    RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1.append(FH1_1)
                    RH2_random.append(RH2_random_1)
                    FH2.append(FH2_1)
                else:
                    for z1 in range(100000):
                        Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                        RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        print("重新选择")
                        if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            print("输出")
                            RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            break
                if ((i + 1) % 2 == 1):
                    sheet1.cell(k, 5, RH1_random[i])
                    sheet1.cell(k + 1, 5, FH1[i])
                    sheet1.cell(k + 2, 5, FH2[i])
                    sheet1.cell(k + 3, 5, RH2_random[i])
                    sheet1.cell(k, 6).value = 'RB'
                    sheet1.cell(k + 1, 6).value = 'RF'
                    sheet1.cell(k + 2, 6).value = 'RF'
                    sheet1.cell(k + 3, 6).value = 'RB'
                else:
                    sheet1.cell(k, 5, FH1[i])
                    sheet1.cell(k + 1, 5, RH1_random[i])
                    sheet1.cell(k + 2, 5, RH2_random[i])
                    sheet1.cell(k + 3, 5, FH2[i])
                    sheet1.cell(k, 6).value = 'RF'
                    sheet1.cell(k + 1, 6).value = 'RB'
                    sheet1.cell(k + 2, 6).value = 'RB'
                    sheet1.cell(k + 3, 6).value = 'RF'
                k = k + 5
            elif (dh1_random < 0):
                RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1_1 = RH1_random_1 - dh1_random
                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                FH2_1 = RH2_random_1 - dh2
                if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                        dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                    print("输出")
                    RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1.append(FH1_1)
                    RH2_random.append(RH2_random_1)
                    FH2.append(FH2_1)
                else:
                    for z1 in range(100000):
                        Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                        RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        print("重新选择")
                        if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            print("输出")
                            RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            break
                if ((i + 1) % 2 == 1):
                    sheet1.cell(k, 5, RH1_random[i])
                    sheet1.cell(k + 1, 5, FH1[i])
                    sheet1.cell(k + 2, 5, FH2[i])
                    sheet1.cell(k + 3, 5, RH2_random[i])
                    sheet1.cell(k, 6).value = 'RB'
                    sheet1.cell(k + 1, 6).value = 'RF'
                    sheet1.cell(k + 2, 6).value = 'RF'
                    sheet1.cell(k + 3, 6).value = 'RB'
                else:
                    sheet1.cell(k, 5, RH1_random[i])
                    sheet1.cell(k + 1, 5, FH1[i])
                    sheet1.cell(k + 2, 5, FH2[i])
                    sheet1.cell(k + 3, 5, RH2_random[i])
                    sheet1.cell(k, 6).value = 'RF'
                    sheet1.cell(k + 1, 6).value = 'RB'
                    sheet1.cell(k + 2, 6).value = 'RB'
                    sheet1.cell(k + 3, 6).value = 'RF'
                k = k + 5
            elif (dh1_random == 0):
                RH1_random_1 = round(random.uniform(0.6, 1.75),
                                     5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1_1 = RH1_random_1 - dh1_random
                RH2_random_1 = RH1_random_1 + Sight_Height_random2
                FH2_1 = RH2_random_1 - dh2
                if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                        dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                    print("输出")
                    RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1.append(FH1_1)
                    RH2_random.append(RH2_random_1)
                    FH2.append(FH2_1)
                else:
                    for z1 in range(100000):
                        Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                        RH1_random_1 = round(random.uniform(0.6, 1.75),
                                             5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1_1 = RH1_random_1 - dh1_random
                        RH2_random_1 = RH1_random_1 + Sight_Height_random2
                        FH2_1 = RH2_random_1 - dh2
                        print("重新选择")
                        if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            print("输出")
                            RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1.append(FH1_1)
                            RH2_random.append(RH2_random_1)
                            FH2.append(FH2_1)
                            break
                if ((i + 1) % 2 == 1):
                    sheet1.cell(k, 5, RH1_random[i])
                    sheet1.cell(k + 1, 5, FH1[i])
                    sheet1.cell(k + 2, 5, FH2[i])
                    sheet1.cell(k + 3, 5, RH2_random[i])
                    sheet1.cell(k, 6).value = 'RB'
                    sheet1.cell(k + 1, 6).value = 'RF'
                    sheet1.cell(k + 2, 6).value = 'RF'
                    sheet1.cell(k + 3, 6).value = 'RB'
                else:
                    sheet1.cell(k, 5, RH1_random[i])
                    sheet1.cell(k + 1, 5, FH1[i])
                    sheet1.cell(k + 2, 5, FH2[i])
                    sheet1.cell(k + 3, 5, RH2_random[i])
                    sheet1.cell(k, 6).value = 'RF'
                    sheet1.cell(k + 1, 6).value = 'RB'
                    sheet1.cell(k + 2, 6).value = 'RB'
                    sheet1.cell(k + 3, 6).value = 'RF'
                k = k + 5
        HD_differ_sum = 0
        k1 = 3
        HDf1 = []
        HDf2 = []
        HDb1 = []
        HDb2 = []
        for i in range(df4.shape[0] - 1):
            HD_random1 = round(random.uniform(-0.2, 0.2), 3)  ##单位m
            HD_random2 = round(random.uniform(-0.2, 0.2), 3)
            HD = df4.iloc[i + 1, 2] / 2
            HDb1.append(HD + HD_random1)
            HDb2.append(HDb1[i] + round(random.uniform(-0.005, 0.005), 3))
            HDf1.append(HD + HD_random2)
            HDf2.append(HDf1[i] + round(random.uniform(-0.005, 0.005), 3))
            HDb = (HDb1[i] + HDb2[i]) / 2
            HDf = (HDf1[i] + HDf2[i]) / 2
            HD_differ = HDb - HDf
            HD_differ_sum = HD_differ_sum + HD_differ
            if (abs(HD_differ) < 1.5 and abs(HD_differ_sum) < 6):
                print('前后视距满足要求')
            else:
                print('前后视距不满足要求')
                print('出现异常将在GUI中提现')
                Exception1 = 1 / 0
            if ((i + 1) % 2 == 1):
                sheet1.cell(k1, 7, HDb1[i])
                sheet1.cell(k1 + 1, 7).value = HDf1[i]
                sheet1.cell(k1 + 2, 7, HDf2[i])
                sheet1.cell(k1 + 3, 7).value = HDb2[i]
                sheet1.cell(k1, 8).value = 'HDB'
                sheet1.cell(k1 + 1, 8).value = 'HDF'
                sheet1.cell(k1 + 2, 8).value = 'HDF'
                sheet1.cell(k1 + 3, 8).value = 'HDB'
                sheet1.cell(k1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
                sheet1.cell(k1 + 1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
                sheet1.cell(k1 + 2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
                sheet1.cell(k1 + 3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
            else:
                sheet1.cell(k1, 7, HDf1[i])
                sheet1.cell(k1 + 1, 7).value = HDb1[i]
                sheet1.cell(k1 + 2, 7, HDb2[i])
                sheet1.cell(k1 + 3, 7).value = HDf2[i]
                sheet1.cell(k1, 8).value = 'HDF'
                sheet1.cell(k1 + 1, 8).value = 'HDB'
                sheet1.cell(k1 + 2, 8).value = 'HDB'
                sheet1.cell(k1 + 3, 8).value = 'HDF'
                sheet1.cell(k1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
                sheet1.cell(k1 + 1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
                sheet1.cell(k1 + 2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
                sheet1.cell(k1 + 3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
            if (HD_differ_sum > 6):
                print(
                    '****************************************************************************************************')
                print('前后视距累计差抄超限')
            k1 = k1 + 5
        # from ce1_8 import position_index
        position = self.position_index(df3)
        Z_name = []
        Z_H = []
        Z_HD = []
        Z_HF = []
        for zz1 in range(len(position)):
            HD_random3 = round(random.uniform(-0.5, 0.5), 3)  ##单位m
            Z_name.append(df3.loc[position[zz1][0], 0])
            Z_HD.append(df3.loc[position[zz1][0], 2] + HD_random3)  ###
            Z_DH1 = df3.loc[dict1[position[zz1][1]], 1] - df3.loc[position[zz1][0], 1]
            Z_H.append(df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - Z_DH1)
            Z_HB = RH1_random[position[zz1][1]]
            Z_HF.append(Z_HB + Z_DH1)
            if (Z_HB + df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - 1.75 < Z_H[zz1] and Z_H[
                zz1] < Z_HB + df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - 0.6):
                print("支点的高程设置合理！！！！！！！！！！！")
            else:
                print("支点的高程设置不合理****************")
        start_num = 3
        start_num_fei = 0
        continue_times = 0
        nn = 0
        kx1 = 0
        flag = 0
        while (flag == 0):
            for zz1 in range(kx1, len(position)):
                for zz2 in range(start_num_fei, position[zz1][1] + 1):
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                    if ((zz2 + 1) % 2 == 1):
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = df4.iloc[0:df4.shape[0], 0].iloc[
                            zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                        name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                        sheet1[name1].font = Font(bold=True)
                    else:
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = df4.iloc[0:df4.shape[0], 0].iloc[
                            zz2 + 1]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                        name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                        sheet1[name1].font = Font(bold=True)
                kx2 = 0
                for zz3 in range(zz1, len(position)):
                    if (position[zz1][1] == position[zz3][1]):
                        kx2 = kx2 + 1
                zz3 = zz1 + kx2
                for zz4 in range(zz3 - zz1):
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 1).value = df3.loc[
                        position[zz1 + zz4][0], 0]  ##中间点的点名
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 5).value = Z_HF[zz1 + zz4]  ##中间点视线高
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 6).value = 'RZ'  ##中间点视线RZ
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 7).value = Z_HD[zz1 + zz4]  ##中间点视距
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 8).value = 'HD'  ##中间点视线高
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 9).value = Z_H[zz1 + zz4]  ##中间点视线高
                kx1 = zz3 - zz1 + kx1
                start_num = start_num + 5 * (zz2 - start_num_fei) + 4 + zz3 - zz1 + 1
                start_num_fei = position[zz1][1] + 1
                if (kx1 == len(position)):
                    ##收尾
                    flag = 1
                    if (position[zz1][1] == df4.shape[0] - 1 - 1):
                        pass
                    else:
                        for zz2 in range(start_num_fei, df4.shape[0] - 1):
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                            if ((zz2 + 1) % 2 == 1):
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                                sheet1[name1].font = Font(bold=True)
                            else:
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                                    df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                                name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                                sheet1[name1].font = Font(bold=True)
                    break
                else:
                    pass
                break

        sheet1['A2'].font = Font(bold=True)
        for i in range(df4.shape[0] - 1):
            name1 = 'A' + str(7 + 5 * i)
            name2 = 'A' + str(7 + 5 * i)
            sheet1[name2] = df4.iloc[i + 1, 0]
            sheet1[name1].font = Font(bold=True)
        book1.save(path1)
        book1.close()
        book2.save(path2)
        book2.close()

    def baoliu(self,a):
        ling1 = '0'
        ling2 = '00'
        ling3 = '000'
        if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数4为 字符串
            ag = a + '0000'
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
            ag = a + ling3
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
            ag = a + ling2
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
            ag = a + ling1
        else:
            ag = a
        return ag

    def baoliu3(self,a):
        ling1 = '0'
        ling2 = '00'
        if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数3为 字符串
            ag = a + '000'
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
            ag = a + ling2
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
            ag = a + ling1
        else:
            ag = a
        return ag

    def baoliu5(self,a):
        ling1 = '0'
        ling2 = '00'
        ling3 = '000'
        ling4 = '0000'
        if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数5为 字符串
            ag = a + '00000'
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
            ag = a + ling4
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
            ag = a + ling3
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
            ag = a + ling2
        elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 4):
            ag = a + ling1
        else:
            ag = a
        return ag
    def function_cj(self):
        # measure_line_path = "D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"#D:\Desktop\测试期次1\测试线路文件1.xlsx##self.editline1.text()
        # dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'#D:\Desktop\测试期次1\测试数据库1.xlsx'#self.editline2.text()#self.editline2.text()
        # path_output1 = 'D:\\Desktop\\测试期次1\\测试output\\'#self.editline3.text()#'D:\Desktop\测试期次1\测试output\self.editline3.text()
        measure_line_path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
        dateset_path =  QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
        src_path=os.path.abspath(os.path.join(measure_line_path, ".."))+'\\' #############################################################
        output_name = "报表"  ###输出的文件夹名字
        if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
            pass
            print(output_name + "已经存在！")
        else:
            print("创建文件夹" + output_name + "!")
            os.makedirs(src_path + output_name)
        path_output1 = src_path + output_name
        qi_shu = self.editline2.text()
        BC_range1 = self.editline1.text()
        BC_range_min = re.findall(r'(.*) ', BC_range1,flags=0)[0]
        BC_range_max = re.findall(r' (.*)', BC_range1,flags=0)[0]
        print(measure_line_path)
        print(dateset_path)
        print(path_output1)
        print(qi_shu)
        print(BC_range_min, BC_range_max)
        RZ_names, RZ_values = self.read_measure_line_from_dataset(dateset_path, measure_line_path)
        for m in range(int(qi_shu)):  # len(RZ_values)
            BC1 = round(random.uniform(float(BC_range_min), float(BC_range_max)), 5)
            path_output2 = path_output1 +'\\'+ str(m + 1) + '期' + '原始数据' + '.xlsx'
            path_output3 = path_output1 + '\\'+str(m + 1) + '期' + '原始数据' + '无支点' + '.xlsx'
            df = pd.read_excel(measure_line_path, '测线')  # 对应的导出为pd.to_excel(***.xlsx)
            df.columns = range(0, 3)
            name = RZ_names
            value = RZ_values[m]
            dict = {name[i]: value[i] for i in range(len(name))}
            for i in range(0, df.shape[0]):
                if (df.loc[i, 0] in name):
                    df.loc[i, 1] = dict[df.loc[i, 0]]
            df1 = df[:1]
            df2 = df[1:len(df)]
            df21 = df2.dropna(how='any')
            df3 = pd.concat([df1, df21], axis=0)
            df3.index = range(df3.shape[0])
            a = df3.iloc[:, [1]]
            b = range(df3.shape[0])
            num = []
            numx = []
            for i in range(df3.shape[0]):
                if (df3.loc[i, 0] in name):
                    numx.append(i)
                    pass
                else:
                    num.append(i)
            instrument_height = 1.4
            df4 = df3.loc[num, [0, 1, 2]]
            dict1 = {i: num[i] for i in range(df4.shape[0])}
            book1 = openpyxl.Workbook()
            sheet1 = book1.create_sheet('原始数据-无支点', 0)
            book2 = openpyxl.Workbook()
            sheet2 = book2.create_sheet('原始数据-支点', 0)
            self.sight_height_distance(df3, df4, book1, sheet1, book2, sheet2, BC1, path_output3, path_output2, num,
                                       dict1)
        pat1 = path_output1+'\\'
        pat_name = os.listdir(pat1)
        SUM_RB = []
        SUM_RF = []
        SUM_RB_CEZHAN = []
        DATE = []
        INDEX3 = []
        for i in range(len(pat_name)):
            if (pat_name[i][-4:] == 'xlsx'):
                if (re.findall(r'数据(.*)', pat_name[i])[0] != '无支点.xlsx'):
                    index3 = re.findall(r'(.*)期', pat_name[i])[0]
                    INDEX3.append(index3)
                    book2 = load_workbook(dateset_path)
                    sheet2 = book2.get_sheet_by_name('建筑沉降成果表')
                    nrows2 = sheet2.max_row
                    for z3 in range(13, nrows2 + 1):
                        if (int(index3) == sheet2.cell(z3, 1).value):
                            DATE.append(sheet2.cell(z3, 2).value)
                            break
                    book1 = load_workbook(pat1 + pat_name[i])
                    sheet1 = book1.get_sheet_by_name('原始数据-支点')
                    nrows = sheet1.max_row
                    ncols = sheet1.max_column
                    sum_RB = 0
                    sum_RF = 0
                    sum_RB_cezhan = 0
                    for j in range(3, nrows + 1):
                        sheet1.cell(j, 3).value = datetime.timedelta(
                            seconds=28800 + (j - 3) * 15)  # print(datetime.timedelta(seconds=28800))
                        if (sheet1.cell(j, 6).value == 'RB'):
                            sum_RB = sum_RB + sheet1.cell(j, 7).value
                            sum_RB_cezhan = sum_RB_cezhan + 1
                        if (sheet1.cell(j, 6).value == 'RF'):
                            sum_RF = sum_RF + sheet1.cell(j, 7).value
                    book1.save(pat1 + pat_name[i])
                    SUM_RB.append(sum_RB)
                    SUM_RF.append(sum_RF)
                    SUM_RB_CEZHAN.append(sum_RB_cezhan)

        num_1 = 0
        for z1 in range(len(pat_name)):
            # path="D:\\Desktop\\测试期次1\\测试output\\10期原始数据1.xlsx"
            path1 = path_output1+'\\'
            if (pat_name[z1][-4:] == 'xlsx'):
                if (re.findall(r'数据(.*)', pat_name[z1])[0] != '无支点.xlsx'):
                    num_1 = num_1 + 1
                    path = path1 + pat_name[z1]
                    data1 = pd.read_excel(path, '原始数据-支点')
                    fid1 = open(path1 + '沉降观测数据第' + INDEX3[num_1 - 1] + '期' + '.txt', 'w')
                    date3 = str(DATE[num_1 - 1])
                    x1 = date3.replace('-', '年', 1)
                    x2 = x1.replace('-', '月', 1)
                    x3 = x2.replace(' 00:00:00', '日', 1)
                    year1 = re.findall(r'(.*)年', x3, flags=0)[0]
                    month1 = re.findall(r'年(.*)月', x3, flags=0)[0]
                    day1 = re.findall(r'月(.*)日', x3, flags=0)[0]
                    if (len(month1) == 1):
                        month1 = '0' + month1
                    if (len(day1) == 1):
                        day1 = '0' + day1
                    date4 = year1 + month1 + day1
                    fid1.write('期次：' + INDEX3[num_1 - 1] + ' ' + '日期：' + x3 + '\n')
                    fid1.write(
                        'For M5|Adr' + '   ' + '  1' + '|' + 'TO' + '  ' + date4 + '.dat' + '               ' + "|" + "                      " + "|" + "                      " + "|" + "                      " + "|" + "\n")
                    fid1.write(
                        'For M5|Adr' + '   ' + '  2' + '|' + 'TO' + '  ' + 'Start-Line' + '      ' + 'aBFFB' + '     ' + '1' + '|' + "                      " + "|" + "                      " + "|" + "                      " + "|" + "\n")
                    rown = data1.shape[0]
                    coln = data1.shape[1]
                    i = 0
                    for i in range(0, rown):
                        dm = data1.iloc[i, 0]
                        if (i + 2 < 10):
                            index1 = '  ' + str(i + 2)
                        elif ((i + 2 > 10 and i + 2 < 100) or i + 2 == 10):
                            index1 = ' ' + str(i + 2)
                        else:
                            index1 = str(i + 2)
                        kn = ''
                        for j in range(9 - len(dm)):
                            kn = kn + ' '
                        dmlength = kn + dm
                        if (i == 0):
                            fid1.write(
                                'For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '                  ' + '1' + '|' + "                      " + '|' + "                      " + '|' + 'Z' + '        ' + self.baoliu(
                                    str(data1.iloc[i, coln - 1])) + ' m' + '    ' + '|' + '\n')
                        else:
                            if ((~numpy.isnan(data1.iloc[i, 4])) and (~numpy.isnan(data1.iloc[i, coln - 1]))):
                                ##支点
                                if (len(self.baoliu3(str(data1.iloc[i, 6]))) == 6):
                                    fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                        data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, 4],
                                                  5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '         ' + self.baoliu3(
                                        str(round(data1.iloc[i, 6], 3))) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, coln - 1], 5))) + ' m' + '   ' + '|' + '\n')
                                else:
                                    fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                        data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, 4],
                                                  5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '          ' + self.baoliu3(
                                        str(round(data1.iloc[i, 6], 3))) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, coln - 1], 5))) + ' m' + '   ' + '|' + '\n')

                            elif ((~numpy.isnan(data1.iloc[i, 4])) and (numpy.isnan(data1.iloc[i, coln - 1]))):
                                ##后视点或者前视点
                                if (len(self.baoliu3(str(data1.iloc[i, 6]))) == 6):
                                    fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                        data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, 4],
                                                  5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '         ' + self.baoliu3(
                                        str(round(data1.iloc[i, 6],
                                                  3))) + ' m' + '   ' + '|' + '                      ' + '|' + '\n')
                                else:
                                    fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                        data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                        str(round(data1.iloc[i, 4],
                                                  5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '          ' + self.baoliu3(
                                        str(round(data1.iloc[i, 6],
                                                  3))) + ' m' + '   ' + '|' + '                      ' + '|' + '\n')
                            else:
                                fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                    data1.iloc[
                                        i, 2]) + '    ' + '1' + '|' + '                      ' + '|' + '                      ' + '|' + 'Z' + '        ' + self.baoliu5(
                                    str(data1.iloc[i, coln - 1])) + ' m' + '   ' + '|' + '\n')
                    index2 = i + 2 + 1
                    index3 = []
                    for z2 in range(3):
                        if (index2 + z2 < 10):
                            index3.append('  ' + str(index2 + z2))
                        elif ((index2 + z2 > 10 and index2 + z2 < 100) or index2 + z2 == 10):
                            index3.append(' ' + str(index2 + z2))
                        else:
                            index3.append(str(index2 + z2))
                    dm1 = data1.iloc[i, 0]
                    kn = ''
                    for j in range(9 - len(dm1)):
                        kn = kn + ' '
                    dmlength1 = kn + dm1
                    if (round(data1.iloc[i, 8] - data1.iloc[0, 8], 5) < 0):
                        fid1.write('For M5|Adr' + '   ' + index3[
                            0] + '|' + 'KD1' + dmlength1 + '                  ' + '1' + '|' + 'Sh' + '       ' + self.baoliu5(
                            '%.5f' % (data1.iloc[i, 8] - data1.iloc[
                                0, 8])) + ' m' + '   ' + '|' + '                      |                      |' + '\n')
                    else:
                        fid1.write('For M5|Adr' + '   ' + index3[
                            0] + '|' + 'KD1' + dmlength1 + '                  ' + '1' + '|' + 'Sh' + '        ' + self.baoliu5(
                            '%.5f' % (data1.iloc[i, 8] - data1.iloc[
                                0, 8])) + ' m' + '   ' + '|' + '                      |                      |' + '\n')
                    czs = int(SUM_RB_CEZHAN[num_1 - 1] / 2)
                    kn = ''
                    for j in range(4 - len(str(czs))):
                        kn = kn + ' '
                    czs1 = str(czs) + kn

                    dbz = SUM_RB[num_1 - 1] / 2
                    dfz = SUM_RF[num_1 - 1] / 2
                    if (len(re.findall(r'(.*)[.]', str(round(dbz, 3)))[0]) == 3):
                        fid1.write('For M5|Adr' + '   ' + index3[
                            1] + '|' + 'KD1' + dmlength1 + '      ' + czs1 + '        ' + '1' + '|' + 'Db' + '        ' + self.baoliu3(
                            str(round(dbz, 3))) + ' m' + '   ' + '|' + 'Df' + '        ' + self.baoliu3(
                            str(round(dfz, 3)) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu(
                                str(data1.iloc[i, coln - 1])) + ' m' + '   ' + '|' + '\n'))
                    else:
                        fid1.write('For M5|Adr' + '   ' + index3[
                            1] + '|' + 'KD1' + dmlength1 + '      ' + czs1 + '        ' + '1' + '|' + 'Db' + '        ' + self.baoliu3(
                            str(round(dbz, 3))) + '  m' + '   ' + '|' + 'Df' + '        ' + self.baoliu3(
                            str(round(dfz, 3)) + '  m' + '   ' + '|' + 'Z' + '        ' + self.baoliu(
                                str(data1.iloc[i, coln - 1])) + ' m' + '   ' + '|' + '\n'))
                    fid1.write('For M5|Adr' + '   ' + index3[
                        1] + '|' + 'TO' + '  ' + 'End-Line' + '                  ' + '1' + '|' + '                      |                      |                      |')
                    fid1.close()
class Jikeng_make_dir(QDialog):
    def __init__(self):
        super(Jikeng_make_dir,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('文件创立')
        self.label1=QLabel('第一次弹出选择：数据库位置',self)
        self.label2=QLabel('第一次弹出选择：原始数据存放的位置',self)
        self.button1=QPushButton(self)
        self.button1.setText('开始创建原始数据文件')
        self.V_layout1=QVBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.create_dir)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def create_dir(self):
        # path1 = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据9x\\"
        # path3 = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据3\\"
        # dateset_path = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据2\\4-2数据库.xlsx"
        dateset_path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
        src_path=os.path.abspath(os.path.join(dateset_path, ".."))+'\\' #############################################################
        output_name = "报表"  ###输出的文件夹名字
        if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
            print(output_name + "已经存在！")
        else:
            print("创建文件夹" + output_name + "!")
            os.makedirs(src_path + output_name)
        path1=src_path + output_name+'\\'
        path3 = os.path.abspath(QFileDialog.getExistingDirectory(self,"选取文件夹",''))+'\\' # 起始路径
        # path1=self.editline2.text()
        # path3=self.editline1.text()
        # dateset_path=self.editline4.text()
        workbook = load_workbook(dateset_path)
        sheet1 = workbook.get_sheet_by_name("日报")
        for zz in range(2, sheet1.max_row + 1):
            if (((sheet1.cell(zz, 2).value is None) or (sheet1.cell(zz, 2).value == '')) or (
                    sheet1.cell(zz, 2).value == [])):  # 根据判断是否为无参数None或者空来判断行数
                row2 = zz - 1
                break
            else:
                row2 = sheet1.max_row  # 根据日报获取期数
        file_names = os.listdir(path3)
        a = '周边道路'  # a='周边道路' #####前提需要是周边道路沉降.dat文件类似的
        b = '周边管线'  # b='周边管线' #####前提需要是周边道路沉降.dat文件类似的
        c = '周边地表'  # c='周边地表' #####前提需要是周边道路沉降.dat文件类似的
        d = '周边建筑'  # d='周边建筑' #####前提需要是周边道路沉降.dat文件类似的
        e = '坡顶沉降'  # e='坡顶沉降' #####前提需要是周边道路沉降.dat文件类似的
        f = '桩顶沉降'
        g = '坡顶位移'
        flagg = 0
        numg = []
        nameg = []
        flagh = 0
        numh = []
        nameh = []
        for i in range(len(file_names)):
            if ('坐标' in file_names[i]):
                numg.append(int(re.findall('(.*)观测记录', file_names[i], flags=0)[0]))
                nameg.append(file_names[i])
                flagg = 1
        for i in range(len(file_names)):
            if ('监测日报' in file_names[i]):
                numh.append(int(re.findall('(.*)监测日报', file_names[i], flags=0)[0]))
                nameh.append(file_names[i])
                flagh = 1
        numa = []
        numb = []
        numc = []
        numd = []
        nume = []
        numf = []
        flaga = 0
        flagb = 0
        flagc = 0
        flagd = 0
        flage = 0
        flagf = 0
        namea = []
        nameb = []
        namec = []
        named = []
        namee = []
        namef = []
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (a in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numa.append(qi_num)
                    namea.append(file_names[i])
                    flaga = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (b in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numb.append(qi_num)
                    nameb.append(file_names[i])
                    flagb = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (c in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numc.append(qi_num)
                    namec.append(file_names[i])
                    flagc = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (d in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numd.append(qi_num)
                    named.append(file_names[i])
                    flagd = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (e in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    nume.append(qi_num)
                    namee.append(file_names[i])
                    flage = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (f in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numf.append(qi_num)
                    namef.append(file_names[i])
                    flagf = 1
        if (flaga == 0):
            numa = [0]
        if (flagb == 0):
            numb = [0]
        if (flagc == 0):
            numc = [0]
        if (flagd == 0):
            numd = [0]
        if (flage == 0):
            nume = [0]
        if (flagf == 0):
            numf = [0]
        if (flagg == 0):
            numg = [0]
        if (flagh == 0):
            numh = [0]
        num_max = np.max(
            [np.max(numa), np.max(numb), np.max(numc), np.max(numd), np.max(nume), np.max(numf), np.max(numg),
             np.max(numh)])  ###共a、b、c、d、e、f、g、h比较
        if (flaga == 0):
            numa = [1000000]
        if (flagb == 0):
            numb = [1000000]
        if (flagc == 0):
            numc = [1000000]
        if (flagd == 0):
            numd = [1000000]
        if (flage == 0):
            nume = [1000000]
        if (flagf == 0):
            numf = [1000000]
        if (flagg == 0):
            numg = [1000000]
        num_min = np.min(
            [np.min(numa), np.min(numb), np.min(numc), np.min(numd), np.min(nume), np.min(numf), np.min(numg),
             np.min(numg)])  ###共a、b、c、d、e、f、g、h比较
        qua_path = "D:\\2021\\基坑监测\\2018巷口\\质量评定.doc"
        for j in range(num_min + 1, num_max + 2):  ##num_max>num_min############需要有两期以上
            date1 = sheet1.cell(j, 2).value
            date3 = "第" + str(j - 1) + "期" + str(date1)
            x1 = date3.replace('-', '年', 1)
            x2 = x1.replace('-', '月', 1)
            x3 = x2.replace(' 00:00:00', '日', 1)  # 通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
            datex2 = path1 + x3
            if (os.path.isdir(datex2)):  ###判断输出文件架是否存在，如果存在不创建
                print(x3+"期已经存在已经存在！")
            else:
                print("创建文件夹" + x3 + "!")
                os.makedirs(datex2)
            datex3 = datex2 + '\\' + '原始数据'
            if (os.path.isdir(datex2 + '\\' + '原始数据')):  ###判断输出文件架是否存在，如果存在不创建
                pass
            else:
                os.makedirs(datex2 + '\\' + '原始数据')
            if (flagh == 1):
                os.makedirs(datex2 + '\\' + '日报')
                for name in nameh:
                    if (re.findall('(.*)监测日报', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '日报')  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagg == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + g)
                for name in nameg:
                    if (re.findall('(.*)观测记录', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + g)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flaga == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + a)
                for name in namea:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + a)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagb == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + b)
                for name in nameb:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + b)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagc == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + c)
                for name in namec:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + c)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagd == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + d)
                for name in named:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + d)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flage == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + e)
                for name in namee:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + e)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagf == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + f)
                for name in namef:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + f)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下

class Jikeng_rename(QDialog):
    def __init__(self):
        super(Jikeng_rename,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('原始数据DAT文件进行重名')
        self.label1=QLabel('第一步确定原始文件DAT文件是否是’水准观测文件',self)
        self.label2=QLabel('第二步确定父文件夹是否是坡顶沉降、桩顶沉降、周边环境沉降、周边道路沉降、周边管线沉降、周边建筑沉降、周边地表沉降',self)
        self.label3=QLabel('第三步选择要重命名文件的父文件夹路径',self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)
    def display_editline(self):
        self.button1.clicked.connect(self.dat_file_rename)
    def dat_file_rename(self):
        path1=os.path.abspath(QFileDialog.getExistingDirectory(self,"选取文件夹",'')) # 起始路径
        src_name=os.path.basename(path1) ####################################
        file_name=os.listdir(path1)
        for name in file_name:
            num=re.findall(r'水准观测文件(.*)',name,flags=0)[0]
            os.rename(path1+'\\'+name,path1+'\\'+src_name+num)

class bihecha_get(QDialog):
    def __init__(self):
        super(bihecha_get, self).__init__()
        self.resize(600, 300)
        self.setWindowTitle('闭合差提取')
        self.label1 = QLabel('第一步：打开沉降原始dat文件所在位置', self)
        self.button1 = QPushButton(self)
        self.button1.setText('开始生成闭合差数据')
        self.V_layout1 = QVBoxLayout()
        self.Layout__init()
        self.display_editline()

    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.button1)
        self.setLayout(self.V_layout1)

    def display_editline(self):
        self.button1.clicked.connect(self.bihecha)
    def bihecha(self):
        path1=os.path.abspath(QFileDialog.getExistingDirectory(self,"选取文件夹",''))+'\\' # 起始路径
        filename=os.listdir(path1)
        book1=openpyxl.Workbook()
        sheet1=book1.create_sheet('闭合差')
        for i in range(len(filename)-1):
            for i in range(len(filename) - 1):
                for j in range(i + 1, len(filename)):
                    a = int(re.findall(r'\d+', filename[i])[0])
                    b = int(re.findall(r'\d+', filename[j])[0])
                    if (a > b):
                        mid1 = filename[i]
                        filename[i] = filename[j]
                        filename[j] = mid1
        k = 1
        for i in range(len(filename)):
            filename1 = path1 + filename[i]
            filename2 = os.listdir(filename1)
            k1 = 1
            for j in range(len(filename2)):
                for z in range(len(filename2)):
                    if ('第' + str(j + 1) + '期' in filename2[z]):
                        filename3 = os.listdir(filename1 + '\\' + filename2[z])
                        for z1 in range(len(filename3)):
                            if ('dat' in filename3[z1]):
                                filename4 = filename3[z1]
                        f = open(filename1 + '\\' + filename2[z] + '\\' + filename4, 'r')
                        file1 = f.readlines()
                        count1 = len(file1)
                        r1 = file1[count1 - 3 - 1][58:66]
                        # print(file1[count1-3-1])
                        f.close()
                        sheet1.cell(k1, k).value = float(r1) * 1000
                        k1 = k1 + 1
            k = k + 1
        book1.save('闭合差1.xlsx')

if __name__== '__main__':
    app=QApplication(sys.argv)
    demo=Data_Process()
    demo.show()
    sys.exit(app.exec_())