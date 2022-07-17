import os
import shutil
from openpyxl import load_workbook
import re
import math
import numpy as np
# path1="D:\\Desktop\\吴昌程还原原始数据\\原始数据\\" #目标文件路径
# path3="D:\\Desktop\\吴昌程还原原始数据\\1吴\\" #源文件路径
# dateset_path="D:\\Desktop\\吴昌程还原原始数据\\抚州保利数据库.xlsx"
path1="D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据2\\原始数据\\"
path3="D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据2\\报表\\"
dateset_path="D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据2\\4-2数据库.xlsx"
workbook = load_workbook(dateset_path)
sheet1=workbook.get_sheet_by_name("日报")
for zz in range(2,sheet1.max_row+1):
    if(((sheet1.cell(zz,2).value is None) or (sheet1.cell(zz,2).value=='')) or (sheet1.cell(zz,2).value==[])):#根据判断是否为无参数None或者空来判断行数
        row2=zz-1
        break
    else:
        row2=sheet1.max_row #根据日报获取期数
file_names = os.listdir(path3)
a='道路'   #a='周边道路'
b='管线'   #b='周边管线'
c='周边环境' #c='周边地表'
d='建筑'    #d='周边建筑'
e='坡顶沉降'   #e='坡顶沉降'
f='桩顶沉降'
g='坡顶位移'
flagg=0
numg=[]
nameg=[]
flagh=0
numh=[]
nameh=[]
for i in range(len(file_names)):
    if('坐标' in file_names[i]):
        numg.append(int(re.findall('(.*)观测记录', file_names[i], flags=0)[0]))
        nameg.append(file_names[i])
        flagg=1
for i in range(len(file_names)):
    if('监测日报' in file_names[i]):
        numh.append(int(re.findall('(.*)监测日报', file_names[i], flags=0)[0]))
        nameh.append(file_names[i])
        flagh=1
numa=[]
numb=[]
numc=[]
numd=[]
nume=[]
numf=[]
flaga=0
flagb=0
flagc=0
flagd=0
flage=0
flagf=0
namea=[]
nameb=[]
namec=[]
named=[]
namee=[]
namef=[]
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(a in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            numa.append(qi_num)
            namea.append(file_names[i])
            flaga=1
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(b in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            numb.append(qi_num)
            nameb.append(file_names[i])
            flagb=1
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(c in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            numc.append(qi_num)
            namec.append(file_names[i])
            flagc=1
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(d in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            numd.append(qi_num)
            named.append(file_names[i])
            flagd=1
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(e in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            nume.append(qi_num)
            namee.append(file_names[i])
            flage=1
for i in range(len(file_names)):
    if('水准观测文件' in file_names[i]): ##属于水准沉降原始文件
        if(f in file_names[i]):
            qi_num=int(re.findall('件(.*)[.]',file_names[i],flags=0)[0])
            numf.append(qi_num)
            namef.append(file_names[i])
            flagf=1
if(flaga==0):
    numa=[0]
if(flagb==0):
    numb=[0]
if(flagc==0):
    numc=[0]
if(flagd==0):
    numd=[0]
if(flage==0):
    nume=[0]
if(flagf==0):
    numf=[0]
if (flagg == 0):
    numg = [0]
if (flagh == 0):
    numh = [0]
num_max=np.max([np.max(numa),np.max(numb),np.max(numc),np.max(numd),np.max(nume),np.max(numf),np.max(numg),np.max(numh)]) ###共a、b、c、d、e、f、g、h比较
if(flaga==0):
    numa=[1000000]
if(flagb==0):
    numb=[1000000]
if(flagc==0):
    numc=[1000000]
if(flagd==0):
    numd=[1000000]
if(flage==0):
    nume=[1000000]
if(flagf==0):
    numf=[1000000]
if(flagg==0):
    numg=[1000000]
num_min=np.min([np.min(numa),np.min(numb),np.min(numc),np.min(numd),np.min(nume),np.min(numf),np.min(numg),np.min(numg)])###共a、b、c、d、e、f、g、h比较
qua_path="D:\\2021\\基坑监测\\2018巷口\\质量评定.doc"
for j in range(num_min+1,num_max+2):  ##num_max>num_min############需要有两期以上
    date1=sheet1.cell(j,2).value
    date3="第"+str(j-1)+"期"+str(date1)
    x1=date3.replace('-','年',1)
    x2=x1.replace('-','月',1)
    x3=x2.replace(' 00:00:00','日',1) #通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
    datex2 = path1 + x3
    os.makedirs(datex2)
    datex3=datex2+'\\'+'原始数据'
    os.makedirs(datex2+'\\'+'原始数据')
    if(flagh==1):
        os.makedirs(datex2 + '\\' + '日报')
        for name in nameh:
            if (re.findall('(.*)监测日报', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '日报')  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flagg==1):
        os.makedirs(datex2 + '\\' + '原始数据' + '\\' + g)
        for name in nameg:
            if (re.findall('(.*)观测记录', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+g)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flaga==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+a)
        for name in namea:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+a)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flagb==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+b)
        for name in nameb:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+b)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flagc==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+c)
        for name in namec:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+c)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flagd==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+d)
        for name in named:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+d)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flage==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+e)
        for name in namee:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+e)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
    if(flagf==1):
        os.makedirs(datex2 + '\\' + '原始数据'+'\\'+f)
        for name in namef:
            if (re.findall('件(.*)[.]', name, flags=0)[0] == str(j - 1)):
                shutil.copy(path3+name, datex2 + '\\' + '原始数据'+'\\'+f)  # 将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下

