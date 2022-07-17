import numpy
import pandas as pd
import numpy as np
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from 沉降观测1 import read_measure_line_from_dataset
import ce1_8
from ce1_8 import position_index1
from ce1_8 import Z_H_function
def sight_height_distance(df3,df4,book1,sheet1,book2,sheet2,BC1,path1,path2,num,dict1):
    sheet1.cell(1,1).value='点位'
    sheet1.cell(1,3).value='时间'
    sheet1.cell(1,5).value='视线高'
    sheet1.cell(1,7).value='视距'
    sheet1.cell(1,9).value='高程'
    sheet1.cell(2,1).value=df3.iloc[0,0]
    sheet1.cell(2,9).value=df4.iloc[0,1]
    sheet2.cell(1,1).value='点位'
    sheet2.cell(1,3).value='时间'
    sheet2.cell(1,5).value='视线高'
    sheet2.cell(1,7).value='视距'
    sheet2.cell(1,9).value='高程'
    sheet2.cell(2,1).value=df3.iloc[0,0]
    sheet2.cell(2,9).value=df4.iloc[0,1]
    k=3
    RH1_random = []
    RH2_random=[]
    FH1=[]
    FH2=[]
    height1=[]
    position1 = ce1_8.position_index(df3)
    position2 = position_index1(df3)
    Z_H1 = Z_H_function(df3, df4, num,BC1)
    for i in range(df4.shape[0]-1):
        dh=df4.iloc[i+1,1]-df4.iloc[i,1]-BC1
        if(dh>1.13):
            print("高差大于1.13m，高差过大") #因为dh1_random的值不能超过1.15m
            xx1=1/0
        height1.append(df4.iloc[i+1,1]-(i+1)*BC1)
        sheet1.cell(7+5*i,9).value=height1[i]
        #'高差值最大是1.8-0.55=1.25m' 且 dh_random的取值范围为0.55+dh~1.8
        Sight_Height_random1=random.randint(-5, 15) * 0.00001  ##单位m
        Sight_Height_random2 = random.randint(-5, 15) * 0.00001  ##单位m0.15mm*2=0.3m
        dh1_random=dh+Sight_Height_random1  ##FH2-FH1=2*Sight_Height_random1+Sight_Height_random2
        dh2=2*dh-dh1_random
        flagxx1=0
        for q in range(len(position2)):
            if(i==position2[q][0]):
                #############加入
                pass
                flagxx1=1
                if (dh1_random > 0):
                    RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu=0
                    for p in range(position2[q][1]):
                        if ( ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006 ) and ( RH1_random_1+df3.loc[dict1[i],1]-i*BC1-1.75 < Z_H1[position2[q][2]-position2[q][1]+p] and Z_H1[position2[q][2]-position2[q][1]+p] < RH1_random_1+df3.loc[dict1[i],1]-i*BC1-0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu=can_shu+1
                    if(can_shu==position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, FH1[i])
                        sheet1.cell(k + 1, 5, RH1_random[i])
                        sheet1.cell(k + 2, 5, RH2_random[i])
                        sheet1.cell(k + 3, 5, FH2[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
                elif (dh1_random < 0):
                    #############加入
                    RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu = 0
                    for p in range(position2[q][1]):
                        if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):
                            # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu = can_shu + 1
                    if (can_shu == position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
                elif (dh1_random == 0):
                    RH1_random_1 = round(random.uniform(0.6, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu=0
                    for p in range(position2[q][1]):
                        if ( ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006 ) and ( RH1_random_1+df3.loc[dict1[i],1]-i*BC1-1.75 < Z_H1[position2[q][2]-position2[q][1]+p] and Z_H1[position2[q][2]-position2[q][1]+p] < RH1_random_1+df3.loc[dict1[i],1]-i*BC1-0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu=can_shu+1
                    if(can_shu==position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
        if(flagxx1==1):
            continue
        if (dh1_random > 0):
            RH1_random_1=round(random.uniform(0.6 + dh1_random,1.75),5)###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 - dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2=random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6 + dh1_random, 1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, FH1[i])
                sheet1.cell(k + 1, 5, RH1_random[i])
                sheet1.cell(k + 2, 5, RH2_random[i])
                sheet1.cell(k + 3, 5, FH2[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k = k + 5
        elif(dh1_random<0):
            RH1_random_1=round(random.uniform(0.6,1.75+dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 -dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6, 1.75 + dh1_random),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k=k+5
        elif(dh1_random==0):
            RH1_random_1=round(random.uniform(0.6,1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 -dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6, 1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k=k+5
    HD_differ_sum=0
    k1=3
    HDf1=[]
    HDf2=[]
    HDb1=[]
    HDb2=[]
    for i in range(df4.shape[0]-1):
        HD_random1 = round(random.uniform(-0.5, 0.5),3)##单位m
        HD_random2 = round(random.uniform(-0.5, 0.5),3)
        HD=df4.iloc[i+1,2]/2
        HDb1.append(HD+HD_random1)
        HDb2.append(HDb1[i]+round(random.uniform(-0.005,0.005),3))
        HDf1.append(HD+HD_random2)
        HDf2.append(HDf1[i]+round(random.uniform(-0.005,0.005),3))
        HDb=(HDb1[i]+HDb2[i])/2
        HDf=(HDf1[i]+HDf2[i])/2
        HD_differ=HDb-HDf
        HD_differ_sum=HD_differ_sum+HD_differ
        if(HD_differ<1.5 and HD_differ_sum<6):
            print('前后视距满足要求')
        else:
            print('前后视距不满足要求')
            print('出现异常将在GUI中提现')
            Exception1=1/0
        if((i+1)%2==1):
            sheet1.cell(k1,7,HDb1[i])
            sheet1.cell(k1+1,7).value=HDf1[i]
            sheet1.cell(k1+2,7,HDf2[i])
            sheet1.cell(k1+3,7).value=HDb2[i]
            sheet1.cell(k1,8).value='HDB'
            sheet1.cell(k1+1,8).value='HDF'
            sheet1.cell(k1+2,8).value='HDF'
            sheet1.cell(k1+3,8).value='HDB'
            sheet1.cell(k1,1).value=df4.iloc[0:df4.shape[0],0].iloc[i]
            sheet1.cell(k1+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
        else:
            sheet1.cell(k1,7,HDf1[i])
            sheet1.cell(k1+1,7).value=HDb1[i]
            sheet1.cell(k1+2,7,HDb2[i])
            sheet1.cell(k1+3,7).value=HDf2[i]
            sheet1.cell(k1,8).value='HDF'
            sheet1.cell(k1+1,8).value='HDB'
            sheet1.cell(k1+2,8).value='HDB'
            sheet1.cell(k1+3,8).value='HDF'
            sheet1.cell(k1,1).value=df4.iloc[0:df4.shape[0],0].iloc[i+1]
            sheet1.cell(k1+1,1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
        if(HD_differ_sum>6):
            print('****************************************************************************************************')
            print('前后视距累计差抄超限')
        k1=k1+5
    from ce1_8 import position_index
    position=position_index(df3)
    Z_name=[]
    Z_H=[]
    Z_HD=[]
    Z_HF=[]
    for zz1 in range(len(position)):
        HD_random3 = round(random.uniform(-0.5, 0.5), 3)  ##单位m
        Z_name.append(df3.loc[position[zz1][0],0])
        Z_HD.append(df3.loc[position[zz1][0],2]+HD_random3) ###
        Z_DH1=df3.loc[dict1[position[zz1][1]],1]-df3.loc[position[zz1][0],1]
        Z_H.append(df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-Z_DH1)
        Z_HB=RH1_random[position[zz1][1]]
        Z_HF.append(Z_HB+Z_DH1)
        if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
            print("支点的高程设置合理！！！！！！！！！！！")
        else:
            print("支点的高程设置不合理****************")
    start_num=3
    start_num_fei=0
    continue_times=0
    nn=0
    kx1=0
    flag=0
    while(flag==0):
        for zz1 in range(kx1,len(position)):
            for zz2 in range(start_num_fei,position[zz1][1]+1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                if((zz2+1)%2==1):
                    sheet2.cell(start_num+5*(zz2-start_num_fei),5).value=RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value =FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                    sheet2.cell(start_num+  5 * (zz2 - start_num_fei) , 1).value = df4.iloc[0:df4.shape[0],0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    name1 = 'A' + str(start_num + 5 *(zz2 - start_num_fei)+4 )
                    sheet1[name1].font = Font(bold=True)
                else:
                    sheet2.cell(start_num+5*(zz2-start_num_fei),5).value=FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                    sheet2.cell(start_num+  5 * (zz2 - start_num_fei) , 1).value = df4.iloc[0:df4.shape[0],0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    name1 = 'A' + str(start_num + 5 *(zz2 - start_num_fei)+4 )
                    sheet1[name1].font = Font(bold=True)
            kx2=0
            for zz3 in range(zz1,len(position)):
                if(position[zz1][1]==position[zz3][1]):
                    kx2=kx2+1
            zz3=zz1+kx2
            for zz4 in range(zz3-zz1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5+zz4, 1).value = df3.loc[position[zz1+zz4][0], 0] ##中间点的点名
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5+zz4, 5).value = Z_HF[zz1+zz4] ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 6).value = 'RZ'  ##中间点视线RZ
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 7).value = Z_HD[zz1+zz4]  ##中间点视距
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 8).value = 'HD'  ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 9).value = Z_H[zz1+zz4]  ##中间点视线高
            kx1=zz3-zz1+kx1
            start_num=start_num + 5 * (zz2 - start_num_fei) + 4 + zz3-zz1+1
            start_num_fei=position[zz1][1]+1
            if(kx1==len(position)):
                ##收尾
                flag=1
                if(position[zz1][1]==df4.shape[0]-1-1):
                    pass
                else:
                    for zz2 in range(start_num_fei, df4.shape[0]-1):
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                        if ((zz2 + 1) % 2 == 1):
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            sheet1[name1].font = Font(bold=True)
                        else:
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            sheet1[name1].font = Font(bold=True)
                break
            else:
                pass
            break

    sheet1['A2'].font=Font(bold=True)
    for i in range(df4.shape[0]-1):
        name1='A'+str(7+5*i)
        name2='A'+str(7+5*i)
        sheet1[name2]=df4.iloc[i+1,0]
        sheet1[name1].font=Font(bold=True)
    book1.save(path1)
    book1.close()
    book2.save(path2)
    book2.close()


# # path1='C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\output1.xlsx'
# path_output1 = 'D:\\Desktop\\测试期次1\\测试output\\'
# # path2='C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\output2.xlsx'
# # dateset_path='C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\测试数据库1.xlsx'
# dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'
# # measure_line_path="C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\测试线路文件1.xlsx"
# measure_line_path = 'D:\\Desktop\\测试期次1\\测试线路文件1.xlsx'
# RZ_names, RZ_values = read_measure_line_from_dataset(dateset_path, measure_line_path)
#
# # "D:\\Desktop\\测试期次1\\测试output\\"
# def mid(x, y):
#     a = x[0] + (y[0] - x[0]) / 2
#     b = x[1] + (y[1] - x[1]) / 2
#     return [a, b]
#
# # path="C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\测试线路文件1.xlsx"
# path = "D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"
# for m in range(46): #len(RZ_values)
#     path_output2=path_output1+str(m+1)+'期'+'原始数据'+'.xlsx'
#     path_output3=path_output1+str(m+1)+'期'+'原始数据'+'无支点'+'.xlsx'
#     df = pd.read_excel(measure_line_path, 'Sheet1')  # 对应的导出为pd.to_excel(***.xlsx)
#     df.columns = range(0, 3)
#     name = RZ_names
#     value = RZ_values[m]
#     dict = {name[i]: value[i] for i in range(len(name))}
#     for i in range(0, df.shape[0]):
#         if (df.loc[i, 0] in name):
#             df.loc[i, 1] = dict[df.loc[i, 0]]
#     df1 = df[:1]
#     df2 = df[1:len(df)]
#     df21 = df2.dropna(how='any')
#     df3 = pd.concat([df1, df21], axis=0)
#     df3.index = range(df3.shape[0])
#     a = df3.iloc[:, [1]]
#     b = range(df3.shape[0])
#     num = []
#     numx = []
#     for i in range(df3.shape[0]):
#         if (df3.loc[i, 0] in name):
#             numx.append(i)
#             pass
#         else:
#             num.append(i)
#
#     # d=df3.iloc[num,[1]]
#     # num2=[]
#     # for i in d.index:
#     #     num2.append([i,d.loc[i,d.columns]])
#     # # print(num2)
#     # num3=[]
#     # for i in range(len(num)-1):
#     #     num3.append(mid(num2[i],num2[i+1]))
#     # fig=pyplot.figure()
#     # ax=fig.add_subplot(111)
#     # ax.plot(numx,df3.loc[numx,1],'b^',markersize=12)
#     # ax.plot(num,df3.loc[num,1],'ro',markersize=12)
#     # ax.plot(d,color='black', linewidth=2.0)
#     # for i in range(len(num3)):
#     #     ax.plot(num3[i][0],num3[i][1],'bX',markersize=12)
#     # for i in range(len(numx)):
#     #     ax.plot([numx[i],num3[1][0]],[df3.loc[numx[i],1],num3[1][1]],color='g', linewidth=2.0)
#     # pyplot.show()
#     # # pyplot.pause(100)
#
#     ##仪器高1.4-1.6m
#     ##目标高0.55-1.8m
#     ##读数范围为如果地势平坦易设置在仪器高附近且不能超目标高范围
#     ##基辅值差0.3mm
#     ##高差之差0.6mm
#     ##闭合差60mm
#     ##改进值为60mm/6=10mm
#     ##前后视距之差为1.5m
#     ##累计视距之差为6m
#     instrument_height = 1.4
#     df4 = df3.loc[num, [0, 1, 2]]
#     dict1 = {i: num[i] for i in range(df4.shape[0])}
#     book1=openpyxl.Workbook()
#     sheet1=book1.create_sheet('原始数据-无支点',0)
#     book2=openpyxl.Workbook()
#     sheet2=book2.create_sheet('原始数据-支点',0)
#     BC1=0.0001#round(random.uniform(-0.005, 0.005),4)
#     sight_height_distance(df3,df4,book1,sheet1,book2,sheet2,BC1,path_output3,path_output2,num,dict1)












