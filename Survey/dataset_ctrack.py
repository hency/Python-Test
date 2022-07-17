import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import datetime
import re
import shutil
import os
path="D:\\2022\\主体沉降\\春天新苑\\春天新苑中间资料.xlsx"
path1="D:\\2022\\主体沉降\\春天新苑\\dataset.xlsx"
path2="D:\\2022\\主体沉降\\春天新苑\\数据库模板.xlsx"
book2=load_workbook(path)
sheet_names=book2.sheetnames
book1 = openpyxl.Workbook()
for z in range(len(sheet_names)):
    data1=pd.read_excel(path,sheet_name=sheet_names[z])
    shutil.copy(path2,os.path.abspath(os.path.join(path, ".."))+'\\'+sheet_names[z]+'数据库.xlsx')
    path3=os.path.abspath(os.path.join(path, ".."))+'\\'+sheet_names[z]+'数据库.xlsx'
    book3=load_workbook(path3)
    sheet3=book3.get_sheet_by_name('建筑沉降成果表')
    sheet4=book3.get_sheet_by_name('日报')
    num=re.findall(r'(.*)#',sheet_names[z],flags=0)[0]
    b=data1.shape[1]
    data1.columns=range(0,b)
    sheet1=book1.create_sheet(sheet_names[z])
    index1=[]
    for i in range(0,data1.shape[0]):
        if(data1.loc[i,0]=='点号'):
            index1=i+1
            pass
        else:
            pass
    index2=[]
    for i in range(index1,data1.shape[0]):
        if('CJ' not in data1.loc[i,0]):
            index2=i
            break
    for i in range(index1,index2):
        sheet1.cell(i-index1+1,1).value=data1.loc[i,1]
        sheet1.cell(i-index1+1,2).value=data1.loc[i,3]
        sheet3.cell(13,i+4-index1).value=data1.loc[i,1]
        sheet3.cell(14,i+4-index1).value=data1.loc[i,3]
        sheet3.cell(12,i+4-index1).value='Y'+num+'-'+'CJ'+str(i-index1+1)
    k=3
    if(re.findall(r'[(](.*)[)]',data1.iloc[1,0],flags=0)==[]):
        sheet1.cell(data1.shape[0], 1).value = re.findall(r'[（](.*)[）]', data1.iloc[1, 0], flags=0)[0]
        sheet3.cell(13, 2).value = re.findall(r'[（](.*)[）]', data1.iloc[1, 0], flags=0)[0]
        sheet3.cell(13,1).value=1
        sheet4.cell(2,2).value=re.findall(r'[（](.*)[）]', data1.iloc[1, 0], flags=0)[0]
        sheet4.cell(2,1).value=1
    else:
        sheet1.cell(data1.shape[0],1).value=re.findall(r'[(](.*)[)]',data1.iloc[1,0],flags=0)[0]
        sheet3.cell(13, 2).value =re.findall(r'[(](.*)[)]',data1.iloc[1,0],flags=0)[0]
        sheet3.cell(13, 1).value = 1
        sheet4.cell(2, 2).value = re.findall(r'[(](.*)[)]',data1.iloc[1,0],flags=0)[0]
        sheet4.cell(2, 1).value = 1
    if(re.findall(r'[(](.*)[)]',data1.iloc[1,2],flags=0)==[]):
        sheet1.cell(data1.shape[0], 2).value = re.findall(r'[（](.*)[）]', data1.iloc[1, 2], flags=0)[0]
        sheet3.cell(14, 2).value = re.findall(r'[（](.*)[）]', data1.iloc[1, 2], flags=0)[0]
        sheet3.cell(14, 1).value = 2
        sheet4.cell(3,2).value= re.findall(r'[（](.*)[）]', data1.iloc[1, 2], flags=0)[0]
        sheet4.cell(3, 1).value = 2
    else:
        sheet1.cell(data1.shape[0],2).value=re.findall(r'[(](.*)[)]',data1.iloc[1,2],flags=0)[0]
        sheet3.cell(14, 2).value = re.findall(r'[(](.*)[)]',data1.iloc[1,2],flags=0)[0]
        sheet3.cell(14, 1).value = 2
        sheet4.cell(3,2).value= re.findall(r'[(](.*)[)]',data1.iloc[1,2],flags=0)[0]
        sheet4.cell(3, 1).value = 2

    for i in range(8,data1.shape[1],7):
        if(re.findall(r'[(](.*)[)]', data1.iloc[1, i-3], flags=0) == []):
            sheet1.cell(data1.shape[0], k).value = re.findall(r'[（](.*)[）]', data1.iloc[1, i - 3], flags=0)[0]
            sheet3.cell(12+k, 2).value = re.findall(r'[（](.*)[）]', data1.iloc[1, i - 3], flags=0)[0]
            sheet3.cell(12 + k, 1).value = k
            sheet4.cell(1+k,2).value = re.findall(r'[（](.*)[）]', data1.iloc[1, i - 3], flags=0)[0]
            sheet4.cell(1 + k, 1).value = k
        else:
            sheet1.cell(data1.shape[0], k).value = re.findall(r'[(](.*)[)]', data1.iloc[1, i - 3], flags=0)[0]
            sheet3.cell(12+k,2).value= re.findall(r'[(](.*)[)]', data1.iloc[1, i - 3], flags=0)[0]
            sheet3.cell(12+k,1).value=k
            sheet4.cell(1 + k, 2).value = re.findall(r'[(](.*)[)]', data1.iloc[1, i - 3], flags=0)[0]
            sheet4.cell(1 + k, 1).value = k
        for j in range(index1,index2):
            sheet1.cell(j-index1+1,k).value=data1.loc[j,i]
            sheet3.cell(12+k,4+j-index1).value=data1.loc[j,i]
        k=k+1
    book3.save(path3)
book1.save(path1)
book1.close()
