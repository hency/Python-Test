import openpyxl
from openpyxl import load_workbook
import os
import re
import numpy
import random
from openpyxl.styles import Font
import datetime

def read_measure_line_from_dataset(dateset_path, measure_line_path):
    mpath = load_workbook(measure_line_path)
    msheet = mpath.get_sheet_by_name('测线')
    msheet = mpath['测线']
    dpath = load_workbook(dateset_path)
    dsheet1 = dpath.get_sheet_by_name('日报')
    sheetname_from_dateset = []
    for name in dpath.sheetnames:
        if ('沉降' in name):
            sheetname_from_dateset.append(name)
            print('沉降数据库有：%s', name)
        else:
            pass
    print('选择指定的数据库进行对应的沉降观测')
    # 由于这里只有一个测线和一个建筑物沉降成果数据库，只对其进行原始数据的还原
    dsheet2 = dpath.get_sheet_by_name(sheetname_from_dateset[0])
    dsheet3 = dpath.get_sheet_by_name('日报')
    # 先进行每一栋每一期的数据进行测试
    max_row = 0
    for i in range(0, dsheet3.max_row):
        if (dsheet3.cell(i + 1, 2).value == '' or dsheet3.cell(i + 1, 3).value is None):
            max_row = dsheet3.max_row
            break
        elif (dsheet3.cell(dsheet3.max_row, 2).value != '' and dsheet3.cell(dsheet3.max_row, 3).value is not None):
            max_row = dsheet3.max_row
    date = []
    max_col = 0
    for i in range(2, max_row + 1):
        date.append(dsheet3.cell(i, 2).value)
        print('第%d期沉降观测' % (i - 1))
    for i in range(1, dsheet2.max_row):
        if (dsheet2.cell(i, 2).value == '日期'):
            date_start = i
    for j in range(1, dsheet2.max_column):
        if (dsheet2.cell(date_start, j) == '' or dsheet2.cell(date_start, j) is None):
            max_col = j - 1
            break
        elif (dsheet2.cell(date_start, dsheet2.max_column) != '' or dsheet2.cell(date_start,
                                                                                 dsheet2.max_column) is not None):
            max_col = dsheet2.max_column
    gc_name = []
    gc_start_num = 0
    for i in range(1, max_col):
        if (dsheet2.cell(date_start, i).value == '时间'):
            gc_start_num = i + 1
            for j in range(0, max_col - gc_start_num + 1):
                gc_name.append(dsheet2.cell(date_start, gc_start_num + j).value)
            break
    cx_make = []
    for i in range(1, len(date) + 1):
        date1 = date_start + i
        print(
            '***********************************%s*******************************' % (dsheet2.cell(date1, 2).value))
        print('提取侧线文件及数据库文件，进行相应复制')
        cx_make1 = []
        for j in range(0, max_col - gc_start_num + 1):
            cx_make1.append(dsheet2.cell(date1, j + 4).value)
        cx_make.append(cx_make1)
    return gc_name, cx_make

def position_index(self,df3):
    position = []
    num_1 = 0
    for i in range(df3.shape[0]):
        if 'Y' in df3.iloc[i, 0]:
            position.append([i, i - num_1 - 1])
            num_1 = num_1 + 1
    return position

def position_index1(self,df3):
    num = []
    position1 = self.position_index(df3)
    flag = 0
    k1 = 0
    k3 = 0
    while (flag == 0):
        for q1 in range(k1, len(position1)):
            k = 0
            k2 = position1[q1][1]
            for q2 in range(len(position1)):
                if (position1[q1][1] == position1[q2][1]):
                    k = k + 1
                    k1 = k1 + 1
                else:
                    pass
            k3 = k3 + k
            num.append([k2, k, k3])
            break
        if (k1 == len(position1)):
            flag = 1
    return num

def Z_H_function(self,df3, df4, num, BC1):
    dict1 = {i: num[i] for i in range(df4.shape[0])}
    position = self.position_index(df3)
    Z_name = []
    Z_H = []
    Z_HD = []
    Z_HF = []
    for zz1 in range(len(position)):
        Z_DH1 = df3.loc[dict1[position[zz1][1]], 1] - df3.loc[position[zz1][0], 1]
        Z_H.append(df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - Z_DH1)
    return Z_H

def sight_height_distance(self,df3, df4, book2, sheet2, BC1, path1, path2, num, dict1):
    # sheet1.cell(1, 1).value = '点位'
    # sheet1.cell(1, 3).value = '时间'
    # sheet1.cell(1, 5).value = '视线高'
    # sheet1.cell(1, 7).value = '视距'
    # sheet1.cell(1, 9).value = '高程'
    # sheet1.cell(2, 1).value = df3.iloc[0, 0]
    # sheet1.cell(2, 9).value = df4.iloc[0, 1]
    sheet2.cell(1, 1).value = '点位'
    sheet2.cell(1, 3).value = '时间'
    sheet2.cell(1, 5).value = '视线高'
    sheet2.cell(1, 7).value = '视距'
    sheet2.cell(1, 9).value = '高程'
    sheet2.cell(2, 1).value = df3.iloc[0, 0]
    sheet2.cell(2, 9).value = df4.iloc[0, 1]
    k = 3
    RH1_random = []
    RH2_random = []
    FH1 = []
    FH2 = []
    height1 = []
    position1 = self.position_index(df3)
    position2 = self.position_index1(df3)
    # BC1=BC2/(df4.shape[0]-1)
    Z_H1 = self.Z_H_function(df3, df4, num, BC1)
    for i in range(df4.shape[0] - 1):
        dh = df4.iloc[i + 1, 1] - df4.iloc[i, 1] - BC1
        if (dh > 1.13):
            print("高差大于1.13m，高差过大")  # 因为dh1_random的值不能超过1.15m
            xx1 = 1 / 0
        height1.append(df4.iloc[i + 1, 1] - (i + 1) * BC1)
        # sheet1.cell(7 + 5 * i, 9).value = height1[i]
        # '高差值最大是1.8-0.55=1.25m' 且 dh_random的取值范围为0.55+dh~1.8
        Sight_Height_random1 = random.randint(-5, 15) * 0.00001  ##单位m
        Sight_Height_random2 = random.randint(-5, 15) * 0.00001  ##单位m0.15mm*2=0.3m
        dh1_random = dh + Sight_Height_random1  ##FH2-FH1=2*Sight_Height_random1+Sight_Height_random2
        dh2 = 2 * dh - dh1_random
        flagxx1 = 0
        for q in range(len(position2)):
            if (i == position2[q][0]):
                #############加入
                pass
                flagxx1 = 1
                if (dh1_random > 0):
                    RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu = 0
                    for p in range(position2[q][1]):
                        if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006) and (
                                RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                            position2[q][2] - position2[q][1] + p] and Z_H1[
                                    position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                    dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu = can_shu + 1
                    if (can_shu == position2[q][1]):
                        RH1_random.append(
                            RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1 = 0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                        FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                        RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                    position2[q][2] - position2[q][1] + p] and Z_H1[
                                            position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                            dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1 = can_shu1 + 1
                            if (can_shu1 == position2[q][1]):
                                RH1_random.append(
                                    RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                elif (dh1_random < 0):
                    #############加入
                    RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu = 0
                    for p in range(position2[q][1]):
                        if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006) and (
                                RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                            position2[q][2] - position2[q][1] + p] and Z_H1[
                                    position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                    dict1[i], 1] - i * BC1 - 0.6)):
                            # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu = can_shu + 1
                    if (can_shu == position2[q][1]):
                        RH1_random.append(
                            RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1 = 0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                        FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                        RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                    position2[q][2] - position2[q][1] + p] and Z_H1[
                                            position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                            dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1 = can_shu1 + 1
                            if (can_shu1 == position2[q][1]):
                                RH1_random.append(
                                    RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                elif (dh1_random == 0):
                    RH1_random_1 = round(random.uniform(0.6, 1.75),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu = 0
                    for p in range(position2[q][1]):
                        if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                                dh1_random - dh2) < 0.0006) and (
                                RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                            position2[q][2] - position2[q][1] + p] and Z_H1[
                                    position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                    dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu = can_shu + 1
                    if (can_shu == position2[q][1]):
                        RH1_random.append(
                            RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1 = 0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75),
                                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(
                                        FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (
                                        RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[
                                    position2[q][2] - position2[q][1] + p] and Z_H1[
                                            position2[q][2] - position2[q][1] + p] < RH1_random_1 + df3.loc[
                                            dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1 = can_shu1 + 1
                            if (can_shu1 == position2[q][1]):
                                RH1_random.append(
                                    RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
        if (flagxx1 == 1):
            continue
        if (dh1_random > 0):
            RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1 = RH1_random_1 - dh1_random
            RH2_random_1 = RH1_random_1 + Sight_Height_random2
            FH2_1 = RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                    dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                            dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
        elif (dh1_random < 0):
            RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1 = RH1_random_1 - dh1_random
            RH2_random_1 = RH1_random_1 + Sight_Height_random2
            FH2_1 = RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                    dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                            dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
        elif (dh1_random == 0):
            RH1_random_1 = round(random.uniform(0.6, 1.75),
                                 5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1 = RH1_random_1 - dh1_random
            RH2_random_1 = RH1_random_1 + Sight_Height_random2
            FH2_1 = RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                    dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1 = round(random.uniform(0.6, 1.75),
                                         5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(
                            dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
    HD_differ_sum = 0
    k1 = 3
    HDf1 = []
    HDf2 = []
    HDb1 = []
    HDb2 = []
    for i in range(df4.shape[0] - 1):
        HD_random1 = round(random.uniform(-0.2, 0.2), 3)  ##单位m
        HD_random2 = round(random.uniform(-0.2, 0.2), 3)
        HD = df4.iloc[i + 1, 2] / 2
        HDb1.append(HD + HD_random1)
        HDb2.append(HDb1[i] + round(random.uniform(-0.005, 0.005), 3))
        HDf1.append(HD + HD_random2)
        HDf2.append(HDf1[i] + round(random.uniform(-0.005, 0.005), 3))
        HDb = (HDb1[i] + HDb2[i]) / 2
        HDf = (HDf1[i] + HDf2[i]) / 2
        HD_differ = HDb - HDf
        HD_differ_sum = HD_differ_sum + HD_differ
        if (abs(HD_differ) < 1.5 and abs(HD_differ_sum) < 6):
            print('前后视距满足要求')
        else:
            print('前后视距不满足要求')
            print('出现异常将在GUI中提现')
            Exception1 = 1 / 0
        k1 = k1 + 5
    position = self.position_index(df3)
    Z_name = []
    Z_H = []
    Z_HD = []
    Z_HF = []
    for zz1 in range(len(position)):
        HD_random3 = round(random.uniform(-0.5, 0.5), 3)  ##单位m
        Z_name.append(df3.loc[position[zz1][0], 0])
        Z_HD.append(df3.loc[position[zz1][0], 2] + HD_random3)  ###
        Z_DH1 = df3.loc[dict1[position[zz1][1]], 1] - df3.loc[position[zz1][0], 1]
        Z_H.append(df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - Z_DH1)
        Z_HB = RH1_random[position[zz1][1]]
        Z_HF.append(Z_HB + Z_DH1)
        if (Z_HB + df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - 1.75 < Z_H[zz1] and Z_H[
            zz1] < Z_HB + df3.loc[dict1[position[zz1][1]], 1] - position[zz1][1] * BC1 - 0.6):
            print("支点的高程设置合理！！！！！！！！！！！")
        else:
            print("支点的高程设置不合理****************")
    start_num = 3
    start_num_fei = 0
    continue_times = 0
    nn = 0
    kx1 = 0
    flag = 0
    while (flag == 0):
        for zz1 in range(kx1, len(position)):
            for zz2 in range(start_num_fei, position[zz1][1] + 1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                if ((zz2 + 1) % 2 == 1):
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = df4.iloc[0:df4.shape[0], 0].iloc[
                        zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                    name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                    # sheet1[name1].font = Font(bold=True)
                else:
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = df4.iloc[0:df4.shape[0], 0].iloc[
                        zz2 + 1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                        df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                    name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                    # sheet1[name1].font = Font(bold=True)
            kx2 = 0
            for zz3 in range(zz1, len(position)):
                if (position[zz1][1] == position[zz3][1]):
                    kx2 = kx2 + 1
            zz3 = zz1 + kx2
            for zz4 in range(zz3 - zz1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 1).value = df3.loc[
                    position[zz1 + zz4][0], 0]  ##中间点的点名
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 5).value = Z_HF[zz1 + zz4]  ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 6).value = 'RZ'  ##中间点视线RZ
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 7).value = Z_HD[zz1 + zz4]  ##中间点视距
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 8).value = 'HD'  ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 9).value = Z_H[zz1 + zz4]  ##中间点视线高
            kx1 = zz3 - zz1 + kx1
            start_num = start_num + 5 * (zz2 - start_num_fei) + 4 + zz3 - zz1 + 1
            start_num_fei = position[zz1][1] + 1
            if (kx1 == len(position)):
                ##收尾
                flag = 1
                if (position[zz1][1] == df4.shape[0] - 1 - 1):
                    pass
                else:
                    for zz2 in range(start_num_fei, df4.shape[0] - 1):
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                        if ((zz2 + 1) % 2 == 1):
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            # sheet1[name1].font = Font(bold=True)
                        else:
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                                df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            # sheet1[name1].font = Font(bold=True)
                break
            else:
                pass
            break
    book2.save(path2)
    book2.close()

def baoliu(self,a):
    ling1 = '0'
    ling2 = '00'
    ling3 = '000'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数4为 字符串
        ag = a + '0000'
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
        ag = a + ling3
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
        ag = a + ling2
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
        ag = a + ling1
    else:
        ag = a
    return ag

def baoliu3(self,a):
    ling1 = '0'
    ling2 = '00'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数3为 字符串
        ag = a + '000'
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
        ag = a + ling2
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
        ag = a + ling1
    else:
        ag = a
    return ag

def baoliu5(self,a):
    ling1 = '0'
    ling2 = '00'
    ling3 = '000'
    ling4 = '0000'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数5为 字符串
        ag = a + '00000'
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
        ag = a + ling4
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
        ag = a + ling3
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
        ag = a + ling2
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 4):
        ag = a + ling1
    else:
        ag = a
    return ag
def function_cj(self):
    # measure_line_path = "D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"#D:\Desktop\测试期次1\测试线路文件1.xlsx##self.editline1.text()
    # dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'#D:\Desktop\测试期次1\测试数据库1.xlsx'#self.editline2.text()#self.editline2.text()
    # path_output1 = 'D:\\Desktop\\测试期次1\\测试output\\'#self.editline3.text()#'D:\Desktop\测试期次1\测试output\self.editline3.text()
    measure_line_path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
    dateset_path =  QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')[0]
    dateset_qishu=re.findall(r'(.*)#',os.path.basename(dateset_path),flags=0)[0]
    src_path=os.path.abspath(os.path.join(measure_line_path, ".."))+'\\' #############################################################
    output_name = "报表"  ###输出的文件夹名字
    if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
        pass
        print(output_name + "已经存在！")
    else:
        print("创建文件夹" + output_name + "!")
        os.makedirs(src_path + output_name)
    path_output1 = src_path + output_name
    qi_shu = self.editline2.text()
    BC_range1 = self.editline1.text()
    BC_range_min = re.findall(r'(.*) ', BC_range1,flags=0)[0]
    BC_range_max = re.findall(r' (.*)', BC_range1,flags=0)[0]
    print(measure_line_path)
    print(dateset_path)
    print(path_output1)
    print(qi_shu)
    print(BC_range_min, BC_range_max)
    RZ_names, RZ_values = self.read_measure_line_from_dataset(dateset_path, measure_line_path)
    for m in range(int(qi_shu)):  # len(RZ_values)
        BC1 = round(random.uniform(float(BC_range_min), float(BC_range_max)), 5)
        path_output2 = path_output1 +'\\'+ str(m + 1) + '期' + '原始数据' + '.xlsx'
        path_output3 = path_output1 + '\\'+str(m + 1) + '期' + '原始数据' + '无支点' + '.xlsx'
        df = pd.read_excel(measure_line_path, '测线')  # 对应的导出为pd.to_excel(***.xlsx)
        df.columns = range(0, 3)
        name = RZ_names
        value = RZ_values[m]
        dict = {name[i]: value[i] for i in range(len(name))}
        for i in range(0, df.shape[0]):
            if (df.loc[i, 0] in name):
                df.loc[i, 1] = dict[df.loc[i, 0]]
        df1 = df[:1]
        df2 = df[1:len(df)]
        df21 = df2.dropna(how='any')
        df3 = pd.concat([df1, df21], axis=0)
        df3.index = range(df3.shape[0])
        a = df3.iloc[:, [1]]
        b = range(df3.shape[0])
        num = []
        numx = []
        for i in range(df3.shape[0]):
            if (df3.loc[i, 0] in name):
                numx.append(i)
                pass
            else:
                num.append(i)
        instrument_height = 1.4
        df4 = df3.loc[num, [0, 1, 2]]
        dict1 = {i: num[i] for i in range(df4.shape[0])}
        # book1 = openpyxl.Workbook()
        # sheet1 = book1.create_sheet('原始数据-无支点', 0)
        book2 = openpyxl.Workbook()
        sheet2 = book2.create_sheet('原始数据-支点', 0)
        BC2 = BC1 / (df4.shape[0] - 1)
        self.sight_height_distance(df3, df4, book2, sheet2, BC2, path_output3, path_output2, num,
                                   dict1)
    pat1 = path_output1+'\\'
    pat_name = os.listdir(pat1)
    SUM_RB = []
    SUM_RF = []
    SUM_RB_CEZHAN = []
    DATE = []
    INDEX3 = []
    for i in range(len(pat_name)):
        if (pat_name[i][-4:] == 'xlsx'):
            if (re.findall(r'数据(.*)', pat_name[i])[0] != '无支点.xlsx'):
                index3 = re.findall(r'(.*)期', pat_name[i])[0]
                INDEX3.append(index3)
                book2 = load_workbook(dateset_path)
                sheet2 = book2.get_sheet_by_name('建筑沉降成果表')
                nrows2 = sheet2.max_row
                for z3 in range(13, nrows2 + 1):
                    if (int(index3) == sheet2.cell(z3, 1).value):
                        DATE.append(sheet2.cell(z3, 2).value)
                        break
                book1 = load_workbook(pat1 + pat_name[i])
                sheet1 = book1.get_sheet_by_name('原始数据-支点')
                nrows = sheet1.max_row
                ncols = sheet1.max_column
                sum_RB = 0
                sum_RF = 0
                sum_RB_cezhan = 0
                for j in range(3, nrows + 1):
                    time_round=random.randint(10,20)
                    sheet1.cell(j, 3).value = datetime.timedelta(
                        seconds=28800 + (j - 3) * time_round)  # print(datetime.timedelta(seconds=28800))
                    if (sheet1.cell(j, 6).value == 'RB'):
                        sum_RB = sum_RB + sheet1.cell(j, 7).value
                        sum_RB_cezhan = sum_RB_cezhan + 1
                    if (sheet1.cell(j, 6).value == 'RF'):
                        sum_RF = sum_RF + sheet1.cell(j, 7).value
                book1.save(pat1 + pat_name[i])
                SUM_RB.append(sum_RB)
                SUM_RF.append(sum_RF)
                SUM_RB_CEZHAN.append(sum_RB_cezhan)

    num_1 = 0
    for z1 in range(len(pat_name)):
        # path="D:\\Desktop\\测试期次1\\测试output\\10期原始数据1.xlsx"
        path1 = path_output1+'\\'
        if (pat_name[z1][-4:] == 'xlsx'):
            if (re.findall(r'数据(.*)', pat_name[z1])[0] != '无支点.xlsx'):
                num_1 = num_1 + 1
                path = path1 + pat_name[z1]
                data1 = pd.read_excel(path, '原始数据-支点')
                # fid1 = open(path1 + '沉降观测数据第' + INDEX3[num_1 - 1] + '期' + '.txt', 'w')
                fid1 = open(path1 + dateset_qishu+'#沉降' + INDEX3[num_1 - 1] + '.dat', 'w')
                date3 = str(DATE[num_1 - 1])
                x1 = date3.replace('-', '年', 1)
                x2 = x1.replace('-', '月', 1)
                x3 = x2.replace(' 00:00:00', '日', 1)
                year1 = re.findall(r'(.*)年', x3, flags=0)[0]
                month1 = re.findall(r'年(.*)月', x3, flags=0)[0]
                day1 = re.findall(r'月(.*)日', x3, flags=0)[0]
                if (len(month1) == 1):
                    month1 = '0' + month1
                if (len(day1) == 1):
                    day1 = '0' + day1
                date4 = year1 + month1 + day1
                fid1.write('期次：' + INDEX3[num_1 - 1] + ' ' + '日期：' + x3 + '\n')
                fid1.write(
                    'For M5|Adr' + '   ' + '  1' + '|' + 'TO' + '  ' + date4 + '.dat' + '               ' + "|" + "                      " + "|" + "                      " + "|" + "                      " + "|" + "\n")
                fid1.write(
                    'For M5|Adr' + '   ' + '  2' + '|' + 'TO' + '  ' + 'Start-Line' + '      ' + 'aBFFB' + '     ' + '1' + '|' + "                      " + "|" + "                      " + "|" + "                      " + "|" + "\n")
                rown = data1.shape[0]
                coln = data1.shape[1]
                i = 0
                for i in range(0, rown):
                    dm = data1.iloc[i, 0]
                    if (i + 2 < 10):
                        index1 = '  ' + str(i + 2)
                    elif ((i + 2 > 10 and i + 2 < 100) or i + 2 == 10):
                        index1 = ' ' + str(i + 2)
                    else:
                        index1 = str(i + 2)
                    kn = ''
                    for j in range(9 - len(dm)):
                        kn = kn + ' '
                    dmlength = kn + dm
                    if (i == 0):
                        fid1.write(
                            'For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '                  ' + '1' + '|' + "                      " + '|' + "                      " + '|' + 'Z' + '        ' + self.baoliu(
                                str(data1.iloc[i, coln - 1])) + ' m' + '    ' + '|' + '\n')
                    else:
                        if ((~numpy.isnan(data1.iloc[i, 4])) and (~numpy.isnan(data1.iloc[i, coln - 1]))):
                            ##支点
                            if (len(self.baoliu3(str(data1.iloc[i, 6]))) == 6):
                                fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                    data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, 4],
                                              5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '         ' + self.baoliu3(
                                    str(round(data1.iloc[i, 6], 3))) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, coln - 1], 5))) + ' m' + '   ' + '|' + '\n')
                            else:
                                fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                    data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, 4],
                                              5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '          ' + self.baoliu3(
                                    str(round(data1.iloc[i, 6], 3))) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, coln - 1], 5))) + ' m' + '   ' + '|' + '\n')

                        elif ((~numpy.isnan(data1.iloc[i, 4])) and (numpy.isnan(data1.iloc[i, coln - 1]))):
                            ##后视点或者前视点
                            if (len(self.baoliu3(str(data1.iloc[i, 6]))) == 6):
                                fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                    data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, 4],
                                              5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '         ' + self.baoliu3(
                                    str(round(data1.iloc[i, 6],
                                              3))) + ' m' + '   ' + '|' + '                      ' + '|' + '\n')
                            else:
                                fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                    data1.iloc[i, 2]) + '    ' + '1' + '|' + data1.iloc[i, 5] + '        ' + self.baoliu5(
                                    str(round(data1.iloc[i, 4],
                                              5))) + ' ' + 'm' + '   ' + '|' + 'HD' + '          ' + self.baoliu3(
                                    str(round(data1.iloc[i, 6],
                                              3))) + ' m' + '   ' + '|' + '                      ' + '|' + '\n')
                        else:
                            fid1.write('For M5|Adr' + '   ' + index1 + '|' + 'KD1' + dmlength + '      ' + str(
                                data1.iloc[
                                    i, 2]) + '    ' + '1' + '|' + '                      ' + '|' + '                      ' + '|' + 'Z' + '        ' + self.baoliu5(str(round(data1.iloc[i, coln - 1],5))) + ' m' + '   ' + '|' + '\n')

                index2 = i + 2 + 1
                index3 = []
                for z2 in range(3):
                    if (index2 + z2 < 10):
                        index3.append('  ' + str(index2 + z2))
                    elif ((index2 + z2 > 10 and index2 + z2 < 100) or index2 + z2 == 10):
                        index3.append(' ' + str(index2 + z2))
                    else:
                        index3.append(str(index2 + z2))
                dm1 = data1.iloc[i, 0]
                kn = ''
                for j in range(9 - len(dm1)):
                    kn = kn + ' '
                dmlength1 = kn + dm1
                if (round(data1.iloc[i, 8] - data1.iloc[0, 8], 5) < 0):
                    fid1.write('For M5|Adr' + '   ' + index3[
                        0] + '|' + 'KD1' + dmlength1 + '                  ' + '1' + '|' + 'Sh' + '       ' + self.baoliu5(
                        '%.5f' % (data1.iloc[i, 8] - data1.iloc[
                            0, 8])) + ' m' + '   ' + '|' + '                      |                      |' + '\n')
                else:
                    fid1.write('For M5|Adr' + '   ' + index3[
                        0] + '|' + 'KD1' + dmlength1 + '                  ' + '1' + '|' + 'Sh' + '        ' + self.baoliu5(
                        '%.5f' % (data1.iloc[i, 8] - data1.iloc[
                            0, 8])) + ' m' + '   ' + '|' + '                      |                      |' + '\n')
                czs = int(SUM_RB_CEZHAN[num_1 - 1] / 2)
                kn = ''
                for j in range(4 - len(str(czs))):
                    kn = kn + ' '
                czs1 = str(czs) + kn

                dbz = SUM_RB[num_1 - 1] / 2
                dfz = SUM_RF[num_1 - 1] / 2
                if (len(re.findall(r'(.*)[.]', str(round(dbz, 3)))[0]) == 3):
                    fid1.write('For M5|Adr' + '   ' + index3[
                        1] + '|' + 'KD1' + dmlength1 + '      ' + czs1 + '        ' + '1' + '|' + 'Db' + '        ' + self.baoliu3(
                        str(round(dbz, 3))) + ' m' + '   ' + '|' + 'Df' + '        ' + self.baoliu3(
                        str(round(dfz, 3)) + ' m' + '   ' + '|' + 'Z' + '        ' + self.baoliu(
                            str(data1.iloc[i, coln - 1])) + ' m' + '   ' + '|' + '\n'))
                else:
                    fid1.write('For M5|Adr' + '   ' + index3[
                        1] + '|' + 'KD1' + dmlength1 + '      ' + czs1 + '        ' + '1' + '|' + 'Db' + '        ' + self.baoliu3(
                        str(round(dbz, 3))) + '  m' + '   ' + '|' + 'Df' + '        ' + self.baoliu3(
                        str(round(dfz, 3)) + '  m' + '   ' + '|' + 'Z' + '        ' + self.baoliu(
                            str(data1.iloc[i, coln - 1])) + ' m' + '   ' + '|' + '\n'))
                fid1.write('For M5|Adr' + '   ' + index3[
                    1] + '|' + 'TO' + '  ' + 'End-Line' + '                  ' + '1' + '|' + '                      |                      |                      |')
                fid1.close()
    tname=os.listdir(src_path + output_name+'\\')
    for i in range(len(tname)):
        if('xlsx' in tname[i]):
            os.remove(src_path + output_name+'\\'+tname[i])