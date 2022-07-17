import numpy
import pandas as pd
import numpy as np
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from 沉降观测1 import read_measure_line_from_dataset
path1='D:\\Desktop\\测试期次1\\output1.xlsx'
def mid(x,y):
    a=x[0]+(y[0]-x[0])/2
    b=x[1]+(y[1]-x[1])/2
    return [a,b]
path="D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"
df=pd.read_excel(path,'Sheet1')#对应的导出为pd.to_excel(***.xlsx)
df.columns=range(0,3)
name=['Y1-a1','Y1-a2','Y1-a3']
value=[19.6425,19.7215,19.7715]
dict = {name[i]:value[i] for i in range(3)}
for i in range(0,df.shape[0]):
    if(df.loc[i,0] in name):
        df.loc[i,1]=dict[df.loc[i,0]]
df1=df[:1]
df2=df[1:len(df)]
df21=df2.dropna(how='any')
df3=pd.concat([df1,df21],axis=0)
df3.index=range(df3.shape[0])
def position_index(df3):
    position = []
    for i in range(df3.shape[0]):
        if 'Y' in df3.iloc[i,0]:
            for j in range(0,i):
                if('Y' not in df3.iloc[i-1-j,0]):
                    position.append([i,i-1-j])
                    break
    return position
position=position_index(df3)
BC1=0.010
numx=[2,3]
num1=[0,1,4,5]
dh1=df3.loc[position[0][1],1]-df3.loc[position[0][0],1]
Hz=df3.loc[1,1]-1*BC1-dh1
SHZ=SHB+dh1
SJ=df3.loc[position[0][0],2]
Z_name=df3.loc[position[0][0],0]

sheet1.cell(7 + 5 * i, 9).value = height1
if ((i + 1) % 2 == 1):
    sheet1.cell(k, 5, RH1_random)
    sheet1.cell(k + 1, 5, FH1)
    sheet1.cell(k + 2, 5, FH2)
    sheet1.cell(k + 3, 5, RH2_random)
    sheet1.cell(k, 6).value = 'RB'
    sheet1.cell(k + 1, 6).value = 'RF'
    sheet1.cell(k + 2, 6).value = 'RF'
    sheet1.cell(k + 3, 6).value = 'RB'
    sheet1.cell(k1, 7, HDb1)
    sheet1.cell(k1 + 1, 7).value = HDf1
    sheet1.cell(k1 + 2, 7, HDf2)
    sheet1.cell(k1 + 3, 7).value = HDb2
    sheet1.cell(k1, 8).value = 'HDB'
    sheet1.cell(k1 + 1, 8).value = 'HDF'
    sheet1.cell(k1 + 2, 8).value = 'HDF'
    sheet1.cell(k1 + 3, 8).value = 'HDB'
    sheet1.cell(k1, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
    sheet1.cell(k1 + 1, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
    sheet1.cell(k1 + 2, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
    sheet1.cell(k1 + 3, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
else:
    sheet1.cell(k, 5, RH1_random)
    sheet1.cell(k + 1, 5, FH1)
    sheet1.cell(k + 2, 5, FH2)
    sheet1.cell(k + 3, 5, RH2_random)
    sheet1.cell(k, 6).value = 'RF'
    sheet1.cell(k + 1, 6).value = 'RB'
    sheet1.cell(k + 2, 6).value = 'RB'
    sheet1.cell(k + 3, 6).value = 'RF'
    sheet1.cell(k1, 7, HDf1)
    sheet1.cell(k1 + 1, 7).value = HDb1
    sheet1.cell(k1 + 2, 7, HDb2)
    sheet1.cell(k1 + 3, 7).value = HDf2
    sheet1.cell(k1, 8).value = 'HDF'
    sheet1.cell(k1 + 1, 8).value = 'HDB'
    sheet1.cell(k1 + 2, 8).value = 'HDB'
    sheet1.cell(k1 + 3, 8).value = 'HDF'
    sheet1.cell(k1, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
    sheet1.cell(k1 + 1, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
    sheet1.cell(k1 + 2, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
    sheet1.cell(k1 + 3, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]