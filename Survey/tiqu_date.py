import pandas
import re
import openpyxl

path="D:\\2022\\吴昌程生成原始数据\\抚州保利S3地块数据\\抚州保利数据库.xlsx"

DATA=pandas.read_excel(path,'日报')
book1=openpyxl.Workbook()
sheet1=book1.create_sheet('日报')
for i in range(DATA.shape[0]):
    sheet1.cell(i+1,1).value=DATA.iloc[i,1]
book1.save("D:\\2022\\吴昌程生成原始数据\\抚州保利S3地块数据\\tiqu.xlsx")