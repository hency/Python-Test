#coding=utf-8
import sys
from PyQt5.QtWidgets import QWidget,QPushButton,QVBoxLayout,QHBoxLayout,QGridLayout,QFormLayout,QLineEdit,QLabel,QMessageBox,QApplication,\
    QDialog,QAction,QDateTimeEdit,QFileDialog
from PyQt5.QtCore import QDate, QTime, QDateTime
import numpy
import pandas as pd
import numpy as np
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import os
import shutil
from openpyxl import load_workbook
import math
class zhu_ti_cj_jiang(QWidget):
    def __init__(self):
        super(zhu_ti_cj_jiang,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降')
        self.label1=QLabel(self)
        self.label1.setText('位移')
        self.pushbutton1=QPushButton(self)
        self.pushbutton1.setText('位移原始数据')
        self.label2=QLabel(self)
        self.label2.setText('沉降')
        self.pushbutton2=QPushButton(self)
        self.pushbutton2.setText('沉降原始数据')
        self.label3=QLabel(self)
        self.label3.setText('原始数据文件夹')
        self.pushbutton3=QPushButton(self)
        self.pushbutton3.setText('创建文件夹')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout_init()
    def Layout_init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout2.addWidget(self.pushbutton1)
        self.V_layout2.addWidget(self.pushbutton2)
        self.V_layout2.addWidget(self.pushbutton3)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.setLayout(self.H_layout1)
        self.pushbutton1.clicked.connect(self.enter_pushbutton1)
        self.pushbutton2.clicked.connect(self.enter_pushbutton2)
        self.pushbutton3.clicked.connect(self.enter_pushbutton3)
    def enter_pushbutton1(self):
        demo1=original_wy()
        demo1.exec_()

    def enter_pushbutton2(self):
        demo1 = original_cj()
        demo1.exec_()

    def enter_pushbutton3(self):
        demo1 = make_dir()
        demo1.exec_()
class original_wy(QDialog):
    def __init__(self):
        super(original_wy,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('水平位移原始数据还原')
        self.label1=QLabel('输入期次范围',self)
        self.editline1=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout2.addWidget(self.editline1)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('请输入期数')
        self.button1.clicked.connect(self.output_excel)
        # self.button2.clicked.connect(self.open_excel_file)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def output_excel(self):
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        src_path=os.path.abspath(os.path.join(openfile_name[0], ".."))+'\\' #############################################################
        dateset_file_path=os.path.abspath(openfile_name[0])
        # src_path = self.editline1.text()  # src_path="D:\\Desktop\\水平位移原始数据还原\\"
        # src_data_name = self.editline2.text()  # src_data_name="4-2数据库.xlsx" ###这里的xlsx文件必须再src_path文件下
        Cnumber = self.editline1.text()   # Cnumber='2-50' ###表示输入期次范围
        # src_path="D:\\Desktop\\水平位移原始数据还原\\"
        # src_data_name="4-2数据库.xlsx" ###这里的xlsx文件必须再src_path文件下
        output_name = "报表"  ###输出的文件夹名字
        if (os.path.isdir(src_path + output_name)):  ###判断输出文件架是否存在，如果存在不创建
            pass
            print(output_name + "已经存在！")
        else:
            print("创建文件夹" + output_name + "!")
            os.makedirs(src_path + output_name)
        book1 = load_workbook(dateset_file_path)
        sheet_names = book1.sheetnames
        wy_name = "坡顶水平位移"
        cj_name = "坡顶沉降"
        wy_name1 = "桩顶水平位移"
        cj_name1 = "桩顶沉降"
        for i in range(len(sheet_names)):
            if (wy_name in sheet_names[i] or wy_name1 in sheet_names[i]):
                wy_sheet = sheet_names[i]
            if (cj_name in sheet_names[i] or cj_name1 in sheet_names[i]):
                cj_sheet = sheet_names[i]
        wy_data = pd.read_excel(dateset_file_path, wy_sheet)
        cj_data = pd.read_excel(dateset_file_path, cj_sheet)
        # Cnumber='2-50' ###表示输入期次范围
        Cnumber1 = int(re.findall(r'(.*)-', Cnumber, flags=0)[0])
        Cnumber2 = int(re.findall(r'-(.*)', Cnumber, flags=0)[0])
        Cdate_range = ''  ###表示日期的范围
        Cznumber = cj_data.shape[0] - 11  ###总的期次 一次性生成完
        Cdname = []
        for i in range(3, cj_data.shape[1]):
            Cdname.append(cj_data.iloc[10, i])
        for i in range(10 + Cnumber1, 10 + Cnumber2 + 1):
            ####创建文件，文件的格式为dat文件或者txt文件
            ###判断文件是否存在
            date1_name = str(i - 10) + '观测记录' + str(cj_data.iloc[i, 1]).replace(" 00:00:00",
                                                                                "") + '坐标.txt'  ###数据库中的日期单元格格式需要保持日期的格式 年月日 ///
            fid = open(src_path + output_name + "\\" + date1_name, 'w')
            fid.write("坡顶水平位移" + str(cj_data.iloc[i, 1]).replace(" 00:00:00", "") + ':' + "\n")
            WY1 = []
            WY2 = []
            WY3 = []
            for j in range(0, cj_data.shape[1] - 3):
                bx = wy_data.iloc[i, 3 + j * 2]
                by = wy_data.iloc[i, 3 + j * 2 + 1]
                if (numpy.isnan(bx)):
                    continue
                wy_x_random1 = round(bx + random.randint(-5, 15) * 0.0001, 4)  ##这里对X方向均值及Y方向的均值进行取值计算
                wy_x_random2 = round(bx + random.randint(-5, 15) * 0.0001, 4)
                wy_x_random3 = round(3 * bx - wy_x_random1 - wy_x_random2, 4)
                wy_y_random1 = round(by + random.randint(-5, 15) * 0.0001, 4)
                wy_y_random2 = round(by + random.randint(-5, 15) * 0.0001, 4)
                wy_y_random3 = round(3 * by - wy_y_random1 - wy_y_random2, 4)
                WY1.append([wy_x_random1, wy_y_random1])
                WY2.append([wy_x_random2, wy_y_random2])
                WY3.append([wy_x_random3, wy_y_random3])
            k = 0
            for j in range(3, cj_data.shape[1]):
                a = cj_data.iloc[i, j]
                if (numpy.isnan(a)):
                    continue
                Cdname1 = Cdname[j - 3]
                name1 = Cdname1 + '-1'
                name2 = Cdname1 + '-2'
                name3 = Cdname1 + '-3'
                cj_random1 = round(a + random.randint(-5, 15) * 0.0001, 4)  ###这里只对沉降数据进行了一次随机数的选取
                cj_random2 = round(a + random.randint(-5, 15) * 0.0001, 4)
                cj_random3 = round(a + random.randint(-5, 15) * 0.0001, 4)
                a = str(WY1[k][0])
                len_num = 100
                ling1 = '0'
                ling2 = '00'
                ling3 = '000'
                if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数4为 字符串
                    ag = a + '0000'
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
                    ag = a + ling3
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 2):
                    ag = a + ling2
                elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 3):
                    ag = a + ling1
                else:
                    ag = a
                b = str(WY1[k][1])
                if (re.findall('[.](.*)', b, flags=0) == []):
                    bg = b + '0000'
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 1):
                    bg = b + ling3
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 2):
                    bg = b + ling2
                elif (len(re.findall('[.](.*)', b, flags=0)[0]) == 3):
                    bg = b + ling1
                else:
                    bg = b
                c = str(cj_random1)
                if (re.findall('[.](.*)', c, flags=0) == []):
                    cg = c + '0000'
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 1):
                    cg = c + ling3
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 2):
                    cg = c + ling2
                elif (len(re.findall('[.](.*)', c, flags=0)[0]) == 3):
                    cg = c + ling1
                else:
                    cg = c
                a1 = str(WY2[k][0])
                if (re.findall('[.](.*)', a1, flags=0) == []):
                    ag1 = a1 + '0000'
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 1):
                    ag1 = a1 + ling3
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 2):
                    ag1 = a1 + ling2
                elif (len(re.findall('[.](.*)', a1, flags=0)[0]) == 3):
                    ag1 = a1 + ling1
                else:
                    ag1 = a1
                b1 = str(WY2[k][1])
                if (re.findall('[.](.*)', b1, flags=0) == []):
                    bg1 = b1 + '0000'
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 1):
                    bg1 = b1 + ling3
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 2):
                    bg1 = b1 + ling2
                elif (len(re.findall('[.](.*)', b1, flags=0)[0]) == 3):
                    bg1 = b1 + ling1
                else:
                    bg1 = b1
                c1 = str(cj_random2)
                if (re.findall('[.](.*)', c1, flags=0) == []):
                    cg1 = c1 + '0000'
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 1):
                    cg1 = c1 + ling3
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 2):
                    cg1 = c1 + ling2
                elif (len(re.findall('[.](.*)', c1, flags=0)[0]) == 3):
                    cg1 = c1 + ling1
                else:
                    cg1 = c1
                a2 = str(WY3[k][0])
                if (re.findall('[.](.*)', a2, flags=0) == []):
                    ag2 = a2 + '0000'
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 1):
                    ag2 = a2 + ling3
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 2):
                    ag2 = a2 + ling2
                elif (len(re.findall('[.](.*)', a2, flags=0)[0]) == 3):
                    ag2 = a2 + ling1
                else:
                    ag2 = a2
                b2 = str(WY3[k][1])
                if (re.findall('[.](.*)', b2, flags=0) == []):
                    bg2 = b2 + '0000'
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 1):
                    bg2 = b2 + ling3
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 2):
                    bg2 = b2 + ling2
                elif (len(re.findall('[.](.*)', b2, flags=0)[0]) == 3):
                    bg2 = b2 + ling1
                else:
                    bg2 = b2
                c2 = str(cj_random3)
                if (re.findall('[.](.*)', c2, flags=0) == []):
                    cg2 = c2 + '0000'
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 1):
                    cg2 = c2 + ling3
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 2):
                    cg2 = c2 + ling2
                elif (len(re.findall('[.](.*)', c2, flags=0)[0]) == 3):
                    cg2 = c2 + ling1
                else:
                    cg2 = c2
                fid.write(name1 + '，' + ag + '，' + bg + '，' + cg + "\n")
                fid.write(name2 + '，' + ag1 + '，' + bg1 + '，' + cg1 + "\n")
                fid.write(name3 + '，' + ag2 + '，' + bg2 + '，' + cg2 + "\n")
                k = k + 1
            fid.close()

class original_cj(QDialog):
    def __init__(self):
        super(original_cj, self).__init__()
        self.resize(600, 300)
        self.setWindowTitle('沉降原始数据还原')
        self.label1 = QLabel('数据库文件路径', self)
        self.editline1 = QLineEdit(self)
        self.label2 = QLabel('数据库名称', self)
        self.editline2 = QLineEdit(self)
        self.label4 = QLabel('输入期次范围', self)
        self.editline4 = QLineEdit(self)
        self.button1 = QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.V_layout3 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout__init()
        self.display_editline()

    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label4)
        self.V_layout2.addWidget(self.editline1)
        self.V_layout2.addWidget(self.editline2)
        self.V_layout2.addWidget(self.editline4)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)

    def display_editline(self):
        self.editline1.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline2.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline4.setPlaceholderText('请输入期数')
        # self.button1.clicked.connect(self.output_excel)
        # self.editline5.setPlaceholderText('从日历选择日期')
class make_dir(QDialog):
    def __init__(self):
        super(make_dir,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('文件创立')
        self.label1=QLabel('原始数据及日报的路径',self)
        self.editline1=QLineEdit(self)
        self.label2=QLabel('目标文件夹路径',self)
        self.editline2=QLineEdit(self)
        self.label4=QLabel('数据路径',self)
        self.editline4=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始创建原始数据文件')
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label4)
        self.V_layout2.addWidget(self.editline1)
        self.V_layout2.addWidget(self.editline2)
        self.V_layout2.addWidget(self.editline4)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline2.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline4.setPlaceholderText('需要严格按照Python路径输入！')
        self.button1.clicked.connect(self.create_dir)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def create_dir(self):
        # path1 = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据9x\\"
        # path3 = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据3\\"
        # dateset_path = "D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据2\\4-2数据库.xlsx"
        path1=self.editline2.text()
        path3=self.editline1.text()
        dateset_path=self.editline4.text()
        workbook = load_workbook(dateset_path)
        sheet1 = workbook.get_sheet_by_name("日报")
        for zz in range(2, sheet1.max_row + 1):
            if (((sheet1.cell(zz, 2).value is None) or (sheet1.cell(zz, 2).value == '')) or (
                    sheet1.cell(zz, 2).value == [])):  # 根据判断是否为无参数None或者空来判断行数
                row2 = zz - 1
                break
            else:
                row2 = sheet1.max_row  # 根据日报获取期数
        file_names = os.listdir(path3)
        a = '周边道路'  # a='周边道路'
        b = '管线'  # b='周边管线'
        c = '周边环境'  # c='周边地表'
        d = '建筑'  # d='周边建筑'
        e = '坡顶沉降'  # e='坡顶沉降'
        f = '桩顶沉降'
        g = '坡顶位移'
        flagg = 0
        numg = []
        nameg = []
        flagh = 0
        numh = []
        nameh = []
        for i in range(len(file_names)):
            if ('坐标' in file_names[i]):
                numg.append(int(re.findall('(.*)观测记录', file_names[i], flags=0)[0]))
                nameg.append(file_names[i])
                flagg = 1
        for i in range(len(file_names)):
            if ('监测日报' in file_names[i]):
                numh.append(int(re.findall('(.*)监测日报', file_names[i], flags=0)[0]))
                nameh.append(file_names[i])
                flagh = 1
        numa = []
        numb = []
        numc = []
        numd = []
        nume = []
        numf = []
        flaga = 0
        flagb = 0
        flagc = 0
        flagd = 0
        flage = 0
        flagf = 0
        namea = []
        nameb = []
        namec = []
        named = []
        namee = []
        namef = []
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (a in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numa.append(qi_num)
                    namea.append(file_names[i])
                    flaga = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (b in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numb.append(qi_num)
                    nameb.append(file_names[i])
                    flagb = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (c in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numc.append(qi_num)
                    namec.append(file_names[i])
                    flagc = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (d in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numd.append(qi_num)
                    named.append(file_names[i])
                    flagd = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (e in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    nume.append(qi_num)
                    namee.append(file_names[i])
                    flage = 1
        for i in range(len(file_names)):
            if ('沉降' in file_names[i]):  ##属于水准沉降原始文降
                if (f in file_names[i]):
                    qi_num = int(re.findall('降(.*)[.]', file_names[i], flags=0)[0])
                    numf.append(qi_num)
                    namef.append(file_names[i])
                    flagf = 1
        if (flaga == 0):
            numa = [0]
        if (flagb == 0):
            numb = [0]
        if (flagc == 0):
            numc = [0]
        if (flagd == 0):
            numd = [0]
        if (flage == 0):
            nume = [0]
        if (flagf == 0):
            numf = [0]
        if (flagg == 0):
            numg = [0]
        if (flagh == 0):
            numh = [0]
        num_max = np.max(
            [np.max(numa), np.max(numb), np.max(numc), np.max(numd), np.max(nume), np.max(numf), np.max(numg),
             np.max(numh)])  ###共a、b、c、d、e、f、g、h比较
        if (flaga == 0):
            numa = [1000000]
        if (flagb == 0):
            numb = [1000000]
        if (flagc == 0):
            numc = [1000000]
        if (flagd == 0):
            numd = [1000000]
        if (flage == 0):
            nume = [1000000]
        if (flagf == 0):
            numf = [1000000]
        if (flagg == 0):
            numg = [1000000]
        num_min = np.min(
            [np.min(numa), np.min(numb), np.min(numc), np.min(numd), np.min(nume), np.min(numf), np.min(numg),
             np.min(numg)])  ###共a、b、c、d、e、f、g、h比较
        qua_path = "D:\\2021\\基坑监测\\2018巷口\\质量评定.doc"
        for j in range(num_min + 1, num_max + 2):  ##num_max>num_min############需要有两期以上
            date1 = sheet1.cell(j, 2).value
            date3 = "第" + str(j - 1) + "期" + str(date1)
            x1 = date3.replace('-', '年', 1)
            x2 = x1.replace('-', '月', 1)
            x3 = x2.replace(' 00:00:00', '日', 1)  # 通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
            datex2 = path1 + x3
            os.makedirs(datex2)
            datex3 = datex2 + '\\' + '原始数据'
            os.makedirs(datex2 + '\\' + '原始数据')
            if (flagh == 1):
                os.makedirs(datex2 + '\\' + '日报')
                for name in nameh:
                    if (re.findall('(.*)监测日报', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '日报')  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagg == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + g)
                for name in nameg:
                    if (re.findall('(.*)观测记录', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + g)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flaga == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + a)
                for name in namea:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + a)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagb == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + b)
                for name in nameb:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + b)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagc == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + c)
                for name in namec:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + c)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagd == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + d)
                for name in named:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + d)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flage == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + e)
                for name in namee:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + e)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下
            if (flagf == 1):
                os.makedirs(datex2 + '\\' + '原始数据' + '\\' + f)
                for name in namef:
                    if (re.findall('降(.*)[.]', name, flags=0)[0] == str(j - 1)):
                        shutil.copy(path3 + name, datex2 + '\\' + '原始数据' + '\\' + f)  # 将通过遍历源目标文降下的dat文降拷贝到对应的目标文降夹下

if __name__== '__main__':
    app=QApplication(sys.argv)
    demo=zhu_ti_cj_jiang()
    demo.show()
    sys.exit(app.exec_())