import os
import shutil
from openpyxl import load_workbook

# path1="D:\\2021\\基坑监测\\2019巷口\\原始数据2\\" #目标文件路径
path1="D:\\2021\\基坑监测\\华侨城万科4-2地块\\原始数据1\\"
# path3="D:\\2021\\基坑监测\\2019巷口\\原始数据1\\" #源文件路径同上
path3="D:\\2021\\基坑监测\\华侨城万科4-2地块\\华侨城4-2还原原始数据\\"
path3="D:\\2021\\基坑监测\\华侨城万科4-2地块\\数据库\\4-2数据库 - 副本 - 副本-报表\\"
# dateset_path="D:\\2021\\基坑监测\\2019巷口\\数据库.xlsx"
dateset_path="D:\\2021\\基坑监测\\华侨城万科4-2地块\\数据库\\4-2数据库 - 副本 - 副本.xlsx"
workbook = load_workbook(dateset_path)
sheet1=workbook.get_sheet_by_name("日报")
for zz in range(2,sheet1.max_row+1):
    if((sheet1.cell(zz,2).value is None) or (sheet1.cell(zz,2).value=='')):#根据判断是否为无参数None或者空来判断行数
        row2=zz-1
        break
    else:
        row2=sheet1.max_row #根据日报获取期数
file_names = os.listdir(path3)
print(len(file_names))
print("+++++++++++++++++++++")
qua_path="D:\\2021\\基坑监测\\2018巷口\\质量评定.doc"
for j in range(130,row2+1):
    date1=sheet1.cell(j,2).value
    date3="第"+str(j-1)+"期"+str(date1)
    x1=date3.replace('-','年',1)
    x2=x1.replace('-','月',1)
    x3=x2.replace(' 00:00:00','日',1) #通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
    datex2 = path1 + x3
    # os.makedirs(datex2)
    datex3=datex2+'\\'+'原始数据'
    # os.makedirs(datex2+'\\'+'原始数据')
    # os.makedirs(datex2 + '\\' + '中间报告')
    for name in file_names:
        path4 = path3 + '\\' + name
        B1=name[-3:]
        if(name[-3:]=='dat'):
            # y2=name.index('降')########根据名称的最后一个字来进行遍历
            # y3=name.index('.')
            # if(str(j-1)==name[y2+1:y3]):
            if (str(j - 1) == name[0:3]):
                shutil.copy(path4, datex3) #将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下
                shutil.copy(qua_path, datex2)