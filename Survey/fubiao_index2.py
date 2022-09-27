import pandas as pd
import openpyxl
import os
import shutil
import re
import math
import numpy
from openpyxl.styles import PatternFill
path="C:\\Users\\Kong\\Desktop\\2022\\滁槎三期二地块基坑数据库\\坡顶沉降成果附表11.xlsx"
book = openpyxl.load_workbook(path)
sheet=book.get_sheet_by_name('坡顶沉降成果表')
data=pd.read_excel(path,'坡顶沉降成果表')
def baoliu1(a):
    ling1 = '0'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数2为 字符串
        ag = a + '0'
    else:
        ag = a
    return ag

def baoliu2(a):
    ling1 = '0'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数2为 字符串
        ag = a + '00'
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
        ag = a + ling1
    else:
        ag = a
    return ag
fill = openpyxl.styles.PatternFill("solid", fgColor="1874CD")
# fill_1 = openpyxl.styles.PatternFill("solid", fgColor="red")
for i in range(2,data.shape[0]-1):
    num=[]
    for j in range(1,data.shape[1]):
        if (isinstance(data.iloc[i, j], str)):
            if(j%2==0):
                sheet.cell(i+2,j+1).value=baoliu1(str(abs(float(data.iloc[i,j]))))
            else:
                sheet.cell(i + 2, j + 1).value =baoliu2(str(abs(float(data.iloc[i, j]))))
    for j in range(1,data.shape[1],2):
        if(isinstance(data.iloc[i,j],str)):
            data.iloc[i,j]=float(data.iloc[i,j])
    for j in range(1, data.shape[1], 2):
        if (numpy.isnan(data.iloc[i, j])):
            pass
        else:
            num.append(abs(float(data.iloc[i,j])))
        if (j + 2 != data.shape[1]):
            if( (~numpy.isnan(data.iloc[i, j]) and numpy.isnan(data.iloc[i, j+2]))):
                break
    max1 = numpy.max(num)
    for j in range(1,data.shape[1],2):
        if(isinstance(data.iloc[i,j],str)):
            data.iloc[i,j]=float(data.iloc[i,j])
        if(abs(float(data.iloc[i,j]))==max1):
            sheet.cell(i+2,j+1).fill=fill


for i in range(2,data.shape[0]-1):
    num=[]
    # for j in range(1,data.shape[1]):
    #     if (isinstance(data.iloc[i, j], str)):
    #         sheet.cell(i+2,j+1).value=abs(float(data.iloc[i,j]))
    for j in range(2,data.shape[1],2):
        if(isinstance(data.iloc[i,j],str)):
            data.iloc[i,j]=float(data.iloc[i,j])
    for j in range(2, data.shape[1], 2):
        if (numpy.isnan(data.iloc[i, j])):
            pass
        else:
            num.append(abs(float(data.iloc[i,j])))
        if (j + 1 != data.shape[1]):
            if( (~numpy.isnan(data.iloc[i, j]) and numpy.isnan(data.iloc[i, j+2]))):
                break
    max1 = numpy.max(num)
    for j in range(2,data.shape[1],2):
        if(isinstance(data.iloc[i,j],str)):
            data.iloc[i,j]=float(data.iloc[i,j])
        if(abs(float(data.iloc[i,j]))==max1):
            sheet.cell(i+2,j+1).fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
book.save(path)
# import pandas as pd
# import openpyxl
# import os
# import shutil
# import re
# import math
# import numpy
# path="D:\\2022\\基坑监测\\明园九龙湾\\成果表\\深层水平位移成果表1.xlsx"
# book = openpyxl.load_workbook(path)
# for p in range(21):
#     name='CX'+str(p+1)
#     sheet=book.get_sheet_by_name(name)
#     data=pd.read_excel(path,name)
#     fill = openpyxl.styles.PatternFill("solid", fgColor="1874CD")
#     for i in range(2,data.shape[0]-1):
#         num=[]
#         for j in range(1,data.shape[1],2):
#             if(isinstance(data.iloc[i,j],str)):
#                 data.iloc[i,j]=float(data.iloc[i,j])
#         for j in range(1, data.shape[1], 2):
#             if (numpy.isnan(data.iloc[i, j])):
#                 pass
#             else:
#                 num.append(abs(float(data.iloc[i,j])))
#             if (j + 2 != data.shape[1]):
#                 if( (~numpy.isnan(data.iloc[i, j]) and numpy.isnan(data.iloc[i, j+2]))):
#                     break
#         max1 = numpy.max(num)
#         for j in range(1,data.shape[1],2):
#             if(isinstance(data.iloc[i,j],str)):
#                 data.iloc[i,j]=float(data.iloc[i,j])
#             if(abs(float(data.iloc[i,j]))==max1):
#                 sheet.cell(i+2,j+1).fill=fill
# book.save(path)

