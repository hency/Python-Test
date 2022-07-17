import numpy
import openpyxl
# import xlwt
# import xlrd
book1=openpyxl.Workbook()
sheet1=book1.create_sheet('sheet')
name1=[11,11,11,11,11,11,11,11]
name2=[1,2,3,5,7,8,9,10]
# name1=[9,9,9,9,9,9]
# name2=['S1','S2','S3','S4','S5','S6'] #######当楼栋号为S时需要简单调动代码
str1='#沉降'
# str1='沉降'
str5='.dat'
str6="D:\\2022\\主体沉降\\润永通\\数据库\\"
str4=[]
r3=[]
k=1
for i in range(0,len(name2)):
    str3=[]
    r2=[]
    k1=1
    for j in range(1,name1[i]+1):
        str2=str(name2[i])+str1+str(j)+str5
        # str2 = name2[i]+ str1 + str(j) + str5
        str7=str6+str2
        f=open(str7,'r')
        file1=f.readlines()
        count1=len(file1)
        r1=file1[count1-3-1][58:66]
        # print(file1[count1-3-1])
        f.close()
        r2.append(float(r1))
        str3.append(str2)
        sheet1.cell(row=k1,column=k).value=float(r1)*1000
        k1=k1+1
        # print(str2)
    str4.append(str3)
    r3.append(r2)
    k=k+1
book1.save(str6+'bihecha.xlsx')

    # print(r3)
# with open(r"E:\文件3\1#沉降1.dat", 'r') as f:
#     count = len(f.readlines())
# # print(count)
# f=open(r"E:\文件3\1#沉降1.dat", 'r')
# file1=f.readlines()
# print(len(file1))
# # for i in range(0,766):
# #     print(file1[i])
# print(file1[766-3-1])
# aa=file1[766-3-1];
# print(aa[58:66])
# f.close()
