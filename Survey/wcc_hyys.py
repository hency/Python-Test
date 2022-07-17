import openpyxl
from openpyxl import load_workbook
import os
import re
import shutil
# path="D:\\Desktop\\吴昌程还原原始数据\\第175期2021.6.27\\原始数据\\测斜\\CX1.xlsx"
path1="D:\\Desktop\\吴昌程还原原始数据\\原始数据4\\"
path2="D:\\Desktop\\吴昌程还原原始数据\\中骏\\中骏世界城基坑数据库.xlsx"
path3="D:\\Desktop\\吴昌程还原原始数据\\原始数据5\\测斜\\"
path4="D:\\Desktop\\吴昌程还原原始数据\\原始数据5\\坡顶沉降\\坡顶水准观测文件292.dat"
path5="D:\\Desktop\\吴昌程还原原始数据\\原始数据5\\位移\\292观测记录2022-1-21坐标.dat"
path6="D:\\Desktop\\吴昌程还原原始数据\\中骏\\日报\\"
path7="D:\\Desktop\\吴昌程还原原始数据\\原始数据5\\管线沉降\\管线水准观测文件292.dat"
book=load_workbook(path2)
sheet=book.get_sheet_by_name('日报')
dataset_name=os.listdir(path6)
for i in range(260,292):
    date=sheet.cell(i+1,2).value
    date3="第"+str(i)+"期"+str(date)
    x1=date3.replace('-','年',1)
    x2=x1.replace('-','月',1)
    x3=x2.replace(' 00:00:00','日',1) #通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
    x4=path1+x3
    os.makedirs(x4)
    x5=x4+'\\'+'日报'
    os.makedirs(x5)
    x6=x4+'\\'+'原始数据'
    os.makedirs(x6)
    x7=x6+'\\'+'测斜'
    os.makedirs(x7)
    x8=x6+'\\'+'坡顶沉降'
    os.makedirs(x8)
    x9=x6+'\\'+'位移'
    os.makedirs(x9)
    x10=x6+'\\'+'管线沉降'
    os.makedirs(x10)
    name1=os.listdir(path3)
    for j in range(len(name1)):
        path_cx1=path3+name1[j]
        book1=load_workbook(path_cx1)
        sheet1=book1.get_sheet_by_name(book1.sheetnames[0])
        sheet1.title=x3
        book1.save(x7+'\\'+'CX'+str(j+1)+'.xlsx')
    a=re.findall('(.*)年',re.findall('期(.*)',x3,flags=0)[0],flags=0)[0]
    b=re.findall('年(.*)月',re.findall('期(.*)',x3,flags=0)[0],flags=0)[0]
    c=re.findall('月(.*)日',re.findall('期(.*)',x3,flags=0)[0],flags=0)[0]
    str1='期次:'+str(i)+' 日期：'+a+'/'+b+'/'+c
    str2=a+b+c+'.dat'
    str3 = str(i) + '观测记录' + a + '-' + b + '-' + c + '坐标' + '.dat'
    ############坡顶沉降
    n = 0
    with open(path4) as f1:
        message = ''
        for line in f1:
            if (n == 0):
                line = str1 + '\n'
            if (n == 1):
                line = re.findall('(.*)T', line, flags=0)[
                           0] + 'TO  ' + str2 + '               |                      |                      |                      |' + '\n'
            message += line
            n = n + 1
    with open(x8 + '\\' + '水准观测文件' + str(i) + '.dat', 'w') as f2:
        f2.write(message)
    ###################管线沉降
    n=0
    with open(path7) as f1:
        message = ''
        for line in f1:
            if (n == 0):
                line = str1 + '\n'
            if (n == 1):
                line = re.findall('(.*)T', line, flags=0)[
                           0] + 'TO  ' + str2 + '               |                      |                      |                      |' + '\n'
            message += line
            n = n + 1
    with open(x10 + '\\' + '水准观测文件' + str(i) + '.dat', 'w') as f2:
        f2.write(message)
    ################################位移
    with open(path5) as f1:
        message = ''
        for line in f1:
            message += line
    with open(x9 + '\\' + str3, 'w') as f2:
        f2.write(message)
    for z in range(len(dataset_name)):
        if(re.findall('(.*)监测日报',dataset_name[z],flags=0)[0]==str(i)):
            shutil.copy(path6+dataset_name[z],x5)



# book1=load_workbook(path)
# sheet1=book1.get_sheet_by_name(book1.sheetnames[0])
# sheet1.title='abc'
# book1.save(path)
# wb = openpyxl.Workbook()
#     #当前打开的sheet页 wb.active
#     ws = wb.active
#
#     # #更改默认名称Sheet`
#     ws.title = "WorkSheetTitle"
#
#     # 定义第二个sheet页
#     ws2 = wb.create_sheet("NewWorkSheet2")
#
#     # 定义第三个sheet页
#     # `0` 的设定 会将该sheet页 置于wb最前面
#     ws3 = wb.create_sheet("NewWorkSheet3", 0)
#
#     # 保存
#     wb.save('example.xlsx')