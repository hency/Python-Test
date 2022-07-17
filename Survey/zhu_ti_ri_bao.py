import numpy
import os
import shutil
import openpyxl
import pandas as pd
from 沉降观测1 import read_measure_line_from_dataset   ##从数据库中提取监测日期、监测日期数据
path="D:\\2021\\主体沉降\\主体沉降代码\\"
path1="D:\\2021\\主体沉降\\主体沉降代码\\测试数据库1.xlsx" ##数据库文件路径
path2="D:\\2021\\主体沉降\\主体沉降代码\\测试线路文件1.xlsx" ##测线文件路径
sheet_name='建筑沉降成果表'
pd1=pd.read_excel(path1,sheet_name)
gc_name,cx_make=read_measure_line_from_dataset(path1,path2)
path3 = "D:\\2021\\主体沉降\\主体沉降代码\\上饶7号楼主体沉降测试2.xlsx"  ##输出的日报的EXCEL文件模板
# name = '第%d栋楼的沉降观测日报.xlsx'.format(i)
path_name1 = "D:\2021\主体沉降\主体沉降代码\上饶7号楼主体沉降测试2_1.xlsx"
shutil.copy(path3, path_name1)
book3 = openpyxl.load_workbook(path_name1)
sheet3 = book3.get_sheet_by_name('7#沉降观测')
for i in range(2,len(cx_make)):###前提是主体沉降的期次大于2
    value=cx_make[i]
    for j in range(len(gc_name)):
        sheet3.cell(3+j,1).value=gc_name[j]
        sheet3.cell(3+j,2).value=cx_make[0][j]
        sheet3.cell(3+j,3).value=value[j]
        if(i==1):
            sheet3.cell(3+j,4).value=cx_make[0][j]##初次的为第一次的高程
            sheet3.cell(3 + j, 5).value = value[j] - cx_make[0][j]
        else:
            sheet3.cell(3+j,4).value=cx_make[i-1][j]###非初次的为上次高程
            sheet3.cell(3+j,5).value=value[j]-cx_make[i-1][j]