import re
import openpyxl
import pandas as pd
import numpy as np
import os
from  datetime import datetime
name1='周边地表沉降'
name2='基坑坡顶部水平位移'
name3='竖向位移'
name4='深层水平位移'
name5='地下水位'
name6='锚杆应力'
book1=openpyxl.Workbook()
sheet_dibiao=book1.create_sheet(name1)
sheet_jikeng_wy=book1.create_sheet(name2)
sheet_cj=book1.create_sheet(name3)
sheet_sc_wy=book1.create_sheet(name4)
sheet_sw=book1.create_sheet(name5)
sheet_zl=book1.create_sheet(name6)
path_Z="D:\\2022\\基坑监测\\明园九龙湾\\日报1--1\\日报"
name_Z=os.listdir(path_Z)
name_Z1=[]
for i in range(len(name_Z)):
    if('xlsm' in name_Z[i]):
        name_Z1.append(name_Z[i])
    else:
        pass

for i in range(0,len(name_Z1)-1):
    for j in range(i,len(name_Z1)):
        a = datetime.strptime(re.findall('(.*)明园', name_Z1[i], flags=0)[0], '%Y.%m.%d')  ################注意一定是的大写的Y
        b=datetime.strptime(re.findall('(.*)明园',name_Z1[j],flags=0)[0],'%Y.%m.%d') ################注意一定是的大写的Y
        if(a.__gt__(b)):
            mid=name_Z1[i]
            name_Z1[i]=name_Z1[j]
            name_Z1[j]=mid

book2=openpyxl.Workbook()
sheet2=book2.create_sheet('Date')
for i in range(len(name_Z1)):
    a=datetime.strptime(re.findall('(.*)明园', name_Z1[i], flags=0)[0], '%Y.%m.%d')
    sheet2.cell(i+1,1).value=a
book2.save(path_Z+'\\'+'DATE.xlsx')

k1=1
k2=1
k3=1
k4=1
k5=1
k6=1
for z in range(len(name_Z1)):
    path=path_Z+'\\'+name_Z1[z]
    #path="D:\\2022\\基坑监测\\明园九龙湾\\2019.12.16明园九龙湾G01地块监测成果表.xlsm"
    data=pd.read_excel(path,'监测日报表')
    cols=data.columns
    dibiao_data=[]
    dibiao_name=[]
    jikeng_wy_data=[]
    jikeng_wy_name=[]
    jikeng_cj_data=[]
    jikeng_cj_name=[]
    sc_wy_data=[]
    sc_wy_name=[]
    sw_data=[]
    sw_name=[]
    zl_data=[]
    zl_name=[]
    kx1 = 1
    kx2 = 1
    kx3 = 1
    kx4 = 1
    kx5 = 1
    kx6 = 1
    for i in range(1,len(cols)):
        if('Unnamed' in cols[i]):
            continue
        else:
            if(name1 in cols[i]):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[j,i],float)):
                        data.iloc[j,i]=str(data.iloc[j,i])
                    if(('CJ' in data.iloc[j,i] and '注' in data.iloc[j,i]) or ('CJ' not in data.iloc[j,i])):
                        break
                    else:
                        dibiao_name.append(data.iloc[j,i])
                        dibiao_data.append(data.iloc[j,i+3])
                        # sheet_dibiao.cell(kx1,k1).value=data.iloc[j,i]
                        sheet_dibiao.cell(kx1,k1).value=data.iloc[j,i+3]
                        kx1=kx1+1
            if(name2 in cols[i] ):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[j,i],float)):
                        data.iloc[j,i]=str(data.iloc[j,i])
                    else:
                        pass
                    if(('WY' in data.iloc[j,i] and '注' in data.iloc[j,i]) or ('WY' not in data.iloc[j,i])):
                        break
                    else:
                        # sheet_jikeng_wy.cell(kx2,k2).value=data.iloc[j,i]
                        sheet_jikeng_wy.cell(kx2,k2).value=data.iloc[j,i+3]
                        kx2=kx2+1
            if(name3 in cols[i]):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[j,i],float)):
                        data.iloc[j,i]=str(data.iloc[j,i])
                    else:
                        pass
                    if(('WY' in data.iloc[j,i] and '注' in data.iloc[j,i]) or ('WY' not in data.iloc[j,i])):
                        break
                    else:
                        # sheet_cj.cell(kx3,k3).value=data.iloc[j,i]
                        sheet_cj.cell(kx3,k3).value=data.iloc[j,i+3]
                        kx3=kx3+1
            if(name5 in cols[i]):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[j,i],float)):
                        data.iloc[j,i]=str(data.iloc[j,i])
                    else:
                        pass
                    if('SW' not in data.iloc[j,i]):
                        break
                    else:
                        # sheet_sw.cell(kx5,k5).value=data.iloc[j,i]
                        sheet_sw.cell(kx5,k5).value=data.iloc[j,i+3]
                        kx5=kx5+1
            if(name6 in cols[i]):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[j,i],float)):
                        data.iloc[j,i]=str(data.iloc[j,i])
                    else:
                        pass
                    if('ZL' not in data.iloc[j,i]):
                        break
                    else:
                        # sheet_zl.cell(kx6,k6).value=data.iloc[j,i]
                        sheet_zl.cell(kx6,k6).value=data.iloc[j,i+3]
                        kx6=kx6+1

    k1 = k1 + 1
    k2 = k2+1
    k3 = k3 + 1
    k5 = k5 + 1
    k6 = k6 + 1
# for i in range(25):
#     sheet1=book1.create_sheet('CX'+str(i+1))
#     kx7 = 1
#     for z in range(len(name_Z1)):#len(name_Z1)
#         path = path_Z + '\\' + name_Z1[z]
#         # path="D:\\2022\\基坑监测\\明园九龙湾\\2019.12.16明园九龙湾G01地块监测成果表.xlsm"
#         data = pd.read_excel(path, '监测日报表')
#         cols = data.columns
#         for j in range(1,len(cols)):
#             if ('Unnamed' in cols[j]):
#                 continue
#             else:
#                 if(name4 in cols[j]):
#                     if(data.iloc[4,j]=='CX'+str(i+1)):
#                         for q in range(4,data.shape[0]):
#                             if(np.isnan(data.iloc[q,j+4])):
#                                 break
#                             else:
#                                 sheet1.cell(q-3,kx7).value=data.iloc[q,j+4]
#                         break
#         kx7 = kx7 + 1
book1.save(path_Z+'\\'+'tiqu11.xlsx')



