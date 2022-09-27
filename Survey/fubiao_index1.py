###该程序是测量公司的基坑总结报告生成的附表
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
import os
import math
import  shutil
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
path_src="D:\\2022\\吕建勋生成原始数据\\金茂海棠悦府数据库-.xlsx"
path_moban="D:\\2022\\吕建勋生成原始数据\\桩顶沉降附表.xlsx"
name1='周边地表'
name2='周边管线'
name3='周边道路'
name4='周边环境'
name5='周边建筑'
name6='坡顶沉降'
name7_1='坡顶沉降'
name7_2='坡顶竖向位移'
name7_3='桩顶沉降'
name7_4='桩顶竖向位移'
name8_1='坡顶水平位移'
name8_2='桩顶水平位移'
name9='深层位移'
name10='轴力'
name11='水位'

book=load_workbook(path_src)
book1=load_workbook(path_moban)
sheet_write_dibiao=book1.get_sheet_by_name(name7_3+'成果表') #####@@@@@@@@@@@@@@@@@@@@@@@@   name
sheetnames=book.sheetnames
sheet_date = '日报'
data_date = pd.read_excel(path_src, sheet_date)
date=[]


def baoliu2(a):
    ling1 = '0'
    if (re.findall('[.](.*)', a, flags=0) == []):  ########保留小数2为 字符串
        ag = a + '00'
    elif (len(re.findall('[.](.*)', a, flags=0)[0]) == 1):
        ag = a + ling1
    else:
        ag = a
    return ag

for i in range(0,data_date.shape[0]):
    date.append(data_date.iloc[i,1])
for i in range(0,len(date)):
    date1=str(date[i]).replace(' 00:00:00','')
    sheet_write_dibiao.merge_cells(start_row=1,start_column=i*2+2,end_row=1,end_column=i*2+2+1)
    sheet_write_dibiao.merge_cells(start_row=2, start_column=i * 2 + 2, end_row=2, end_column=i * 2 + 2 + 1)
    sheet_write_dibiao.cell(2,i*2+2).value=date1
    sheet_write_dibiao.cell(1,i*2+2).value='第'+str(i+1)+'期'
    sheet_write_dibiao.column_dimensions[get_column_letter(i*2+2)].width=16
    sheet_write_dibiao.column_dimensions[get_column_letter(i * 2 + 2+1)].width = 16
    sheet_write_dibiao.cell(1, i * 2 + 2).alignment = Alignment(horizontal="center", vertical="center")
    sheet_write_dibiao.cell(2,i*2+2).alignment=Alignment(horizontal="center", vertical="center")
    sheet_write_dibiao.cell(3,i * 2 + 2).value='速率（mm/d）'
    sheet_write_dibiao.cell(3,i * 2 + 2).alignment=Alignment(wrap_text=True)
    sheet_write_dibiao.cell(3, i * 2 + 2).alignment=Alignment(horizontal="center", vertical="center")
    sheet_write_dibiao.cell(3,i*2+2+1).value='累计变化量（mm）'
    sheet_write_dibiao.cell(3, i * 2 + 2 + 1).alignment=Alignment(wrap_text=True)
    sheet_write_dibiao.cell(3, i * 2 + 2 + 1).alignment=Alignment(horizontal="center", vertical="center")
    #cell.alignment = Alignment(horizontal="center", vertical="center")
for i in range(len(sheetnames)):
    if(name7_3 in sheetnames[i]):   ####周边地表沉降数据的数据提取#####@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@  name
        sheet1=sheetnames[i]
        data1=pd.read_excel(path_src,sheet1)
        dnames=[]
        for j in range(3,data1.shape[1]):
            dnames.append(data1.iloc[10,j])
            sheet_write_dibiao.cell(j+1,1).value=data1.iloc[10,j]
        for j in range(3,data1.shape[1]):
            asum=0
            for z in range(11,11+len(date)-1):
                if((~np.isnan(data1.iloc[z,j]) and ~np.isnan(data1.iloc[z+1,j])) and (z==11)):
                    ###初始值1
                        sheet_write_dibiao.cell(j+1,2).value='0.00'
                        sheet_write_dibiao.cell(j + 1, 3).value = '0.00'
                        sheet_write_dibiao.cell(j + 1, 4).value = baoliu2(str(round(((data1.iloc[z+1, j ] - data1.iloc[z, j]) / int(re.findall('(.*) days',str(date[1] - date[0]))[0]))*1000,2)))
                        sheet_write_dibiao.cell(j + 1, 5).value = baoliu2(str(round(((data1.iloc[z+1, j] - data1.iloc[z, j]))*1000,2)))
                        sheet_write_dibiao.cell(j+1,2).alignment=Alignment(horizontal="center", vertical="center")
                        sheet_write_dibiao.cell(j + 1, 3).alignment=Alignment(horizontal="center", vertical="center")
                        sheet_write_dibiao.cell(j + 1, 4).alignment=Alignment(horizontal="center", vertical="center")
                        sheet_write_dibiao.cell(j + 1, 5).alignment=Alignment(horizontal="center", vertical="center")
                        asum=asum+(data1.iloc[z+1, j] - data1.iloc[z, j])*1000
                if((~np.isnan(data1.iloc[z,j]) and ~np.isnan(data1.iloc[z+1,j])) and (z!=11) ):
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2).value = baoliu2(str(round(((data1.iloc[z+1,  j] - data1.iloc[z, j]) / int(re.findall('(.*) days',str(date[z-11+1] - date[z-11]))[0]))*1000,2)))
                    asum = asum + (data1.iloc[z+1,  j] - data1.iloc[z, j])*1000
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2+1).value = baoliu2(str(round(asum,2)))
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2).alignment=Alignment(horizontal="center", vertical="center")
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2+ 1).alignment=Alignment(horizontal="center", vertical="center")
                if((np.isnan(data1.iloc[z,j]) and ~np.isnan(data1.iloc[z+1,j]))):
                    sheet_write_dibiao.cell(j+1,(z-11+2)*2).value='0.00'
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2+1).value = '0.00'
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2).alignment=Alignment(horizontal="center", vertical="center")
                    sheet_write_dibiao.cell(j + 1, (z-11+2)*2 + 1).alignment=Alignment(horizontal="center", vertical="center")
                if((~np.isnan(data1.iloc[z,j]) and np.isnan(data1.iloc[z,j]))):
                    break

book1.save(path_moban)