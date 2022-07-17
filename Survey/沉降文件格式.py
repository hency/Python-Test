import copy

import numpy
from openpyxl import load_workbook

#头文件Header-file 文件格式为M5
#命名格式为：水准观测文件
ss1='For M5'
ss2='AAA'
def header(times,date):
    file_name="水准观测文件"
    path="D:\\Desktop\\原始文件\\"+str(times)+"期"+file_name+'.dat'
    with open(path,'w+') as f:
            f.write("期次："+str(times)+'\n'+'日期：'+date+'\n')
            str1=[]
            for i in range(0, 124):
                str1.append(' ')
            str1[6] = '|'
            str1[16] = '|'
            str1[48] = '|'
            str1[71] = '|'
            str1[94] = '|'
            str1[117] = '|'
            str4=copy.deepcopy(str1)
            str1[0:len(ss1)]=ss1
            ss3='TO  ' + str(times) + "期" + file_name + '.dat'
            str1[7:len(ss2)+7]=ss2
            str1[17:len(ss3)+17]=ss3
            f.write(''.join(str1)+'\n')
            str2=[]
            for i in range(0, 123):
                str2.append(' ')
            str2[6] = '|'
            str2[16] = '|'
            str2[53] = '|'
            str2[76] = '|'
            str2[99] = '|'
            str2[122] = '|'
            str2[0:len(ss1)] = ss1
            str2[7:len(ss2)+7]=ss2
            ss3 = 'TO  start-line   aBBFF'
            str2[17:len(ss3) + 17]=ss3
            f.write(''.join(str2))
            str4[104:116]='END OF HEADER'
            f.write('\n'+''.join(str4))
workbook=load_workbook('E:\\文件3\\1#数据库.xlsx')
sheet1 = workbook.get_sheet_by_name("日报")
for zz in range(2,sheet1.max_row+1):
    if((sheet1.cell(zz,2).value is None) or (sheet1.cell(zz,2).value=='')):#根据判断是否为无参数None或者空来判断行数
        row1=zz-1
        break
    else:
        row1=sheet1.max_row #根据日报获取期数
for i in range(2,row1):
    date=sheet1.cell(i,2).value
    date1 = str(date)
    x1 = date1.replace('-', '年', 1)
    x2 = x1.replace('-', '月', 1)
    x3 = x2.replace(' 00:00:00', '日', 1)  # 通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
    times=i-1
    header(times,x3)



#打开水准测线文件

#从监测数据库调取高程

#剔除无高程的支点

#前后视距差应小于0.8 #累计差小于2.8

#水准尺与仪器高 水准尺设置为2m尺长 读数0.3~1.8 仪器高度0.6~1.7

#读数为读基辅（红黑） 基辅差为301550 读数相差为0.2mm 读数精度+-0.01mm


#非固定点高程波动  （表示除监测点外高程具有波动性）

#平差选择：固定的闭合差【3.6】 和 变化闭合差【-1~1】
#测站定权或者距离定权
#高差为平差后的高差

#观测顺序BBFF 或者aBBFF


#开始观测08：00：00 测站间隔：30左右 测尺间隔6s左右 需要加入人为因素

#测线号：1-10

#读书重复：1

#加入随机错误 出错概率：1/20

