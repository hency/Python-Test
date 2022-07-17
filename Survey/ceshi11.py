import openpyxl
from openpyxl import load_workbook
path="D:\\2022\\主体沉降\\润永通\\测线 - 副本.xlsx"
book1=load_workbook(path)
sheet1=book1.get_sheet_by_name('测线')
print(sheet1.cell(3,3).value)

import pandas as pd
pd1=pd.read_excel(path,'测线')
