import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
import xlsxwriter
from openpyxl import load_workbook
path="D:\\2021\\基坑监测\\赣电中心\\赣电中心数据库-报表\\3监测日报.xls"
path1="D:\\2021\\基坑监测\\赣电中心\\赣电中心数据库-报表\\"
pathce="D:\\2021\\基坑监测\\华侨城万科4-2地块\\6-1地块计算表.xlsx"
workbook = xlrd.open_workbook(path)
sheet = workbook.sheet_by_index(3)
text=sheet.cell(0,0).value
re_text='附表1：'+text
workbook1=copy(workbook)
workbook1.save(path1+'2.xls')
# sheet1=workbook1.sheet_by_name('周边道路沉降BFZ(1)')
# sheet1.write(1,1,re_text)
# workbook1.save('2.xls')
# workbook1.close()

# wb = xlsxwriter.Workbook(path1+"test.xlsx")
# sheet1 = wb.add_worksheet()
# for row in range(sheet.nrows):
#     for col in range(sheet.ncols):
#         sheet1.write(row, col, sheet.cell(row,col).value)
# wb.close()
# sheet1=workbook1.index()
# sheet1=workbook1.get_sheet_by_name('周边道路沉降BFZ(1)')
# sheet1.cell(row=1,column=1).value=re_text
# workbook1.save(path1+'1.xlsx')