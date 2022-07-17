###没有阴影部分的，没有监测点重复的，提取CW沉降观测值
import time
import openpyxl
import pandas as pd
import os
import numpy
import xlwt
import xlrd
from datetime import datetime
# path="D:\\2021\\基坑监测\\巷口\\报告"
path="D:\\2021\\基坑监测\\高新安置\\日报"
####遍历PATH文件下所有的EXCEL文件
subpath1_names=os.listdir(path)
subpath2_names=[]
subpath3_names=[]
for name in subpath1_names:
    if(name[-4:]=='xlsm'):
        subpath3_names.append(name)
        subpath2_names.append(path+"\\"+name)
    else:
        pass
for i in range(0,len(subpath3_names)):
    if('~' in subpath3_names[i]):
        print(subpath3_names[i])
# for i in range(0,len(subpath3_names)):
#     if('2020.12.2~23 - 巷口祥和嘉园监测阶段报告.xlsm' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath3x_names=subpath3_names[0:144] ###########文件的个数
# subpath3_names=[]
# subpath3_names=subpath3x_names
# print("+++++++++++++++++++++++++++++++++++++++")
# for i in range(0,len(subpath3_names)):
#     if('~' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath2x_names=subpath2_names[0:144]
# subpath2_names=[]
# subpath2_names=subpath2x_names
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def bubble_sort(nums,nums1,DATE2):
    for i in range(len(nums) - 1):  # 这个循环负责设置冒泡排序进行的次数
        for j in range(len(nums) - i - 1):  # j为列表下标
            if nums[j] > nums[j + 1]:
                nums[j], nums[j + 1] = nums[j + 1], nums[j]
                nums1[j],nums1[j+1]=nums1[j+1],nums1[j]
                DATE2[j], DATE2[j + 1] = DATE2[j + 1], DATE2[j]
                # temp=nums[j]
                # nums[j]=nums[j+1]
                # nums[j+1]=temp
    return nums,nums1,DATE2
DATE=[]
DATE1=[]
# subpath3_names[144]='2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm'
for i in range(0,len(subpath3_names)):
    for j in range(1,len(subpath3_names[i])):
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]==' '):
            num1=j
            date=subpath3_names[i][0:j-1]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]!=' '):
            date=subpath3_names[i][0:j]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
####根据时间戳进行排序
# subpath2_names[144]='D:\\2021\\基坑监测\\巷口\\报告\\2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm' #####特殊情况说明：出现了~$符号
bubble_sort(DATE,subpath2_names,DATE1)
name1=path+"\\"+subpath2_names[0]
book1 = openpyxl.Workbook()
sheet1 = book1.create_sheet('CW_CJ')
sheet2=book1.create_sheet('CW_WY')
sheet3=book1.create_sheet('date')
def get_CW_value(name1):
    pd2 = []
    cw_name1 = []
    pd1 = pd.read_excel(name1)
    cols = pd1.columns
    numm=0
    index1=[]
    start_index1=[]
    end_index1=[]
    for i in range(len(cols)):
        if("桩顶竖向位移监测日报表" in cols[i]):#边坡顶部
            numm=numm+1
    for i in range(len(cols)):
        if ("桩顶竖向位移监测日报表" in cols[i]):
            index1.append( i + 2 )
            for j in range(pd1.shape[0]):
                if(isinstance(pd1.iloc[j,i],float)):
                    pd1.iloc[j,i]=str(pd1.iloc[j,i])
                if( 'CW' in pd1.iloc[j,i]):
                    start_index1.append(j)
                    start_index_1=j
                    break
            for j in range(start_index_1, pd1.shape[0]):
                if (isinstance(pd1.iloc[j, i], float)):
                    pd1.iloc[j,i]=str(pd1.iloc[j,i])
            for j in range(start_index_1,pd1.shape[0]):
                if ('CW' in pd1.iloc[j, i] and '注' in pd1.iloc[j + 1, i]):
                    end_index1.append(j)
                    end_index_1=j
                    break
                if 'CW' not in pd1.iloc[j, i]:
                    end_index1.append(j - 1)
                    end_index_1=j-1
                    break
            pd2.append(pd1.iloc[start_index_1:end_index_1+1,i+3])
            cw_name1.append(pd1.iloc[start_index_1:end_index_1+1,i])
    pd_concat=[]
    cw_concat=[]
    if(numm==1):
        pd_concat=pd2[0]
        cw_concat=cw_name1[0]
    elif(numm==2 or numm>2):
        pd_concat=pd.concat([pd2[0],pd2[1]],axis=0)
        cw_concat=pd.concat([cw_name1[0],cw_name1[1]],axis=0)
    else:
        pass
    if(numm>2):
        for i in range(numm-2):
            pd_concat=pd.concat([pd_concat,pd2[2+i]],axis=0)###
            cw_concat=pd.concat([cw_concat,cw_name1[2+i]],axis=0)
    return pd_concat,cw_concat
# pd_concat,cw_concat=get_CW_value(subpath2_names[0])
for j in range(1,77):
    sheet1.cell(1,j).value='CW'+str(j) ###根据实际的CW个数进行赋值这里是CW1-CW76
k=2
for j in range(len(subpath2_names)):   #len(subpath2_names)len(subpath2_names)
    pd_concat=[]
    cw_concat=[]
    pd_concat,cw_concat=get_CW_value(subpath2_names[j])
    # nn1=int(cw_concat1.iloc[0][2:])
    for i in range(0, pd_concat.shape[0]):
        for z in range(1,77):
            if(sheet1.cell(1,z).value==cw_concat.iloc[i]):
                sheet1.cell(k,z).value=pd_concat.iloc[i]
                break
            else:
                pass
    k=k+1
    sheet3.cell(j+1,1).value=DATE1[j]
    print(j)
book1.save(path+"\\"+'CW_高新安置_2019.xlsx')
book1.close()
###############60存在问题
