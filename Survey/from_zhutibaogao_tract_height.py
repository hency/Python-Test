import pandas as pd
import openpyxl
import xlrd
import re
import math
import numpy as np
src_path="D:\\2022\\主体沉降\\润永通\\润永通（1#-3#、5#、7#-10#楼）数据.xlsx"
book1=openpyxl.load_workbook(src_path)
# book2=xlrd.open_workbook(src_path)
sheet_names=book1.sheetnames
sheet_name=sheet_names[7]#############################改动sheet的排列序号
pd1=pd.read_excel(src_path,sheet_name)
book2=openpyxl.Workbook()
sheet2=book2.create_sheet('data')
sheet3=book2.create_sheet('date')
date=[]
k1=1
for i in range(pd1.shape[1]):
    if(isinstance(pd1.iloc[1,i],float)):
        pd1.iloc[1,i]=str(pd1.iloc[1,i])
    if(('第' in pd1.iloc[1,i]) and ('次' in pd1.iloc[1,i]) ):
        if('(' in pd1.iloc[1,i]):
            date1=pd1.iloc[1,i]
            date2=re.findall(r'[(](.*)[)]',date1)
            sheet3.cell(k1,1).value=date2[0]
            date.append(date2[0])
            k1=k1+1
        if('（' in pd1.iloc[1,i]):
            date1=pd1.iloc[1,i]
            date2=re.findall(r'[（](.*)[）]',date1)
            sheet3.cell(k1,1).value=date2[0]
            date.append(date2[0])
            k1=k1+1
for i in range(3,pd1.shape[0]):
    if('CJ' not in pd1.iloc[i,0]):
        num_CJ=i-3
        break
height=[]
k=1
for i in range(pd1.shape[1]):
    height1=[]
    if(isinstance(pd1.iloc[2,i],float)):
        pd1.iloc[2,i]=str(pd1.iloc[2,i])
    if('本次高程' in pd1.iloc[2,i]):    ##################如果提取的是“本次高程就填写本次高程”；如果填写的是“累计变化量就填累计变化量”
        for j in range(3,3+num_CJ):
            height1.append(pd1.iloc[j,i])
            sheet2.cell(j-2,k).value=pd1.iloc[j,i]
        k=k+1
        height.append(height1)
out_path="D:\\2022\\主体沉降\\润永通\\" + sheet_name+'本次高程'+'.xlsx'
book2.save(out_path)


