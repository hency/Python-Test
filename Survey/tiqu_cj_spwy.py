import time
import openpyxl
import pandas as pd
import os
import numpy
import xlwt
import xlrd
from datetime import datetime
# path="D:\\2021\\基坑监测\\巷口\\报告"
path="D:\\2021\\基坑监测\\高新安置\\日报1"
####遍历PATH文件下所有的EXCEL文件
subpath1_names=os.listdir(path)
subpath2_names=[]
subpath3_names=[]
for name in subpath1_names:
    if(name[-4:]=='xlsm'):
        subpath3_names.append(name)
        subpath2_names.append(path+"\\"+name)
    else:
        pass
for i in range(0,len(subpath3_names)):
    if('~' in subpath3_names[i]):
        print(subpath3_names[i])
# for i in range(0,len(subpath3_names)):
#     if('2020.12.2~23 - 巷口祥和嘉园监测阶段报告.xlsm' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath3x_names=subpath3_names[0:144] ###########文件的个数
# subpath3_names=[]
# subpath3_names=subpath3x_names
# print("+++++++++++++++++++++++++++++++++++++++")
# for i in range(0,len(subpath3_names)):
#     if('~' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath2x_names=subpath2_names[0:144]
# subpath2_names=[]
# subpath2_names=subpath2x_names
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def bubble_sort(nums,nums1,DATE2):
    for i in range(len(nums) - 1):  # 这个循环负责设置冒泡排序进行的次数
        for j in range(len(nums) - i - 1):  # j为列表下标
            if nums[j] > nums[j + 1]:
                nums[j], nums[j + 1] = nums[j + 1], nums[j]
                nums1[j],nums1[j+1]=nums1[j+1],nums1[j]
                DATE2[j], DATE2[j + 1] = DATE2[j + 1], DATE2[j]
                # temp=nums[j]
                # nums[j]=nums[j+1]
                # nums[j+1]=temp
    return nums,nums1,DATE2
DATE=[]
DATE1=[]
# subpath3_names[144]='2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm'
for i in range(0,len(subpath3_names)):
    for j in range(1,len(subpath3_names[i])):
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]==' '):
            num1=j
            date=subpath3_names[i][0:j-1]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]!=' '):
            date=subpath3_names[i][0:j]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
####根据时间戳进行排序
# subpath2_names[144]='D:\\2021\\基坑监测\\巷口\\报告\\2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm' #####特殊情况说明：出现了~$符号
bubble_sort(DATE,subpath2_names,DATE1)
name1=path+"\\"+subpath2_names[0]
book1 = openpyxl.Workbook()
sheet1 = book1.create_sheet('CW_CJ')
sheet2=book1.create_sheet('CW_WY')
sheet3=book1.create_sheet('date')
def get_CW_value(name1):
    pd1 = pd.read_excel(name1)
    cols = pd1.columns
    numm=0
    for i in range(len(cols)):
        if("周边地表、建筑竖向位移监测日报表" in cols[i]):#边坡顶部  周边地表、建筑竖向位移监测日报表 周边地表、建筑竖向位移监测日报表
            numm=numm+1
    if(numm==2):
        gc_name1 = "周边地表、建筑竖向位移监测日报表"
        gc_name2 = gc_name1 + '.1'
        gc_name3 = "桩顶水平位移监测日报表"
        gc_name4 = gc_name3 + ".1"
        for i in range(0, len(cols)):
            if (cols[i] == gc_name1):
                index1 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name2):
                index2 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name3):
                index3 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name4):
                index4 = i + 2
                break
        #########################################边坡顶部竖向位移监测点日报表
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
            if 'JGC' in pd1.iloc[i,index1-2]:
                start_index1=i
                break
            else:
                pass
        for i in range(start_index1,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
        for i in range(start_index1, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index1-2] and '注' in pd1.iloc[i+1,index1-2]):
                end_index1=i
                break
            if 'JGC' not in pd1.iloc[i,index1-2]:
                end_index1=i-1
                break
        #######################################边坡顶部竖向位移监测点日报表.1
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
            if 'JGC' in pd1.iloc[i,index2-2]:
                start_index2=i
                break
            else:
                pass
        for i in range(start_index2,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
        for i in range(start_index2, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index2-2] and '注' in pd1.iloc[i+1,index2-2]):
                end_index2=i
                break
            if 'JGC' not in pd1.iloc[i,index2-2]:
                end_index2=i-1
                break
        #######################################边坡顶部水平位移监测日报表
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index3-2],float)):
                pd1.iloc[i,index3-2]=str(pd1.iloc[i,index3-2])
            if 'CW' in pd1.iloc[i,index3-2]:
                start_index3=i
                break
            else:
                pass
        for i in range(start_index3,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index3-2],float)):
                pd1.iloc[i,index3-2]=str(pd1.iloc[i,index3-2])
        for i in range(start_index3, pd1.shape[0]):
            if('CW' in pd1.iloc[i,index3-2] and '注' in pd1.iloc[i+1,index3-2]):
                end_index3=i
                break
            if 'CW' not in pd1.iloc[i,index3-2]:
                end_index3=i-1
                break
        ###########################################  边坡顶部水平位移监测日报表.1
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index4-2],float)):
                pd1.iloc[i,index4-2]=str(pd1.iloc[i,index4-2])
            if 'CW' in pd1.iloc[i,index4-2]:
                start_index4=i
                break
            else:
                pass
        for i in range(start_index4,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index4-2],float)):
                pd1.iloc[i,index4-2]=str(pd1.iloc[i,index4-2])
        for i in range(start_index4, pd1.shape[0]):
            if('CW' in pd1.iloc[i,index4-2] and '注' in pd1.iloc[i+1,index4-2]):
                end_index4=i
                break
            if 'CW' not in pd1.iloc[i,index4-2]:
                end_index4=i-1
                break
        pd2 = pd1.iloc[start_index1:end_index1+1, index1+1]
        pd3 = pd1.iloc[start_index2:end_index2+1, index2+1]
        pd4 = pd1.iloc[start_index3:end_index3+1, index3+1]
        pd5 = pd1.iloc[start_index4:end_index4+1, index4+1]
        cw_name1=pd1.iloc[start_index1:end_index1+1, index1-2]
        cw_name2=pd1.iloc[start_index2:end_index2+1, index2-2]
        cw_name3=pd1.iloc[start_index3:end_index3+1, index3-2]
        cw_name4=pd1.iloc[start_index4:end_index4+1, index4-2]
        pd_concat1 = pd.concat([pd2, pd3], axis=0)
        pd_concat2 = pd.concat([pd4, pd5], axis=0)
        cw_concat1=pd.concat([cw_name1,cw_name2],axis=0)
        cw_concat2=pd.concat([cw_name3,cw_name4],axis=0)
        # return pd_concat1,pd_concat2,cw_concat1,cw_concat2
        return pd_concat1, cw_concat1
    elif(numm==3):
        gc_name1 = "周边地表、建筑竖向位移监测日报表"
        gc_name2 = gc_name1 + '.1'
        gc_name3 = gc_name1 + '.2'
        gc_name4 = "桩顶水平位移监测日报表"
        gc_name5 = gc_name4 + ".1"
        gc_name6 = gc_name4 + ".2"
        for i in range(0, len(cols)):
            if (cols[i] == gc_name1):
                index1 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name2):
                index2 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name3):
                index3 = i + 2
                break
        # for i in range(0, len(cols)):
        #     if (cols[i] == gc_name4):
        #         index4 = i + 2
        #         break
        # for i in range(0, len(cols)):
        #     if (cols[i] == gc_name5):
        #         index5 = i + 2
        #         break
        # for i in range(0, len(cols)):
        #     if (cols[i] == gc_name6):
        #         index6 = i + 2
        #         break
        #########################################边坡顶部竖向位移监测点日报表
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
            if 'JGC' in pd1.iloc[i,index1-2]:
                start_index1=i
                break
            else:
                pass
        for i in range(start_index1,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
        for i in range(start_index1, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index1-2] and '注' in pd1.iloc[i+1,index1-2]):
                end_index1=i
                break
            if 'JGC' not in pd1.iloc[i,index1-2]:
                end_index1=i-1
                break
        #######################################边坡顶部竖向位移监测点日报表.1
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
            if 'JGC' in pd1.iloc[i,index2-2]:
                start_index2=i
                break
            else:
                pass
        for i in range(start_index2,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
        for i in range(start_index2, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index2-2] and '注' in pd1.iloc[i+1,index2-2]):
                end_index2=i
                break
            if 'JGC' not in pd1.iloc[i,index2-2]:
                end_index2=i-1
                break
        #######################################边坡顶部竖向位移监测点日报表.2
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index3-2],float)):
                pd1.iloc[i,index3-2]=str(pd1.iloc[i,index3-2])
            if 'JGC' in pd1.iloc[i,index3-2]:
                start_index3=i
                break
            else:
                pass
        for i in range(start_index3,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index3-2],float)):
                pd1.iloc[i,index3-2]=str(pd1.iloc[i,index3-2])
        for i in range(start_index3, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index3-2] and '注' in pd1.iloc[i+1,index3-2]):
                end_index3=i
                break
            if 'JGC' not in pd1.iloc[i,index3-2]:
                end_index3=i-1
                break
        ###########################################边坡顶部水平位移监测日报表
        # for i in range(pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index4-2],float)):
        #         pd1.iloc[i,index4-2]=str(pd1.iloc[i,index4-2])
        #     if 'CW' in pd1.iloc[i,index4-2]:
        #         start_index4=i
        #         break
        #     else:
        #         pass
        # for i in range(start_index4,pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index4-2],float)):
        #         pd1.iloc[i,index4-2]=str(pd1.iloc[i,index4-2])
        # for i in range(start_index4, pd1.shape[0]):
        #     if('CW' in pd1.iloc[i,index4-2] and '注' in pd1.iloc[i+1,index4-2]):
        #         end_index4=i
        #         break
        #     if 'CW' not in pd1.iloc[i,index4-2]:
        #         end_index4=i-1
        #         break
        # ##########################################边坡顶部水平位移监测日报表.1
        # for i in range(pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index5-2],float)):
        #         pd1.iloc[i,index5-2]=str(pd1.iloc[i,index5-2])
        #     if 'CW' in pd1.iloc[i,index5-2]:
        #         start_index5=i
        #         break
        #     else:
        #         pass
        # for i in range(start_index5,pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index5-2],float)):
        #         pd1.iloc[i,index5-2]=str(pd1.iloc[i,index5-2])
        # for i in range(start_index5, pd1.shape[0]):
        #     if('CW' in pd1.iloc[i,index5-2] and '注' in pd1.iloc[i+1,index5-2]):
        #         end_index5=i
        #         break
        #     if 'CW' not in pd1.iloc[i,index5-2]:
        #         end_index5=i-1
        #         break
        # ##########################################边坡顶部水平位移监测日报表.2
        # for i in range(pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index6-2],float)):
        #         pd1.iloc[i,index6-2]=str(pd1.iloc[i,index6-2])
        #     if 'CW' in pd1.iloc[i,index6-2]:
        #         start_index6=i
        #         break
        #     else:
        #         pass
        # for i in range(start_index6,pd1.shape[0]):
        #     if(isinstance(pd1.iloc[i,index6-2],float)):
        #         pd1.iloc[i,index6-2]=str(pd1.iloc[i,index6-2])
        # for i in range(start_index6, pd1.shape[0]):
        #     if('CW' in pd1.iloc[i,index6-2] and '注' in pd1.iloc[i+1,index6-2]):
        #         end_index6=i
        #         break
        #     if 'CW' not in pd1.iloc[i,index6-2]:
        #         end_index6=i-1
        #         break
        ##########################################
        pd2 = pd1.iloc[start_index1:end_index1+1, index1+1]
        pd3 = pd1.iloc[start_index2:end_index2+1, index2+1]
        pd4 = pd1.iloc[start_index3:end_index3+1, index3+1]
        # pd5 = pd1.iloc[start_index4:end_index4+1, index4+1]
        # pd6 = pd1.iloc[start_index5:end_index5+1, index5+1]
        # pd7 = pd1.iloc[start_index6:end_index6+1, index6+1]
        cw_name1=pd1.iloc[start_index1:end_index1+1, index1-2]
        cw_name2=pd1.iloc[start_index2:end_index2+1, index2-2]
        cw_name3=pd1.iloc[start_index3:end_index3+1, index3-2]
        # cw_name4=pd1.iloc[start_index4:end_index4+1, index4-2]
        # cw_name5=pd1.iloc[start_index5:end_index5+1, index5-2]
        # cw_name6=pd1.iloc[start_index6:end_index6+1, index6-2]
        pd_concat1 = pd.concat([pd2, pd3,pd4], axis=0)
        # pd_concat2 = pd.concat([pd4, pd5,pd6], axis=0)
        cw_concat1=pd.concat([cw_name1,cw_name2],axis=0)
        # cw_concat2=pd.concat([cw_name3,cw_name4],axis=0)
        # return pd_concat1,pd_concat2,cw_concat1,cw_concat2
        return pd_concat1, cw_concat1
    elif(numm==1):
        gc_name1 = "周边地表、建筑竖向位移监测日报表"
        gc_name2 = "桩顶水平位移监测日报表"
        for i in range(0, len(cols)):
            if (cols[i] == gc_name1):
                index1 = i + 2
                break
        for i in range(0, len(cols)):
            if (cols[i] == gc_name2):
                index2 = i + 2
                break
        #########################################边坡顶部竖向位移监测点日报表
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
            if 'JGC' in pd1.iloc[i,index1-2]:
                start_index1=i
                break
            else:
                pass
        for i in range(start_index1,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index1-2],float)):
                pd1.iloc[i,index1-2]=str(pd1.iloc[i,index1-2])
        for i in range(start_index1, pd1.shape[0]):
            if('JGC' in pd1.iloc[i,index1-2] and '注' in pd1.iloc[i+1,index1-2]):
                end_index1=i
                break
            if 'JGC' not in pd1.iloc[i,index1-2]:
                end_index1=i-1
                break
        #######################################边坡顶部水平位移监测日报表
        for i in range(pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
            if 'CW' in pd1.iloc[i,index2-2]:
                start_index2=i
                break
            else:
                pass
        for i in range(start_index2,pd1.shape[0]):
            if(isinstance(pd1.iloc[i,index2-2],float)):
                pd1.iloc[i,index2-2]=str(pd1.iloc[i,index2-2])
        for i in range(start_index2, pd1.shape[0]):
            if('CW' in pd1.iloc[i,index2-2] and '注' in pd1.iloc[i+1,index2-2]):
                end_index2=i
                break
            if 'CW' not in pd1.iloc[i,index2-2]:
                end_index2=i-1
                break
        pd2 = pd1.iloc[start_index1:end_index1+1, index1+1]
        pd3 = pd1.iloc[start_index2:end_index2+1, index2+1]
        cw_name1=pd1.iloc[start_index1:end_index1+1, index1-2]
        cw_name2=pd1.iloc[start_index2:end_index2+1, index2-2]
        pd_concat1 = pd2
        pd_concat2 = pd3
        cw_concat1=cw_name1
        cw_concat2=cw_name2
        return pd_concat1,pd_concat2,cw_concat1,cw_concat2
k=2
for j in range(0,len(subpath2_names)):   #len(subpath2_names)
    pd_concat1=[]
    # pd_concat2=[]
    # pd_concat1,pd_concat2,cw_concat1,cw_concat2=get_CW_value(subpath2_names[j])
    pd_concat1,  cw_concat1 = get_CW_value(subpath2_names[j])
    # nn1=int(cw_concat1.iloc[0][2:])
    for i in range(0, pd_concat1.shape[0]):
        sheet1.cell(1, i+1).value =cw_concat1.iloc[i]                   #'CW' + str(i)
        sheet1.cell(k, i+1).value = pd_concat1.iloc[i]
        # sheet2.cell(1, i+1).value =cw_concat2.iloc[i]                   #'CW' + str(i)
        # sheet2.cell(k, i+1).value = pd_concat2.iloc[i]
    k=k+1
    sheet3.cell(j+1,1).value=DATE1[j]
    print(j)
book1.save(path+"\\"+'CW_高新安置_2019.xlsx')
book1.close()
###############60存在问题
