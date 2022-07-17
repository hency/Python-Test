import pandas
import openpyxl
from openpyxl import load_workbook
import os
import re
import numpy
import time
from datetime import datetime


path="D:\\2021\\基坑监测\\高新安置\\日报"
####遍历PATH文件下所有的EXCEL文件
subpath1_names=os.listdir(path)
subpath2_names=[]
subpath3_names=[]
for name in subpath1_names:
    if(name[-4:]=='xlsm'):
        subpath3_names.append(name)
        subpath2_names.append(path+"\\"+name)
    else:
        pass
for i in range(0,len(subpath3_names)):
    if('~' in subpath3_names[i]):
        print(subpath3_names[i])
# for i in range(0,len(subpath3_names)):
#     if('2020.12.2~23 - 巷口祥和嘉园监测阶段报告.xlsm' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath3x_names=subpath3_names[0:144] ###########文件的个数
# subpath3_names=[]
# subpath3_names=subpath3x_names
# print("+++++++++++++++++++++++++++++++++++++++")
# for i in range(0,len(subpath3_names)):
#     if('~' in subpath3_names[i]):
#         print(subpath3_names[i])
# subpath2x_names=subpath2_names[0:144]
# subpath2_names=[]
# subpath2_names=subpath2x_names
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def parse_ymd(s):
    year_s, mon_s, day_s = s.split('.')
    return datetime(int(year_s), int(mon_s), int(day_s))
def bubble_sort(nums,nums1,DATE2):
    for i in range(len(nums) - 1):  # 这个循环负责设置冒泡排序进行的次数
        for j in range(len(nums) - i - 1):  # j为列表下标
            if nums[j] > nums[j + 1]:
                nums[j], nums[j + 1] = nums[j + 1], nums[j]
                nums1[j],nums1[j+1]=nums1[j+1],nums1[j]
                DATE2[j], DATE2[j + 1] = DATE2[j + 1], DATE2[j]
                # temp=nums[j]
                # nums[j]=nums[j+1]
                # nums[j+1]=temp
    return nums,nums1,DATE2
DATE=[]
DATE1=[]
# subpath3_names[144]='2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm'
for i in range(0,len(subpath3_names)):
    for j in range(1,len(subpath3_names[i])):
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]==' '):
            num1=j
            date=subpath3_names[i][0:j-1]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
        if(subpath3_names[i][j]=='-' and subpath3_names[i][j-1]!=' '):
            date=subpath3_names[i][0:j]
            date1=time.strptime(date,"%Y.%m.%d")
            date3=int(time.mktime(date1))
            DATE1.append(parse_ymd(date))
            DATE.append(date3)
            break
####根据时间戳进行排序
# subpath2_names[144]='D:\\2021\\基坑监测\\巷口\\报告\\2018.6.2 - 巷口祥和嘉园监测成果表（报告）.xlsm' #####特殊情况说明：出现了~$符号
bubble_sort(DATE,subpath2_names,DATE1)
book=openpyxl.Workbook()
sheet=book.create_sheet('周边环境沉降')
path_save="D:\\2021\\基坑监测\\高新安置\\日报\\1.xlsx"
path__1="D:\\2021\\基坑监测\\高新安置\\日报\\"
name='周边地表、建筑竖向位移监测日报表'
k=1

for pp in range(len(subpath2_names)):
    path=subpath2_names[pp]
    data=pandas.read_excel(path,'监测日报表')
    title=data.columns
    for i in range(len(title)):
        if(isinstance(title[i],float)):
            continue
        else:
            if(name in title[i]):
                for j in range(4,data.shape[0]):
                    if(isinstance(data.iloc[33,i],float)):
                        if(~isinstance(data.iloc[j,i],float) and isinstance(data.iloc[j+1,i],float)):
                            sheet.cell(int(re.findall('JGC(.*)', data.iloc[j, i], flags=0)[0]), k).value = data.iloc[j, i+3]
                            break
                        else:
                            sheet.cell(int(re.findall('JGC(.*)',data.iloc[j,i],flags=0)[0]),k).value=data.iloc[j,i+3]
                    else:
                        if(('JGC' in data.iloc[j,i]) and ('注' not in data.iloc[j,i])):
                            sheet.cell(int(re.findall(r'JGC(.*)',data.iloc[j,i],flags=0)[0]),k).value=data.iloc[j,i+3]
                        else:
                            break
    k=k+1
book.save(path_save)
