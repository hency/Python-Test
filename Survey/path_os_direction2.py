# -*- coding: utf-8 -*-
##还需改进
import os
import shutil
from openpyxl import load_workbook


# path1="D:\\Desktop\\测x\\" #目标文件路径
# path3="E:\\中航城三期主体沉降观测--数据库--测线"   #源文件路径
# pathx1="E:\\中航城三期主体沉降观测--数据库--测线\\" #源文件路径同上
path1="D:\\2022\\主体沉降\\润永通\\数据库\\"  #目标文件路径
path3="D:\\2022\\主体沉降\\润永通\\数据库\\"   #源文件路径
a=[1,2,3,5,7,8,9,10] #楼栋号
# a=['S1','S2','S3','S4','S5','S6']
for i in range(0,len(a)):
    os.makedirs(path1+str(a[i])+'#') #建立**#
    # os.makedirs(path1 + a[i] + '#')  # 建立**#
    path2=path1+str(a[i])+'#'+'\\'
    # path2 = path1 + a[i] + '#' + '\\'
    file_names=os.listdir(path3)
    workbook = load_workbook(path1+str(a[i])+"#数据库.xlsx")
    # workbook = load_workbook(path1 + a[i] + "#数据库.xlsx")
    # book1=xlrd.open_workbook(pathx1+str(a[i])+"#数据库.xlsx")
    sheet1=workbook.get_sheet_by_name("日报")
    # sheet1=book1.sheet_by_name('日报')
    for zz in range(2,sheet1.max_row+1):
        if((sheet1.cell(zz,2).value is None) or (sheet1.cell(zz,2).value=='')):#根据判断是否为无参数None或者空来判断行数
            row2=zz-1
            break
        else:
            row2=sheet1.max_row #根据日报获取期数
    # row1=sheet1.max_row
    for j in range(2,row2+1):
        date1=sheet1.cell(j,2).value
        date3="第"+str(j-1)+"期"+str(date1)
        date4=path1+str(a[i])+'#'+'\\'+date3
        x1=date3.replace('-','年',1)
        x2=x1.replace('-','月',1)
        x3=x2.replace(' 00:00:00','日',1) #通过将日期datetime格式转换成字符串的形式将对应的2018/9/20 00：00：00转换成汉字2018年9月20日
        datex2 = path1 + str(a[i]) + '#' + '\\' + x3
        # datex2 = path1 + a[i] + '#' + '\\' + x3
        os.makedirs(datex2) #创建各栋楼的各自的期数文件夹
        for name in file_names:
            path4 = path1 + name
            if(name[-3:]=='dat'):
                y1=name.index('#')
                y2=name.index('降')
                y3=name.index('.')
                if (name[0:y1] == (str(a[i]))):
                # if (name[0:2] == (a[i])):
                    if(str(j-1)==name[y2+1:y3]):
                        shutil.copy(path4, datex2) #将通过遍历源目标文件下的dat文件拷贝到对应的目标文件夹下

##下面的代码是判断源文件下的文件属于文件夹还是文件，进而进行目标文件的剪切
# def isFile(folder_name=""):
#     file_names = os.listdir(folder_name)
#     for name in file_names:
#         if os.path.isfile(folder_name + "\\" + name):
#             return True
#
#     return False
#
#
# '''
# folder_name:文件夹名称，D:\\xxx\\xxx\\xx
# move_file : 移入的文件夹名称
# n:向文件夹移入的文件数量
# '''
#
#
# def fileMove(folder_name="", move_file="", n=0):
#     j = 0
#     while isFile(folder_name):
#         j += 1
#         dir = move_file + str(j)
#         dirpath = folder_name + "\\" + dir
#         os.makedirs(dirpath)
#         file_names = os.listdir(folder_name)
#         i = 0
#         for name in file_names:
#
#             if move_file in name:
#                 pass
#             else:
#                 try:
#                     if i >= n:
#                         break
#                     file_name = folder_name + "\\" + name
#                     shutil.move(file_name, dirpath)
#                     print(name + '=》' + dirpath)
#                     i += 1
#                 except BaseException as e:
#                     print(e)
#                     print(name)
#                     with open('log.txt', 'a+', encoding='utf-8') as f:
#                         f.write(name + '\n')
#
#
# if __name__ == '__main__':
#     # 需要处理的文件夹路径
#     folder_name = r'D:\xxx\xxx\xx'
#     # 文件名
#     move_file = 'tmp'
#     # 控制文件放入文件夹个数
#     n = 35
#     fileMove(folder_name, move_file, n)
# path1='\\Clgs-aoc\工作\\2018\\沉降\\CL2018-118海德公园\\新建文件夹\\原始数据1\\'