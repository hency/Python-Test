import numpy
import pandas as pd
import numpy as np
import xlrd
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from 沉降观测1 import read_measure_line_from_dataset
path1="D:\\Desktop\\测试期次1\\output1.xlsx"
def mid(x,y):
    a=x[0]+(y[0]-x[0])/2
    b=x[1]+(y[1]-x[1])/2
    return [a,b]
path="D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"
df=pd.read_excel(path,'Sheet1')#对应的导出为pd.to_excel(***.xlsx)
df.columns=range(0,3)
name=['Y1-a1','Y1-a2','Y1-a3']
value=[19.6425,19.7215,19.7715]
dict = {name[i]:value[i] for i in range(3)}
for i in range(0,df.shape[0]):
    if(df.loc[i,0] in name):
        df.loc[i,1]=dict[df.loc[i,0]]
df1=df[:1]
df2=df[1:len(df)]
df21=df2.dropna(how='any')
df3=pd.concat([df1,df21],axis=0)
df3.index=range(df3.shape[0])
a=df3.iloc[:,[1]]
b=range(df3.shape[0])
c=a
num=[]
numx=[]
for i in range(df3.shape[0]):
    if(df3.loc[i,0] in name):
        numx.append(i)
        pass
    else:
        num.append(i)
d=df3.iloc[num,[1]]
num2=[]
for i in d.index:
    num2.append([i,d.loc[i,d.columns]])
# print(num2)
num3=[]
for i in range(len(num)-1):
    num3.append(mid(num2[i],num2[i+1]))
fig=pyplot.figure()
ax=fig.add_subplot(111)
ax.plot(numx,df3.loc[numx,1],'b^',markersize=12)
ax.plot(num,df3.loc[num,1],'ro',markersize=12)
ax.plot(d,color='black', linewidth=2.0)
for i in range(len(num3)):
    ax.plot(num3[i][0],num3[i][1],'bX',markersize=12)
for i in range(len(numx)):
    ax.plot([numx[i],num3[1][0]],[df3.loc[numx[i],1],num3[1][1]],color='g', linewidth=2.0)
# pyplot.show()
# pyplot.pause(100)

##仪器高1.4-1.6m
##目标高0.55-1.8m
##读数范围为如果地势平坦易设置在仪器高附近且不能超目标高范围
##基辅值差0.3mm
##高差之差0.6mm
##闭合差60mm
##改进值为60mm/6=10mm
##前后视距之差为1.5m
##累计视距之差为6m
instrument_height=1.4
Sight_Height_random1=[]
Sight_Height_random2=[]
Sight_Height_random3=[]
Sight_Height_random4=[]
for i in range(6):
    Sight_Height_random1.append(random.randint(-5, 15)*0.00001)##单位m
    # Sight_Height_random1.append(random.uniform(0.00005,0.00015))
df4=df3.loc[num,[0,1,2]]
def sight_height_distance(cx_df4,book1,sheet1,BC1):
    sheet1.cell(1,1).value='点位'
    sheet1.cell(1,3).value='时间'
    sheet1.cell(1,5).value='视线高'
    sheet1.cell(1,7).value='视距'
    sheet1.cell(1,9).value='高程'
    sheet1.cell(2,1).value='KZ1'
    sheet1.cell(2,9).value=df4.iloc[0,1]
    k=3
    for i in range(cx_df4.shape[0]-1):
        dh=cx_df4.iloc[i+1,1]-cx_df4.iloc[i,1]-BC1
        if(dh>1.25):
            print("高差大于1.25m，高差过大")
            xx1=1/0
        height1=cx_df4.iloc[i+1,1]-(i+1)*BC1
        sheet1.cell(7+5*i,9).value=height1
        #'高差值最大是1.8-0.55=1.25m' 且 dh_random的取值范围为0.55+dh~1.8
        dh1_random=dh+Sight_Height_random1[i]
        dh2=2*dh-dh1_random
        if (dh > 0):
            RH1_random = round(random.uniform(0.6 + dh1_random,1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1 = RH1_random - dh1_random
            RH2_random = RH1_random + Sight_Height_random1[i]
            FH2 = RH2_random - dh2
            if ((abs(RH1_random - RH2_random) < 0.3 and abs(FH1 - FH2) < 0.3) and abs(dh1_random - dh2) < 0.6):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
            else:
                for z1 in range(100000):
                    dh1_random = dh + Sight_Height_random1[i]
                    dh2 = 2 * dh - dh1_random
                    if (dh > 0):
                        RH1_random = round(random.uniform(0.6 + dh1_random, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    else:
                        RH1_random = round(random.uniform(0.6 + abs(dh1_random), 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1 = RH1_random - dh1_random
                    RH2_random = RH1_random + Sight_Height_random1[i]
                    FH2 = RH2_random - dh2
                    if ((abs(RH1_random - RH2_random) < 0.3 and abs(FH1 - FH2) < 0.3) and abs(dh1_random - dh2) < 0.6):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        break
                print("重新选择")
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random)
                sheet1.cell(k + 1, 5, FH1)
                sheet1.cell(k + 2, 5, FH2)
                sheet1.cell(k + 3, 5, RH2_random)
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, FH1)
                sheet1.cell(k + 1, 5, RH1_random)
                sheet1.cell(k + 2, 5, RH2_random)
                sheet1.cell(k + 3, 5, FH2)
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k = k + 5
        else:
            FH1_random = round(random.uniform(0.6 + abs(dh1_random),1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            RH1 = FH1_random - abs(dh1_random)
            FH2_random = FH1_random + Sight_Height_random1[i]
            RH2 = FH2_random - abs(dh2)
            if ((abs(RH1 - RH2) < 0.3 and abs(FH1_random - FH2_random) < 0.3) and abs(dh1_random - dh2) < 0.6):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
            else:
                for z1 in range(100000):
                    dh1_random = dh + Sight_Height_random1[i]
                    dh2 = 2 * dh - dh1_random
                    FH1_random = round(random.uniform(0.6 + abs(dh1_random), 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    RH1 = FH1_random - abs(dh1_random)
                    FH2_random = FH1_random + Sight_Height_random1[i]
                    RH2 = FH2_random - abs(dh2)
                    if ((abs(RH1 - RH2) < 0.3 and abs(FH1_random - FH2_random) < 0.3) and abs(dh1_random - dh2) < 0.6):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        break
                print("重新选择")
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1)
                sheet1.cell(k + 1, 5, FH1_random)
                sheet1.cell(k + 2, 5, FH2_random)
                sheet1.cell(k + 3, 5, RH2)
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, FH1_random)
                sheet1.cell(k + 1, 5, RH1)
                sheet1.cell(k + 2, 5, RH2)
                sheet1.cell(k + 3, 5, FH2_random)
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k=k+5
    HD_differ_sum=0
    k1=3
    for i in range(cx_df4.shape[0]-1):
        HD_random1 = round(random.uniform(-0.5, 0.5),3)##单位m
        HD_random2 = round(random.uniform(-0.5, 0.5),3)
        HD=cx_df4.iloc[i+1,2]/2
        HDb1=HD+HD_random1
        HDb2=HDb1+round(random.uniform(-0.005,0.005),3)
        HDf1=HD+HD_random2
        HDf2=HDf1+round(random.uniform(-0.005,0.005),3)
        HDb=(HDb1+HDb2)/2
        HDf=(HDf1+HDf2)/2
        HD_differ=HDb-HDf
        HD_differ_sum=HD_differ_sum+HD_differ
        if(HD_differ<1.5 and HD_differ_sum<6):
            print('前后视距满足要求')
        else:
            print('前后视距不满足要求')
            print('出现异常将在GUI中提现')
            Exception1=1/0
        if((i+1)%2==1):
            sheet1.cell(k1,7,HDb1)
            sheet1.cell(k1+1,7).value=HDf1
            sheet1.cell(k1+2,7,HDf2)
            sheet1.cell(k1+3,7).value=HDb2
            sheet1.cell(k1,8).value='HDB'
            sheet1.cell(k1+1,8).value='HDF'
            sheet1.cell(k1+2,8).value='HDF'
            sheet1.cell(k1+3,8).value='HDB'
            sheet1.cell(k1,1).value=cx_df4.iloc[0:cx_df4.shape[0],0].iloc[i]
            sheet1.cell(k1+1, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+2, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+3, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
        else:
            sheet1.cell(k1,7,HDf1)
            sheet1.cell(k1+1,7).value=HDb1
            sheet1.cell(k1+2,7,HDb2)
            sheet1.cell(k1+3,7).value=HDf2
            sheet1.cell(k1,8).value='HDF'
            sheet1.cell(k1+1,8).value='HDB'
            sheet1.cell(k1+2,8).value='HDB'
            sheet1.cell(k1+3,8).value='HDF'
            sheet1.cell(k1,1).value=cx_df4.iloc[0:cx_df4.shape[0],0].iloc[i+1]
            sheet1.cell(k1+1,1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+2, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+3, 1).value = cx_df4.iloc[0:cx_df4.shape[0], 0].iloc[i + 1]
        if(HD_differ_sum>6):
            print('****************************************************************************************************')
            print('前后视距累计差抄超限')
        k1=k1+5
    sheet1['A2'].font=Font(bold=True)
    for i in range(cx_df4.shape[0]-1):
        name1='A'+str(7+5*i)
        name2='A'+str(7+5*i)
        sheet1[name2]=cx_df4.iloc[i+1,0]
        sheet1[name1].font=Font(bold=True)
    book1.save(path1)
    book1.close()
dateset_path='D:\\Desktop\\测试期次1\\测试数据库1.xlsx'
measure_line_path="D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"
RZ_names,RZ_values=read_measure_line_from_dataset(dateset_path,measure_line_path)
numx1=[]
for i in range(df1.shape[0]):
    if(df1.iloc[i,1] in RZ_names):
        numx1.append(i)
flag=0
numx2=[]
for i in range(1,len(numx1)):
    if(numx1[i]-numx1[flag]==i-flag):
        if(i==len(numx1)-1):
            numx2.append(numx1[flag:i+1])
        pass
    else:
        if(i-flag==1):
            numx2.append(numx1[flag])
            if(i==len(numx1)-1):
                numx2.append(numx1[i])
        else:
            numx2.append(numx1[flag:flag+i-flag])
        flag=i
if(len(numx2)==1):
    if(isinstance(numx2[0],int)):
        index_up_name = df1.iloc[numx2[0] - 1, 0]
        index_low_name = df1.iloc[numx2[0] + 1, 0]
        for j in range(df1.shape[0]):
            if(index_up_name==df1.iloc[j,0]):
                index_up_num=j
            if(index_low_name==df1.iloc[j,0]):
                index_low_num=j
    else:
        index_up_name=df1.iloc[numx2[0]-1,0]
        index_low_name = df1.iloc[numx2[-1] + 1, 0]
        for j in range(df1.shape[0]):
            if (index_up_name == df1.iloc[j, 0]):
                index_up_num = j
            if (index_low_name == df1.iloc[j, 0]):
                index_low_num = j
else:
    for i in range(len(numx2)):
        nname1=numx2[i]
        if(isinstance(nname1,int)):
            index_up_name = df1.iloc[numx2[0] - 1, 0]
            index_low_name = df1.iloc[numx2[0] + 1, 0]
            for j in range(df1.shape[0]):
                if (index_up_name == df1.iloc[j, 0]):
                    index_up_num = j
                if (index_low_name == df1.iloc[j, 0]):
                    index_low_num = j
        else:
            index_up_name = df1.iloc[nname1[0] - 1, 0]
            index_low_name = df1.iloc[nname1[-1] + 1, 0]
            for j in range(df1.shape[0]):
                if (index_up_name == df1.iloc[j, 0]):
                    index_up_num = j
                if (index_low_name == df1.iloc[j, 0]):
                    index_low_num = j
book1=openpyxl.Workbook()
sheet=book1.create_sheet('Sheet2')
BC1=10/1000
sight_height_distance(df4,book1,sheet,BC1)











