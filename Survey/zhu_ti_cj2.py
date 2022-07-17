#coding=utf-8
import sys
from PyQt5.QtWidgets import QWidget,QPushButton,QVBoxLayout,QHBoxLayout,QGridLayout,QFormLayout,QLineEdit,QLabel,QMessageBox,QApplication,\
    QDialog,QAction,QDateTimeEdit
from PyQt5.QtCore import QDate, QTime, QDateTime
from 沉降观测1 import read_measure_line_from_dataset
import numpy
import pandas as pd
import numpy as np
import random
from matplotlib import pyplot
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from ce1_7 import sight_height_distance
import re
class zhu_ti_cj_jiang(QWidget):
    def __init__(self):
        super(zhu_ti_cj_jiang,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降')
        self.label1=QLabel(self)
        self.label1.setText('主体沉降原始数据')
        self.pushbutton1=QPushButton(self)
        self.pushbutton1.setText('原始数据')
        self.label2=QLabel(self)
        self.label2.setText('主体沉降日报表(不含曲线图)')
        self.pushbutton2=QPushButton(self)
        self.pushbutton2.setText('日报表1')
        self.label3=QLabel(self)
        self.label3.setText('主体沉降日报表(含曲线图)')
        self.pushbutton3=QPushButton(self)
        self.pushbutton3.setText('日报表2')
        self.label4=QLabel(self)
        self.label4.setText('提取闭合差')
        self.pushbutton4=QPushButton(self)
        self.pushbutton4.setText('闭合差')
        self.label5=QLabel(self)
        self.label5.setText('原始数据的放置')
        self.pushbutton5=QPushButton(self)
        self.pushbutton5.setText('原始数据放置')
        self.V_layout1 = QVBoxLayout()
        self.V_layout2 = QVBoxLayout()
        self.H_layout1 = QHBoxLayout()
        self.Layout_init()
    def Layout_init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout1.addWidget(self.label4)
        self.V_layout1.addWidget(self.label5)
        self.V_layout2.addWidget(self.pushbutton1)
        self.V_layout2.addWidget(self.pushbutton2)
        self.V_layout2.addWidget(self.pushbutton3)
        self.V_layout2.addWidget(self.pushbutton4)
        self.V_layout2.addWidget(self.pushbutton5)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.setLayout(self.H_layout1)
        self.pushbutton1.clicked.connect(self.enter_pushbutton1)
    def enter_pushbutton1(self):
        demo1=original()
        demo1.exec_()
class original(QDialog):
    def __init__(self):
        super(original,self).__init__()
        self.resize(600,300)
        self.setWindowTitle('主体沉降原始数据还原')
        self.label1=QLabel('测线路径',self)
        self.editline1=QLineEdit(self)
        self.label2=QLabel('数据库路径',self)
        self.editline2=QLineEdit(self)
        self.label3=QLabel('EXCEL输出路径',self)
        self.editline3=QLineEdit(self)
        self.label4=QLabel('输入总期数',self)
        self.editline4=QLineEdit(self)
        self.label5=QLabel('手动键入闭合差范围',self)
        self.editline5=QLineEdit(self)
        self.label6=QLabel('从日历选择日期',self)
        # self.editline5=QLineEdit(self)
        self.button1=QPushButton(self)
        self.button1.setText('开始生成原始数据')
        self.datetime_1 = QDateTimeEdit(QDate.currentDate(), self)              # 3
        self.datetime_1.dateTimeChanged.connect(lambda: print('DateTime Changed!'))
        self.datetime_1.setCalendarPopup(True)
        self.V_layout1=QVBoxLayout()
        self.V_layout2=QVBoxLayout()
        self.V_layout3=QVBoxLayout()
        self.H_layout1=QHBoxLayout()
        self.Layout__init()
        self.display_editline()
    def Layout__init(self):
        self.V_layout1.addWidget(self.label1)
        self.V_layout1.addWidget(self.label2)
        self.V_layout1.addWidget(self.label3)
        self.V_layout1.addWidget(self.label4)
        self.V_layout1.addWidget(self.label5)
        self.V_layout1.addWidget(self.label6)
        self.V_layout2.addWidget(self.editline1)
        self.V_layout2.addWidget(self.editline2)
        self.V_layout2.addWidget(self.editline3)
        self.V_layout2.addWidget(self.editline4)
        self.V_layout2.addWidget(self.editline5)
        self.V_layout2.addWidget(self.datetime_1)
        # self.V_layout2.addWidget(self.editline5)
        self.H_layout1.addLayout(self.V_layout1)
        self.H_layout1.addLayout(self.V_layout2)
        self.V_layout3.addLayout(self.H_layout1)
        self.V_layout3.addWidget(self.button1)
        # self.V_layout3.addWidget(self.datetime_1)
        self.setLayout(self.V_layout3)
    def display_editline(self):
        self.editline1.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline2.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline3.setPlaceholderText('需要严格按照Python路径输入！')
        self.editline4.setPlaceholderText('请输入期数')
        self.editline5.setPlaceholderText('按照"**mm-**mm"')
        self.button1.clicked.connect(self.output_excel)
        # self.editline5.setPlaceholderText('从日历选择日期')
    def output_excel(self):
        measure_line_path = "D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"#D:\Desktop\测试期次1\测试线路文件1.xlsx##self.editline1.text()
        dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'#D:\Desktop\测试期次1\测试数据库1.xlsx'#self.editline2.text()#self.editline2.text()
        path_output1 = 'D:\\Desktop\\测试期次1\\测试output\\'#self.editline3.text()#'D:\Desktop\测试期次1\测试output\self.editline3.text()
        qi_shu = self.editline4.text()
        BC_range1 = self.editline5.text()
        BC_range_min = re.findall('(.*)/', BC_range1)
        BC_range_max = re.findall('/(.*)', BC_range1)
        print(measure_line_path)
        print(dateset_path)
        print(path_output1)
        print(qi_shu)
        print(BC_range_min,BC_range_max)
        BC1=round(random.uniform(float(BC_range_min[0]), float(BC_range_max[0])), 4)
        RZ_names, RZ_values = read_measure_line_from_dataset(dateset_path, measure_line_path)
        for m in range(int(qi_shu)):#len(RZ_values)
            path_output2 = path_output1 + str(m + 1) + '期' + '原始数据' + '.xlsx'
            path_output3 = path_output1 + str(m + 1) + '期' + '原始数据' + '无支点' + '.xlsx'
            df = pd.read_excel(measure_line_path, 'Sheet1')  # 对应的导出为pd.to_excel(***.xlsx)
            df.columns = range(0, 3)
            name = RZ_names
            value = RZ_values[m]
            dict = {name[i]: value[i] for i in range(len(name))}
            for i in range(0, df.shape[0]):
                if (df.loc[i, 0] in name):
                    df.loc[i, 1] = dict[df.loc[i, 0]]
            df1 = df[:1]
            df2 = df[1:len(df)]
            df21 = df2.dropna(how='any')
            df3 = pd.concat([df1, df21], axis=0)
            df3.index = range(df3.shape[0])
            a = df3.iloc[:, [1]]
            b = range(df3.shape[0])
            num = []
            numx = []
            for i in range(df3.shape[0]):
                if (df3.loc[i, 0] in name):
                    numx.append(i)
                    pass
                else:
                    num.append(i)
            instrument_height = 1.4
            df4 = df3.loc[num, [0, 1, 2]]
            dict1 = {i: num[i] for i in range(df4.shape[0])}
            book1 = openpyxl.Workbook()
            sheet1 = book1.create_sheet('原始数据-无支点', 0)
            book2 = openpyxl.Workbook()
            sheet2 = book2.create_sheet('原始数据-支点', 0)
            sight_height_distance(df3, df4, book1, sheet1, book2, sheet2, BC1, path_output3, path_output2,num,dict1)



if __name__== '__main__':
    app=QApplication(sys.argv)
    demo=zhu_ti_cj_jiang()
    demo.show()
    sys.exit(app.exec_())