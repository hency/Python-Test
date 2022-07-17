#coding-utf8
import numpy
from openpyxl import load_workbook
import os
import openpyxl
import string

def read_measure_line_from_dataset(dateset_path,measure_line_path):
    mpath=load_workbook(measure_line_path)
    msheet=mpath.get_sheet_by_name('Sheet1')
    msheet=mpath['Sheet1']
    dpath=load_workbook(dateset_path)
    dsheet1=dpath.get_sheet_by_name('日报')
    sheetname_from_dateset=[]
    for name in dpath.sheetnames:
        if('沉降' in name):
            sheetname_from_dateset.append(name)
            print('沉降数据库有：%s',name)
        else:
            pass
    print('选择指定的数据库进行对应的沉降观测')
    #由于这里只有一个测线和一个建筑物沉降成果数据库，只对其进行原始数据的还原
    dsheet2=dpath.get_sheet_by_name(sheetname_from_dateset[0])
    dsheet3=dpath.get_sheet_by_name('日报')
    #先进行每一栋每一期的数据进行测试
    max_row=0
    for i in range(0,dsheet3.max_row):
        if(dsheet3.cell(i+1,2).value=='' or dsheet3.cell(i+1,3).value is None):
            max_row=dsheet3.max_row
            break
        elif(dsheet3.cell(dsheet3.max_row,2).value!='' and dsheet3.cell(dsheet3.max_row,3).value is not None):
            max_row=dsheet3.max_row
    date=[]
    max_col=0
    for i in range(2,max_row+1):
        date.append(dsheet3.cell(i, 2).value)
        print('第%d期沉降观测'%(i-1))
    for i in range(1,dsheet2.max_row):
        if(dsheet2.cell(i,2).value=='日期'):
            date_start=i
    for j in range(1,dsheet2.max_column):
        if(dsheet2.cell(date_start,j)=='' or dsheet2.cell(date_start,j) is None ):
            max_col=j-1
            break
        elif(dsheet2.cell(date_start,dsheet2.max_column)!='' or dsheet2.cell(date_start,dsheet2.max_column) is not None):
            max_col=dsheet2.max_column
    gc_name=[]
    gc_start_num=0
    for i in range(1,max_col):
        if (dsheet2.cell(date_start, i).value == '时间'):
            gc_start_num = i + 1
            for j in range(0,max_col-gc_start_num+1):
                gc_name.append(dsheet2.cell(date_start,gc_start_num+j).value)
            break
    cx_make=[]
    for i in range(1, len(date) + 1):
        date1 = date_start + i
        print('***********************************%s*******************************'%(dsheet2.cell(date1,2).value))
        print('提取侧线文件及数据库文件，进行相应复制')
        cx_make1=[]
        for j in range(0,max_col-gc_start_num+1):
            cx_make1.append(dsheet2.cell(date1,j+4).value)
        cx_make.append(cx_make1)
    return gc_name,cx_make
# dateset_path = 'D:\\Desktop\\测试期次1\\测试数据库1.xlsx'
# # measure_line_path="C:\\Users\\Kong\\Desktop\\项目\\测试期次1\\测试线路文件1.xlsx"
# measure_line_path = 'D:\\Desktop\\测试期次1\\测试线路文件1.xlsx'
# RZ_names, RZ_values = read_measure_line_from_dataset(dateset_path, measure_line_path)
def position_index(df3):
    position = []
    num_1=0
    for i in range(df3.shape[0]):
        if 'Y' in df3.iloc[i,0]:
            position.append([i,i-num_1-1])
            num_1 = num_1 + 1
    return position
def pingcha_height(cx_date1):
    pass
def time_jiange(cx_date1):
    pass
def pingcha_hou(cx_date1):
    pass
def pingcha_qian(cx_date1):
    pass
def check_BIAS():
    pass
def output_original_sight():
    pass
def output_original_height():
    pass
def output_summary1():
    pass
def output_summary2():
    pass
def create_original_file_dat():
    pass
def create_xlsx_RMS():
    pass
def create_house_date_dat_dir():
    pass
def sight_height_distance(df3,df4,book1,sheet1,book2,sheet2,BC1,path1,path2):
    sheet1.cell(1,1).value='点位'
    sheet1.cell(1,3).value='时间'
    sheet1.cell(1,5).value='视线高'
    sheet1.cell(1,7).value='视距'
    sheet1.cell(1,9).value='高程'
    sheet1.cell(2,1).value=df3.iloc[0,0]
    sheet1.cell(2,9).value=df4.iloc[0,1]
    sheet2.cell(1,1).value='点位'
    sheet2.cell(1,3).value='时间'
    sheet2.cell(1,5).value='视线高'
    sheet2.cell(1,7).value='视距'
    sheet2.cell(1,9).value='高程'
    sheet2.cell(2,1).value=df3.iloc[0,0]
    sheet2.cell(2,9).value=df4.iloc[0,1]
    k=3
    RH1_random = []
    RH2_random=[]
    FH1=[]
    FH2=[]
    height1=[]
    from ce1_6 import position_index
    from ce1_6 import position_index1
    from ce1_6 import Z_H_function
    position1 = position_index(df3)
    position2 = position_index1(df3)
    Z_H1 = Z_H_function(df3, df4, num)
    for i in range(df4.shape[0]-1):
        dh=df4.iloc[i+1,1]-df4.iloc[i,1]-BC1
        if(dh>1.13):
            print("高差大于1.13m，高差过大") #因为dh1_random的值不能超过1.15m
            xx1=1/0
        height1.append(df4.iloc[i+1,1]-(i+1)*BC1)
        sheet1.cell(7+5*i,9).value=height1[i]
        #'高差值最大是1.8-0.55=1.25m' 且 dh_random的取值范围为0.55+dh~1.8
        Sight_Height_random1=random.randint(-5, 15) * 0.00001  ##单位m
        Sight_Height_random2 = random.randint(-5, 15) * 0.00001  ##单位m0.15mm*2=0.3m
        dh1_random=dh+Sight_Height_random1  ##FH2-FH1=2*Sight_Height_random1+Sight_Height_random2
        dh2=2*dh-dh1_random
        flagxx1=0
        for q in range(len(position2)):
            if(i==position2[q][0]):
                #############加入
                pass
                flagxx1=1
                if (dh1_random > 0):
                    RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu=0
                    for p in range(position2[q][1]):
                        if ( ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006 ) and ( RH1_random_1+df3.loc[dict1[i],1]-i*BC1-1.75 < Z_H1[position2[q][2]-position2[q][1]+p] and Z_H1[position2[q][2]-position2[q][1]+p] < RH1_random_1+df3.loc[dict1[i],1]-i*BC1-0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu=can_shu+1
                    if(can_shu==position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6 + dh1_random, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, FH1[i])
                        sheet1.cell(k + 1, 5, RH1_random[i])
                        sheet1.cell(k + 2, 5, RH2_random[i])
                        sheet1.cell(k + 3, 5, FH2[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
                elif (dh1_random < 0):
                    #############加入
                    RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu = 0
                    for p in range(position2[q][1]):
                        if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):
                            # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu = can_shu + 1
                    if (can_shu == position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75 + dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
                elif (dh1_random == 0):
                    RH1_random_1 = round(random.uniform(0.6, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1 = RH1_random_1 - dh1_random
                    RH2_random_1 = RH1_random_1 + Sight_Height_random2
                    FH2_1 = RH2_random_1 - dh2
                    can_shu=0
                    for p in range(position2[q][1]):
                        if ( ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006 ) and ( RH1_random_1+df3.loc[dict1[i],1]-i*BC1-1.75 < Z_H1[position2[q][2]-position2[q][1]+p] and Z_H1[position2[q][2]-position2[q][1]+p] < RH1_random_1+df3.loc[dict1[i],1]-i*BC1-0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                            can_shu=can_shu+1
                    if(can_shu==position2[q][1]):
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        print("输出")
                    else:
                        for z1 in range(100000):
                            can_shu1=0
                            Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                            RH1_random_1 = round(random.uniform(0.6, 1.75),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                            FH1_1 = RH1_random_1 - dh1_random
                            RH2_random_1 = RH1_random_1 + Sight_Height_random2
                            FH2_1 = RH2_random_1 - dh2
                            print("重新选择")
                            for p in range(position2[q][1]):
                                if (((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006) and (RH1_random_1 + df3.loc[dict1[i], 1] - i * BC1 - 1.75 < Z_H1[position2[q][2] - position2[q][1]+p] and Z_H1[position2[q][2] - position2[q][1]+p] < RH1_random_1 +df3.loc[dict1[i], 1] - i * BC1 - 0.6)):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                                    can_shu1=can_shu1+1
                            if(can_shu1==position2[q][1]):
                                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                                FH1.append(FH1_1)
                                RH2_random.append(RH2_random_1)
                                FH2.append(FH2_1)
                                print("输出")
                                break
                            else:
                                continue
                    if ((i + 1) % 2 == 1):
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RB'
                        sheet1.cell(k + 1, 6).value = 'RF'
                        sheet1.cell(k + 2, 6).value = 'RF'
                        sheet1.cell(k + 3, 6).value = 'RB'
                    else:
                        sheet1.cell(k, 5, RH1_random[i])
                        sheet1.cell(k + 1, 5, FH1[i])
                        sheet1.cell(k + 2, 5, FH2[i])
                        sheet1.cell(k + 3, 5, RH2_random[i])
                        sheet1.cell(k, 6).value = 'RF'
                        sheet1.cell(k + 1, 6).value = 'RB'
                        sheet1.cell(k + 2, 6).value = 'RB'
                        sheet1.cell(k + 3, 6).value = 'RF'
                    k = k + 5
        if(flagxx1==1):
            continue
        if (dh1_random > 0):
            RH1_random_1=round(random.uniform(0.6 + dh1_random,1.75),5)###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 - dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2=random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6 + dh1_random, 1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, FH1[i])
                sheet1.cell(k + 1, 5, RH1_random[i])
                sheet1.cell(k + 2, 5, RH2_random[i])
                sheet1.cell(k + 3, 5, FH2[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k = k + 5
        elif(dh1_random<0):
            RH1_random_1=round(random.uniform(0.6,1.75+dh1_random),5)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 -dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6, 1.75 + dh1_random),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k=k+5
        elif(dh1_random==0):
            RH1_random_1=round(random.uniform(0.6,1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
            FH1_1=RH1_random_1 -dh1_random
            RH2_random_1=RH1_random_1 + Sight_Height_random2
            FH2_1=RH2_random_1 - dh2
            if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                print("输出")
                RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                FH1.append(FH1_1)
                RH2_random.append(RH2_random_1)
                FH2.append(FH2_1)
            else:
                for z1 in range(100000):
                    Sight_Height_random2 = random.randint(-5, 15) * 0.00001
                    RH1_random_1=round(random.uniform(0.6, 1.75),5) ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                    FH1_1=RH1_random_1 - dh1_random
                    RH2_random_1=RH1_random_1 + Sight_Height_random2
                    FH2_1=RH2_random_1 - dh2
                    print("重新选择")
                    if ((abs(RH1_random_1 - RH2_random_1) < 0.0003 and abs(FH1_1 - FH2_1) < 0.0003) and abs(dh1_random - dh2) < 0.0006):  # 进行判断基辅差与高差之差，由于仪器取位是5位，这里的随机数据选择五位
                        print("输出")
                        RH1_random.append(RH1_random_1)  ###为了保证RH2_random的取值范围在0.55~1.8之间这里把目标范围设置为0.6~1.75之间
                        FH1.append(FH1_1)
                        RH2_random.append(RH2_random_1)
                        FH2.append(FH2_1)
                        break
            if ((i + 1) % 2 == 1):
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RB'
                sheet1.cell(k + 1, 6).value = 'RF'
                sheet1.cell(k + 2, 6).value = 'RF'
                sheet1.cell(k + 3, 6).value = 'RB'
            else:
                sheet1.cell(k, 5, RH1_random[i])
                sheet1.cell(k + 1, 5, FH1[i])
                sheet1.cell(k + 2, 5, FH2[i])
                sheet1.cell(k + 3, 5, RH2_random[i])
                sheet1.cell(k, 6).value = 'RF'
                sheet1.cell(k + 1, 6).value = 'RB'
                sheet1.cell(k + 2, 6).value = 'RB'
                sheet1.cell(k + 3, 6).value = 'RF'
            k=k+5
    HD_differ_sum=0
    k1=3
    HDf1=[]
    HDf2=[]
    HDb1=[]
    HDb2=[]
    for i in range(df4.shape[0]-1):
        HD_random1 = round(random.uniform(-0.5, 0.5),3)##单位m
        HD_random2 = round(random.uniform(-0.5, 0.5),3)
        HD=df4.iloc[i+1,2]/2
        HDb1.append(HD+HD_random1)
        HDb2.append(HDb1[i]+round(random.uniform(-0.005,0.005),3))
        HDf1.append(HD+HD_random2)
        HDf2.append(HDf1[i]+round(random.uniform(-0.005,0.005),3))
        HDb=(HDb1[i]+HDb2[i])/2
        HDf=(HDf1[i]+HDf2[i])/2
        HD_differ=HDb-HDf
        HD_differ_sum=HD_differ_sum+HD_differ
        if(HD_differ<1.5 and HD_differ_sum<6):
            print('前后视距满足要求')
        else:
            print('前后视距不满足要求')
            print('出现异常将在GUI中提现')
            Exception1=1/0
        if((i+1)%2==1):
            sheet1.cell(k1,7,HDb1[i])
            sheet1.cell(k1+1,7).value=HDf1[i]
            sheet1.cell(k1+2,7,HDf2[i])
            sheet1.cell(k1+3,7).value=HDb2[i]
            sheet1.cell(k1,8).value='HDB'
            sheet1.cell(k1+1,8).value='HDF'
            sheet1.cell(k1+2,8).value='HDF'
            sheet1.cell(k1+3,8).value='HDB'
            sheet1.cell(k1,1).value=df4.iloc[0:df4.shape[0],0].iloc[i]
            sheet1.cell(k1+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
            sheet1.cell(k1+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
        else:
            sheet1.cell(k1,7,HDf1[i])
            sheet1.cell(k1+1,7).value=HDb1[i]
            sheet1.cell(k1+2,7,HDb2[i])
            sheet1.cell(k1+3,7).value=HDf2[i]
            sheet1.cell(k1,8).value='HDF'
            sheet1.cell(k1+1,8).value='HDB'
            sheet1.cell(k1+2,8).value='HDB'
            sheet1.cell(k1+3,8).value='HDF'
            sheet1.cell(k1,1).value=df4.iloc[0:df4.shape[0],0].iloc[i+1]
            sheet1.cell(k1+1,1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i]
            sheet1.cell(k1+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[i + 1]
        if(HD_differ_sum>6):
            print('****************************************************************************************************')
            print('前后视距累计差抄超限')
        k1=k1+5
    from ce1_6 import position_index
    position=position_index(df3)
    Z_name=[]
    Z_H=[]
    Z_HD=[]
    Z_HF=[]
    for zz1 in range(len(position)):
        HD_random3 = round(random.uniform(-0.5, 0.5), 3)  ##单位m
        Z_name.append(df3.loc[position[zz1][0],0])
        Z_HD.append(df3.loc[position[zz1][0],2]+HD_random3) ###
        Z_DH1=df3.loc[dict1[position[zz1][1]],1]-df3.loc[position[zz1][0],1]
        Z_H.append(df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-Z_DH1)
        Z_HB=RH1_random[position[zz1][1]]
        Z_HF.append(Z_HB+Z_DH1)
        if(Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-1.75 < Z_H[zz1] and Z_H[zz1] < Z_HB+df3.loc[dict1[position[zz1][1]],1]-position[zz1][1]*BC1-0.6):
            print("支点的高程设置合理！！！！！！！！！！！")
        else:
            print("支点的高程设置不合理****************")
    start_num=3
    start_num_fei=0
    continue_times=0
    nn=0
    kx1=0
    flag=0
    while(flag==0):
        for zz1 in range(kx1,len(position)):
            for zz2 in range(start_num_fei,position[zz1][1]+1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                if((zz2+1)%2==1):
                    sheet2.cell(start_num+5*(zz2-start_num_fei),5).value=RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value =FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                    sheet2.cell(start_num+  5 * (zz2 - start_num_fei) , 1).value = df4.iloc[0:df4.shape[0],0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    name1 = 'A' + str(start_num + 5 *(zz2 - start_num_fei)+4 )
                    sheet1[name1].font = Font(bold=True)
                else:
                    sheet2.cell(start_num+5*(zz2-start_num_fei),5).value=FH1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                    sheet2.cell(start_num+  5 * (zz2 - start_num_fei) , 1).value = df4.iloc[0:df4.shape[0],0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+1, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+2, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei)+3, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = df4.iloc[0:df4.shape[0], 0].iloc[zz2+1]
                    name1 = 'A' + str(start_num + 5 *(zz2 - start_num_fei)+4 )
                    sheet1[name1].font = Font(bold=True)
            kx2=0
            for zz3 in range(zz1,len(position)):
                if(position[zz1][1]==position[zz3][1]):
                    kx2=kx2+1
            zz3=zz1+kx2
            for zz4 in range(zz3-zz1):
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5+zz4, 1).value = df3.loc[position[zz1+zz4][0], 0] ##中间点的点名
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5+zz4, 5).value = Z_HF[zz1+zz4] ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 6).value = 'RZ'  ##中间点视线RZ
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 7).value = Z_HD[zz1+zz4]  ##中间点视距
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 8).value = 'HD'  ##中间点视线高
                sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 5 + zz4, 9).value = Z_H[zz1+zz4]  ##中间点视线高
            kx1=zz3-zz1+kx1
            start_num=start_num + 5 * (zz2 - start_num_fei) + 4 + zz3-zz1+1
            start_num_fei=position[zz1][1]+1
            if(kx1==len(position)):
                ##收尾
                flag=1
                if(position[zz1][1]==df4.shape[0]-1-1):
                    pass
                else:
                    for zz2 in range(start_num_fei, df4.shape[0]-1):
                        sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 9).value = height1[zz2]
                        if ((zz2 + 1) % 2 == 1):
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            sheet1[name1].font = Font(bold=True)
                        else:
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 5).value = FH1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 5).value = RH1_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 5).value = RH2_random[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 5).value = FH2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 8).value = 'HDB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 8).value = 'HDF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 6).value = 'RB'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 6).value = 'RF'
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 7).value = HDf1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 7).value = HDb1[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 7).value = HDb2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 7).value = HDf2[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei), 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 1, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 2, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 3, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            sheet2.cell(start_num + 5 * (zz2 - start_num_fei) + 4, 1).value = \
                            df4.iloc[0:df4.shape[0], 0].iloc[zz2 + 1]
                            name1 = 'A' + str(start_num + 5 * (zz2 - start_num_fei) + 4)
                            sheet1[name1].font = Font(bold=True)
                break
            else:
                pass
            break

    sheet1['A2'].font=Font(bold=True)
    for i in range(df4.shape[0]-1):
        name1='A'+str(7+5*i)
        name2='A'+str(7+5*i)
        sheet1[name2]=df4.iloc[i+1,0]
        sheet1[name1].font=Font(bold=True)
    book1.save(path1)
    book1.close()
    book2.save(path2)
    book2.close()
if __name__=='__main__':
    dateset_path='D:\\Desktop\\测试期次1\\测试数据库1.xlsx'
    measure_line_path="D:\\Desktop\\测试期次1\\测试线路文件1.xlsx"
    gc_name,cx_make,=read_measure_line_from_dataset(dateset_path,measure_line_path)
    pass
